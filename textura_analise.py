#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
textura_analise.py — fase 2: estatística e gráficos sobre Excel revisto
======================================================================

Não reextrai. Exige 0_Instrucoes e colunas de revisão da fase 1.

Uso:
    python textura_analise.py --xlsx resultado_near.xlsx
    python textura_analise.py --xlsx resultado_near.xlsx --desduplicacao nenhuma
    python textura_analise.py --xlsx resultado_near.xlsx --legendas legendas.json
"""

from __future__ import annotations

import argparse
import hashlib
import math
import re
import sys
from collections import Counter
from datetime import datetime
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill
from scipy import stats

import textura_legendas as tleg
import textura_lexico as tlex
import textura_near as tn
import textura_plots as tplot
import textura_triagem as ttri

try:
    import textura_stats as tst
except ImportError:
    tst = None

MODOS_DEDUPE = (
    "nenhuma", "candidatos", "contexto", "obra_termo",
    "ocorrencia", "ocorrencia_termo",
)

RELACOES_VALIDAS = sorted(
    set(tn.RELACOES_NUCLEARES) | set(ttri.RELACOES_NAO_NUCLEARES)
)
COLS_EDITAVEIS = [
    "relacao_sintactica", "nuclear", "polaridade", "eixo",
    "dominio", "motivo_exclusao",
]


def _cfg_consola():
    for s in (sys.stdout, sys.stderr):
        try:
            s.reconfigure(encoding="utf-8", errors="replace")
        except Exception:
            pass


def _meta_linha(unidade: str, n: int) -> pd.DataFrame:
    return pd.DataFrame({"unidade": [unidade], "N": [n]})


def _escrever_folha(
    xw,
    nome: str,
    df: pd.DataFrame,
    unidade: str,
    n: int,
    *,
    com_meta: bool = True,
):
    """Escreve folha estatística.

    ``com_meta=False`` para folhas de concordância reutilizáveis (sem banner
    ``unidade``/``N`` nas primeiras linhas — esse banner partia
    ``pd.read_excel`` em re-análises).
    """
    if not com_meta:
        if df is not None and len(df):
            df.to_excel(xw, sheet_name=nome, index=False)
        return
    meta = _meta_linha(unidade, n)
    meta.to_excel(xw, sheet_name=nome, index=False, startrow=0)
    if df is not None and len(df):
        df.to_excel(xw, sheet_name=nome, index=False, startrow=3)


def _parece_banner_meta(cols) -> bool:
    c = [str(x).strip().lower() for x in list(cols)]
    return len(c) >= 2 and c[0] == "unidade" and c[1] == "n"


def ler_folha_concordancia(xlsx, sheet: str = "8_Concordancia") -> pd.DataFrame:
    """Lê concordância fase 1 ou export de análise (com banner meta legado)."""
    conc = pd.read_excel(xlsx, sheet_name=sheet)
    if "relacao_sintactica" in conc.columns:
        return conc
    if _parece_banner_meta(conc.columns):
        # Export antigo de textura_analise: meta nas linhas 0–2, cabeçalho na 3
        conc2 = pd.read_excel(xlsx, sheet_name=sheet, header=3)
        if "relacao_sintactica" in conc2.columns:
            print(
                f"  [aviso] {sheet}: formato de saída de análise (banner meta); "
                "a ler cabeçalho na linha 4. Prefira o Excel *_near* / "
                "*_revisto* como entrada da fase 2.",
                flush=True,
            )
            return conc2
    return conc


def _sugerir_entrada_revisao(xlsx: Path) -> str:
    """Se o utilizador passou *_analise.xlsx, apontar o Excel de revisão."""
    nome = xlsx.name
    low = nome.lower()
    if "_analise" not in low:
        return ""
    stem = re.sub(r"_analise(\.xlsx)?$", "", nome, flags=re.I)
    if not stem.lower().endswith(".xlsx"):
        cand = xlsx.with_name(stem + ".xlsx")
    else:
        cand = xlsx.with_name(stem)
    if cand.exists():
        return str(cand)
    # Compósitaa_near_revisto_LR_v2_analise → …_v2.xlsx
    alt = xlsx.with_name(re.sub(r"_analise\.xlsx$", ".xlsx", nome, flags=re.I))
    if alt.exists() and alt != xlsx:
        return str(alt)
    return ""


def _nuclear_true_count(serie: pd.Series) -> int:
    return int(serie.map(
        lambda v: v is True or str(v).lower() in {"true", "1", "sim"}
    ).sum())


def sincronizar_hits_com_concordancia(xlsx: Path) -> dict:
    """Copia ``8_Concordancia`` → ``8_Concordancia_Hits`` (arquivo alinhado).

    A revisão edita só Concordancia; Hits é espelho e pode ficar velha.
    Devolve ``{ok, n_conc, n_hits_antes, n_hits_depois, mensagem}``.
    """
    xlsx = Path(xlsx)
    with pd.ExcelFile(xlsx) as xl:
        sheets = set(xl.sheet_names)
    if "8_Concordancia" not in sheets:
        return {"ok": False, "mensagem": "falta 8_Concordancia"}
    conc = ler_folha_concordancia(xlsx, "8_Concordancia")
    n_conc = _nuclear_true_count(conc["nuclear"]) if "nuclear" in conc.columns else 0
    n_antes = None
    if "8_Concordancia_Hits" in sheets:
        hits = pd.read_excel(xlsx, sheet_name="8_Concordancia_Hits")
        if "nuclear" in hits.columns:
            n_antes = _nuclear_true_count(hits["nuclear"])
    with pd.ExcelWriter(
        xlsx, engine="openpyxl", mode="a", if_sheet_exists="replace",
    ) as xw:
        conc.to_excel(xw, sheet_name="8_Concordancia_Hits", index=False)
    return {
        "ok": True,
        "n_conc": n_conc,
        "n_hits_antes": n_antes,
        "n_hits_depois": n_conc,
        "mensagem": (
            f"8_Concordancia_Hits ← 8_Concordancia "
            f"(nuclear True: {n_antes} → {n_conc})"
            if n_antes is not None else
            f"8_Concordancia_Hits criada (nuclear True={n_conc})"
        ),
    }


def cramers_v(tab: pd.DataFrame) -> float:
    if tab.size == 0 or min(tab.shape) < 2:
        return float("nan")
    chi2 = stats.chi2_contingency(tab.values)[0]
    n = tab.values.sum()
    r, k = tab.shape
    return float(np.sqrt(chi2 / (n * min(r - 1, k - 1)))) if n else float("nan")


def colinear_deterministica(df, a: str, b: str) -> bool:
    """True se V de Cramer ~= 1 (uma variavel funcao da outra)."""
    if a not in df.columns or b not in df.columns:
        return False
    tab = pd.crosstab(df[a], df[b])
    if tab.empty:
        return False
    # uma unica celula nao-nula por linha
    if (tab.gt(0).sum(axis=1) <= 1).all():
        return True
    v = cramers_v(tab)
    return v == v and v >= 0.999


def teste_contingencia(tab: pd.DataFrame, n_perm: int = 20000, semente=20260725,
                       min_validas: int = 200):
    """χ² ou Monte Carlo se alguma celula esperada < 5.

    Simulações degeneradas (expected zero) são saltadas; o denominador de p
    usa só o número efectivo de draws válidos. Se esse número ficar abaixo
    de ``min_validas``, cai-se para o teste exacto de Fisher (2×2) com aviso,
    em vez de reportar um p com resolução muito inferior ao ``n_perm`` nominal.
    """
    if tab.shape[0] < 2 or tab.shape[1] < 2:
        return {"metodo": "inaplicavel", "p": float("nan"),
                "estatistica": "", "cramer_v": float("nan"),
                "n_validas": 0}
    chi2, p_asym, gl, esperado = stats.chi2_contingency(tab.values)
    v = cramers_v(tab)
    if (esperado < 5).any():
        rng = np.random.default_rng(semente)
        obs = tab.values
        n = int(obs.sum())
        rprop = obs.sum(1) / n
        cprop = obs.sum(0) / n
        estat = chi2
        maiores = 0
        validas = 0
        max_tentativas = max(n_perm * 20, n_perm + 100)
        tentativas = 0
        while validas < n_perm and tentativas < max_tentativas:
            tentativas += 1
            flat = rng.multinomial(n, np.outer(rprop, cprop).ravel())
            sim = flat.reshape(obs.shape)
            try:
                c2 = stats.chi2_contingency(sim, correction=False)[0]
            except ValueError:
                continue
            validas += 1
            if c2 >= estat - 1e-12:
                maiores += 1

        # Poucas sims válidas → Fisher exacto em tabelas 2×2 (resolução honesta)
        if validas < min_validas and obs.shape == (2, 2):
            p_fish = float(stats.fisher_exact(obs).pvalue)
            print(
                f"AVISO: teste_contingencia — só {validas} sims MC válidas "
                f"(pedido {n_perm}); a usar Fisher exacto.",
                flush=True,
            )
            return {
                "metodo": (
                    f"Fisher exacto (fallback; {validas} MC validas/"
                    f"{tentativas} tentativas < {min_validas})"
                ),
                "p": p_fish,
                "estatistica": f"χ²({gl}) = {chi2:.3f}; Fisher",
                "cramer_v": round(v, 4),
                "n_validas": validas,
            }
        if validas == 0:
            return {"metodo": "Monte Carlo (inaplicavel: 0 sims validas)",
                    "p": float("nan"),
                    "estatistica": f"χ²({gl}) = {chi2:.3f}",
                    "cramer_v": round(v, 4),
                    "n_validas": 0}
        p = (maiores + 1) / (validas + 1)
        metodo = (
            f"Monte Carlo ({validas} perm. validas/{tentativas} tentativas; "
            f"celula esperada < 5)"
        )
        return {"metodo": metodo, "p": p,
                "estatistica": f"χ²({gl}) = {chi2:.3f}",
                "cramer_v": round(v, 4),
                "n_validas": validas}
    return {"metodo": "χ² assintotico", "p": p_asym,
            "estatistica": f"χ²({gl}) = {chi2:.3f}",
            "cramer_v": round(v, 4),
            "n_validas": None}


def logdice(o11: int, o12: int, n_janelas_no: int) -> float:
    """logDice com denominador = janelas do no (A5)."""
    dice = 2 * o11 / (n_janelas_no + o11 + o12) if (
        n_janelas_no + o11 + o12) else float("nan")
    return 14 + math.log2(dice) if dice and dice > 0 else float("nan")


def polaridade_nulo_banda(res_nuc: pd.DataFrame, hits_banda: dict | None,
                          campo: dict) -> float:
    """Proporcao esperada de estabilidade na banda (opcao 2, omissao)."""
    if not hits_banda:
        # fallback: proporcao de padroes E no lexico
        n_e = sum(1 for t in campo if t in tlex.POLO_ESTABILIDADE)
        return n_e / max(len(campo), 1)
    tot_e = sum(hits_banda.get(t, 0) for t in campo
                if t in tlex.POLO_ESTABILIDADE)
    tot = sum(hits_banda.get(t, 0) for t in campo)
    if tot == 0:
        n_e = sum(1 for t in campo if t in tlex.POLO_ESTABILIDADE)
        return n_e / max(len(campo), 1)
    return tot_e / tot


def validar_fase1(xlsx: Path, *, estrito: bool = False) -> dict:
    xlsx = Path(xlsx)
    with pd.ExcelFile(xlsx) as xl:
        sheets = list(xl.sheet_names)
        if "0_Instrucoes" not in sheets:
            raise SystemExit(
                "Ficheiro sem folha 0_Instrucoes — nao passou pela fase 1 "
                "(extraccao para revisao). Corra textura_near.py primeiro.")
        if "8_Concordancia" not in sheets:
            raise SystemExit("Falta a folha 8_Concordancia.")
        eh_saida_analise = (
            "1_Resumo" in sheets
            or "3_Testes" in sheets
            or "_analise" in xlsx.name.lower()
        )
        if eh_saida_analise:
            alt = _sugerir_entrada_revisao(xlsx)
            print(
                "  [aviso] Este Excel parece saída da fase 2 (*_analise / "
                "folhas 1_Resumo…). A entrada correcta é o Excel de revisão "
                "(*_near* / *_revisto*) com 8_Concordancia editável.",
                flush=True,
            )
            if alt:
                print(f"  [aviso] Use em vez disso:\n    {alt}", flush=True)
        conc = ler_folha_concordancia(xl, "8_Concordancia")
        meta = {}
        try:
            inst = pd.read_excel(xl, sheet_name="0_Instrucoes")
            if "chave" in inst.columns and "valor" in inst.columns:
                meta = dict(zip(inst["chave"].astype(str), inst["valor"]))
        except Exception:
            pass
    for c in ("relacao_sintactica", "nuclear", "canonical_term"):
        if c not in conc.columns:
            alt = _sugerir_entrada_revisao(xlsx)
            tip = (
                f"\nProvável causa: seleccionou um *_analise.xlsx (saída da "
                f"fase 2), não o Excel revisto.\n"
                f"Use: {alt or xlsx.with_name(xlsx.stem.replace('_analise', '') + '.xlsx')}"
            )
            raise SystemExit(f"Coluna obrigatoria em falta: {c}.{tip}")
    # validar taxonomia
    for i, row in conc.iterrows():
        rel = str(row.get("relacao_sintactica", ""))
        if rel and rel not in RELACOES_VALIDAS and rel != "nan":
            raise SystemExit(
                f"Valor invalido em relacao_sintactica linha {i+2}: {rel!r}. "
                f"Admissiveis: {RELACOES_VALIDAS}")

    # R95 — checklist de revisão (avisos; erros só com --estrito)
    checklist = None
    try:
        import textura_triagem as ttri
        checklist = ttri.checklist_revisao(conc)
        for a in checklist.get("avisos") or []:
            print(f"  [revisão] aviso: {a}", flush=True)
        for e in checklist.get("erros") or []:
            print(f"  [revisão] ERRO: {e}", flush=True)
        print(
            f"  [revisão] checklist score={checklist.get('score')} "
            f"nuclear={checklist.get('n_nuclear')}/{checklist.get('n')}",
            flush=True,
        )
        if estrito and not checklist.get("ok"):
            raise SystemExit(
                "Checklist de revisão falhou (--estrito). "
                "Corrija 8_Concordancia ou use textura_doctor.py."
            )
    except SystemExit:
        raise
    except Exception as exc:  # noqa: BLE001
        print(f"  [revisão] checklist indisponível: {exc}", flush=True)

    return {
        "conc": conc, "meta": meta, "sheets": sheets, "checklist": checklist,
    }


def comparar_revisao(conc: pd.DataFrame, meta: dict) -> dict:
    """Taxa de concordancia automatico vs humano (se houver snapshot)."""
    n = len(conc)
    alteradas = 0
    if "revisto_por_humano" in conc.columns:
        alteradas = int(conc["revisto_por_humano"].astype(str).str.strip()
                        .replace("", np.nan).notna().sum())
    return {
        "linhas": n,
        "marcadas_revisto_por_humano": alteradas,
        "taxa_marcacao": round(alteradas / n, 4) if n else 0,
    }


def _flag_candidato(serie: pd.Series) -> pd.Series:
    s = serie.astype(str).str.strip()
    return s.ne("") & ~s.str.lower().isin({"nan", "none", "nat", "false", "0"})


def aplicar_desduplicacao(nuc: pd.DataFrame, modo: str,
                          col_doc: str) -> tuple[pd.DataFrame, str]:
    """Desduplicação opcional na fase 2 (a revisão manual é a fonte de verdade)."""
    modo = (modo or "nenhuma").strip().lower()
    if modo not in MODOS_DEDUPE:
        raise ValueError(
            f"desduplicacao desconhecida: {modo!r}; "
            f"escolha entre {', '.join(MODOS_DEDUPE)}")
    if modo == "nenhuma" or len(nuc) == 0:
        return nuc.copy(), "nenhuma"
    if modo == "ocorrencia":
        if "texture_occurrence_id" not in nuc.columns:
            return nuc.copy(), "ocorrencia(col em falta→nenhuma)"
        return (nuc.drop_duplicates(subset=["texture_occurrence_id"])
                .copy(), "ocorrencia")
    if modo == "ocorrencia_termo":
        cols = [c for c in ("texture_occurrence_id", "canonical_term")
                if c in nuc.columns]
        if len(cols) < 2:
            return nuc.copy(), "ocorrencia_termo(cols em falta→nenhuma)"
        return nuc.drop_duplicates(subset=cols).copy(), "ocorrencia_termo"
    if modo == "obra_termo":
        cols = [c for c in (col_doc, "canonical_term") if c in nuc.columns]
        if len(cols) < 2:
            return nuc.copy(), "obra_termo(cols em falta→nenhuma)"
        return nuc.drop_duplicates(subset=cols).copy(), "obra_termo"
    if modo == "contexto":
        if "contexto" not in nuc.columns:
            return nuc.copy(), "contexto(indisponivel→nenhuma)"
        chave = nuc["contexto"].astype(str).str.strip()
        return (nuc.assign(_k=chave).drop_duplicates(subset=["_k"])
                .drop(columns=["_k"]).copy(), "contexto")
    # candidatos: entre linhas assinaladas na fase 1, fica 1 por contexto;
    # se não houver flag, equivale a desduplicar por contexto exacto.
    if "candidato_duplicado" in nuc.columns and "contexto" in nuc.columns:
        flagged = _flag_candidato(nuc["candidato_duplicado"])
        if flagged.any():
            keep = []
            visto: set[str] = set()
            for i, row in nuc.iterrows():
                if not bool(flagged.loc[i]):
                    keep.append(i)
                    continue
                k = str(row.get("contexto", "")).strip()
                if k in visto:
                    continue
                visto.add(k)
                keep.append(i)
            return nuc.loc[keep].copy(), "candidatos"
    return aplicar_desduplicacao(nuc, "contexto", col_doc)[0], "candidatos→contexto"


def analisar(xlsx: Path, saida: Path | None = None,
             nulo_polaridade: str = "banda",
             cooc_unidade: str = "obra",
             desduplicacao: str = "nenhuma",
             legendas: Path | None = None,
             relacoes: list[str] | None = None,
             estrito: bool = False) -> int:
    _cfg_consola()
    saida = saida or xlsx
    info = validar_fase1(xlsx, estrito=estrito)
    conc = info["conc"]
    meta = info["meta"]
    rev = comparar_revisao(conc, meta)
    consulta = str(meta.get("consulta", meta.get("query", "")) or "")
    leg = tleg.carregar(legendas, consulta=consulta)

    # conjuntos
    conc["nuclear"] = conc["nuclear"].map(
        lambda v: v is True or str(v).lower() in {"true", "1", "sim"})
    brutas = conc
    nuc = conc.loc[conc["nuclear"]].copy()
    if relacoes:
        rels = {r.strip().lower() for r in relacoes if r and r.strip()}
        if "relacao_sintactica" in nuc.columns and rels:
            antes = len(nuc)
            nuc = nuc.loc[
                nuc["relacao_sintactica"].astype(str).str.strip().str.lower()
                .isin(rels)
            ].copy()
            print(f"Filtro --relacao {sorted(rels)}: {antes} → {len(nuc)}",
                  flush=True)
    col_doc = "doc_id" if "doc_id" in nuc.columns else "caminho_ficheiro"
    dedup, modo_dedupe = aplicar_desduplicacao(nuc, desduplicacao, col_doc)

    n_bruto = len(brutas)
    n_nuc = len(nuc)
    n_dedup = len(dedup)
    n_janelas_no = int(float(meta.get("janelas_kwic_processadas", n_bruto) or n_bruto))
    tot_near = float(meta.get("tot_near", 0) or 0)
    tot_banda = float(meta.get("tot_banda", 0) or 0)

    # Sempre recontar a partir das nucleares *actuais* (após revisão).
    # Os n_* em 0_Instrucoes são da extracção fase 1 e ficam obsoletos
    # quando o utilizador muda nuclear TRUE↔FALSE.
    if "texture_occurrence_id" in dedup.columns:
        n_occ_nuc = int(dedup["texture_occurrence_id"].nunique())
    elif "texture_occurrence_id" in nuc.columns:
        n_occ_nuc = int(nuc["texture_occurrence_id"].nunique())
    else:
        n_occ_nuc = n_dedup
    n_occ_meta = meta.get("n_ocorrencias_nucleares", meta.get("n_ocorrencias"))
    try:
        n_occ_meta_i = int(float(n_occ_meta)) if n_occ_meta not in (None, "") else None
    except (TypeError, ValueError):
        n_occ_meta_i = None
    if n_occ_meta_i is not None and n_occ_meta_i != n_occ_nuc:
        avisos_pre = (
            f"meta fase 1 tinha n_ocorrencias_nucleares={n_occ_meta_i}; "
            f"recontado após revisão = {n_occ_nuc} (usado nos resultados)."
        )
    else:
        avisos_pre = ""
    print(f"Fase 2 | brutas={n_bruto} nucleares(hits)={n_nuc} "
          f"dedup({modo_dedupe})={n_dedup} "
          f"ocorrencias_nucleares={n_occ_nuc} janelas_no={n_janelas_no}",
          flush=True)
    if avisos_pre:
        print(f"  [aviso] {avisos_pre}", flush=True)
    if legendas:
        print(f"Legendas: {legendas} | {tleg.resumo_titulos(leg)}", flush=True)

    # --- A9 colinearidade -------------------------------------------------
    avisos = []
    if avisos_pre:
        avisos.append(avisos_pre)
    # Folha Hits desactualizada → sincronizar espelho no Excel de revisão
    if "8_Concordancia_Hits" in info.get("sheets", []):
        try:
            hits_df = pd.read_excel(xlsx, sheet_name="8_Concordancia_Hits")
            if "nuclear" in hits_df.columns:
                h_true = _nuclear_true_count(hits_df["nuclear"])
                if h_true != int(n_nuc):
                    sync = sincronizar_hits_com_concordancia(xlsx)
                    msg = sync.get("mensagem") or (
                        f"Hits sincronizada ({h_true} → {n_nuc})"
                    )
                    print(f"  [sync] {msg}", flush=True)
                    avisos.append(f"sincronizado: {msg}")
                    # folha Hits no workbook de entrada já está alinhada
                    info["sheets"] = list(pd.ExcelFile(xlsx).sheet_names)
        except Exception as exc:  # noqa: BLE001
            avisos.append(
                f"8_Concordancia_Hits desactualizada e sync falhou: {exc}"
            )
            print(f"  [aviso] sync Hits falhou: {exc}", flush=True)
    if colinear_deterministica(nuc, "canonical_term", "eixo"):
        avisos.append(
            "BLOQUEADO: eixo e funcao deterministica de canonical_term "
            "(V de Cramer = 1). Testes que cruzam eixo foram omitidos. "
            "Proposta de revisao do eixo: varied -> invariancia_diacronica "
            "(ou ambos); uniform -> homogeneidade_sincronica; "
            "static -> invariancia_diacronica.")
    if colinear_deterministica(nuc, "canonical_term", "polaridade"):
        avisos.append(
            "polaridade é função de canonical_term no léxico adjudicado "
            "(cada termo tem uma só polaridade). Esperado em classes "
            "mono-polares; não é falha da análise.")

    # --- 2_Frequencias (A2) ----------------------------------------------
    freq_tok = (nuc.groupby("canonical_term")
                .agg(ocorrencias_token=("canonical_term", "size"),
                     obras_com_ocorrencia=(col_doc, "nunique"))
                .sort_values("ocorrencias_token", ascending=False)
                .reset_index())
    # H, J, 1/D sobre ocorrencias_token
    cont_tok = nuc["canonical_term"].value_counts()
    S = int((cont_tok > 0).sum())
    H = tn.shannon(cont_tok.values) if S else float("nan")
    J = tn.pielou(cont_tok.values) if S else float("nan")
    invD = tn.simpson_inverso(cont_tok.values) if S else float("nan")

    # --- 3_Testes (A4, A8, A9) -------------------------------------------
    testes = []
    pol = dedup.replace({"polaridade": {"": np.nan}}).dropna(subset=["polaridade"])
    if len(pol):
        n_est = int((pol["polaridade"] == "estabilidade").sum())
        n_tot = len(pol)
        # nulo
        hits_banda = {}
        if "hits_banda" in meta:
            try:
                hits_banda = eval(meta["hits_banda"], {"__builtins__": {}})
            except Exception:
                hits_banda = {}
        campo = {}
        if "campo_tipos" in meta:
            campo = {t: [] for t in str(meta["campo_tipos"]).split(",") if t}
        if nulo_polaridade == "lexico":
            p0 = sum(1 for t in campo if t in tlex.POLO_ESTABILIDADE) / max(
                len(campo), 1)
            rotulo_nulo = f"proporcional ao lexico (p0={p0:.3f})"
        else:
            p0 = polaridade_nulo_banda(nuc, hits_banda, campo or {
                t: [] for t in nuc["canonical_term"].unique()})
            rotulo_nulo = f"calibrado pela banda de referencia (p0={p0:.3f})"
        bt = stats.binomtest(n_est, n_tot, p0, alternative="two-sided")
        if tst:
            ic = tn.bootstrap_proporcao(
                (pol["polaridade"] == "estabilidade").values)
        else:
            ic = (float("nan"), float("nan"))
        testes.append({
            "familia": "polaridade",
            "teste": f"Binomial estabilidade vs nulo ({rotulo_nulo})",
            "estatistica": f"{n_est}/{n_tot} = {n_est/n_tot:.3f} "
                           f"[IC95% {ic[0]:.3f}-{ic[1]:.3f}]",
            "dimensao_efeito": f"prop={n_est/n_tot:.3f}; nulo={p0:.3f}",
            "p": bt.pvalue,
            "metodo": "binomial exacto",
        })
        # Bayes -> 3_Testes (A3)
        if tst:
            bf = tst.bayes_factor_proporcao(n_est, n_tot)
            testes.append({
                "familia": "polaridade",
                "teste": "Factor de Bayes (proporcao vs 0,5; prior Beta(1,1))",
                "estatistica": f"BF10={bf}",
                "dimensao_efeito": f"BF={bf}",
                "p": float("nan"),
                "metodo": "Bayes factor proporcao",
            })

    col_rel = "relacao_sintactica"
    if len(pol) and col_rel in pol.columns:
        if not colinear_deterministica(pol, col_rel, "polaridade"):
            tab = pd.crosstab(pol[col_rel], pol["polaridade"])
            r = teste_contingencia(tab)
            testes.append({
                "familia": "estrutura_sintactica",
                "teste": "relacao_sintactica × polaridade",
                "estatistica": r["estatistica"],
                "dimensao_efeito": f"V_Cramer={r['cramer_v']}",
                "p": r["p"],
                "metodo": r["metodo"],
            })
        else:
            avisos.append(
                "teste relação×polaridade omitido: tabela degenerada "
                "(só uma polaridade nas nucleares → colinearidade). "
                "Não é falha; o teste χ² não se aplica.")

    if ("eixo" in dedup.columns and col_rel in dedup.columns
            and not any("eixo e funcao" in a for a in avisos)):
        if colinear_deterministica(dedup, "canonical_term", "eixo"):
            avisos.append("Teste relacao×eixo bloqueado (eixo=f(canonical_term)).")
        else:
            tab = pd.crosstab(dedup[col_rel], dedup["eixo"])
            r = teste_contingencia(tab)
            testes.append({
                "familia": "estrutura_sintactica",
                "teste": "relacao_sintactica × eixo",
                "estatistica": r["estatistica"],
                "dimensao_efeito": f"V_Cramer={r['cramer_v']}",
                "p": r["p"],
                "metodo": r["metodo"],
            })

    tst_df = pd.DataFrame(testes)
    if len(tst_df):
        # BH por familia (A14)
        tst_df["p_ajustado_BH"] = np.nan
        for fam, idx in tst_df.groupby("familia").groups.items():
            sub = tst_df.loc[idx, "p"].astype(float)
            mask = sub.notna()
            if mask.any():
                adj = tn.benjamini_hochberg(sub[mask].values)
                tst_df.loc[sub[mask].index, "p_ajustado_BH"] = adj

    # --- 4_Sintaxe sobre dedup -------------------------------------------
    sintaxe = (pd.crosstab(dedup["canonical_term"], dedup[col_rel])
               if len(dedup) else pd.DataFrame())

    # --- 5_Coocorrencia por obra (A12) -----------------------------------
    cooc = pd.DataFrame()
    if cooc_unidade == "obra" and len(nuc):
        tipos = sorted(nuc["canonical_term"].unique())
        mat = pd.DataFrame(0, index=tipos, columns=tipos, dtype=int)
        for _, g in nuc.groupby(col_doc):
            presentes = sorted(g["canonical_term"].unique())
            for a in presentes:
                for b in presentes:
                    mat.at[a, b] += 1
        # so escrever se houver off-diagonal
        if (mat.values.sum() - np.trace(mat.values)) > 0:
            cooc = mat
        else:
            avisos.append(
                "5_Coocorrencia omitida: matriz diagonal "
                f"(unidade={cooc_unidade}; nenhum par de tipos na mesma obra).")

    # --- 9_Associacao (A5, A6) -------------------------------------------
    # O11 sobre nucleares; O12 da meta se existir
    hits_banda = {}
    if "hits_banda" in meta:
        try:
            hits_banda = eval(meta["hits_banda"], {"__builtins__": {}})
        except Exception:
            hits_banda = {}
    filas = []
    for etq, o11 in Counter(nuc["canonical_term"]).items():
        o12 = int(hits_banda.get(etq, 0))
        r1 = tot_near or max(n_janelas_no * 3, 1)  # fallback
        r2 = tot_banda or r1
        m = tn.medidas_associacao(o11, o12, r1, r2, n_janelas_no)
        if not m:
            continue
        # corrigir logDice (A5)
        m["logDice"] = round(logdice(o11, o12, n_janelas_no), 3)
        m["canonical_term"] = etq
        m["obras"] = int(nuc.loc[nuc["canonical_term"] == etq, col_doc].nunique())
        # leitura (A6)
        orr = m.get("razao_possib", float("nan"))
        dp = m.get("DeltaP", float("nan"))
        m["leitura"] = (
            f"OR={orr}: o termo e {orr:.2f}x mais provavel na janela NEAR "
            f"do que na banda; ΔP={dp:.5f} (diferenca de probabilidades)."
        )
        m["r1_tokens_near"] = r1
        m["r2_tokens_banda"] = r2
        m["n_janelas_no"] = n_janelas_no
        m["formula_logDice"] = "14+log2(2*O11/(n_janelas_no+O11+O12))"
        # dispersao so se obras >= 5
        if m["obras"] < 5:
            m["DP_Gries"] = float("nan")
            m["D_Juilland"] = float("nan")
            m["partes_nao_nulas"] = m["obras"]
        else:
            partes = Counter(nuc.loc[nuc["canonical_term"] == etq, col_doc])
            tam = Counter(nuc[col_doc])
            m["DP_Gries"] = tn.dispersao_gries_dp(partes, tam)
            m["D_Juilland"] = tn.juilland_d(partes, max(len(tam), 1))
            m["partes_nao_nulas"] = int((np.array(list(partes.values())) > 0).sum())
        filas.append(m)
    assoc = pd.DataFrame(filas)
    if len(assoc):
        # ordenar por dimensao de efeito (OR), nao por G2
        assoc = assoc.sort_values("razao_possib", ascending=False)
        if "MI3" in assoc.columns:
            assoc = assoc.rename(columns={"MI3": "MI3_heuristica"})

    # --- 15_Perfis sobre dedup (A1, A11) ---------------------------------
    perfil = pd.DataFrame()
    if tst and len(dedup) >= 3:
        d = dedup.copy()
        # Excel fase 1 pode já ter termo_tipo e canonical_term — evitar
        # colunas duplicadas no rename (value_counts falha com DataFrame).
        if "canonical_term" in d.columns:
            d["termo_tipo"] = d["canonical_term"]
        elif "termo_tipo" not in d.columns:
            d["termo_tipo"] = d.get("matched_form", pd.Series(dtype=str))
        try:
            perfil, _ = tst.perfis_e_dendrograma(d, None, min_ocorr=1)
        except Exception as exc:
            avisos.append(f"15_Perfis omitido: {exc}")
            perfil = pd.DataFrame()

    # --- 13_Regressao: prever nuclear (A10) ------------------------------
    reg_tab, reg_aj = pd.DataFrame(), {}
    if tst and len(brutas) > 20:
        dreg = brutas.copy()
        dreg["polaridade"] = np.where(dreg["nuclear"], "estabilidade", "variabilidade")
        # abuso do alvo: usamos polaridade como proxy binario nuclear
        # melhor: criar coluna alvo_nuclear
        dreg = dreg.rename(columns={"polaridade": "_pol_old"})
        dreg["alvo_nuclear"] = np.where(dreg["nuclear"], "nuclear", "incidental")
        # so preditores nao circulares
        preds = [c for c in ("distancia", "lado", "censurado_esq", "censurado_dir")
                 if c in dreg.columns]
        if preds:
            # adaptar regressao para alvo custom
            tab, aj = tst.regressao_logistica(
                dreg.rename(columns={"alvo_nuclear": "polaridade"}),
                alvo="polaridade", positivo="nuclear",
                preditores=tuple(preds))
            # bloquear coeficientes absurdos (separacao)
            if len(tab) and (
                    (tab["coef_log_odds"].abs() > 10).any()
                    or (tab["erro_padrao"] > 100).any()):
                avisos.append(
                    "13_Regressao abortada: separacao completa "
                    "(|coef|>10 ou EP>100).")
                reg_tab, reg_aj = pd.DataFrame(), {"aviso": "separacao"}
            else:
                reg_tab, reg_aj = tab, aj

    # --- A3 formas por tipo (sem indices de riqueza) ---------------------
    formas = (nuc.groupby(["canonical_term", "matched_form"])
              .agg(n=("matched_form", "size"),
                   obras=(col_doc, "nunique"))
              .reset_index())

    # --- Graficos (B) ----------------------------------------------------
    base = saida.parent
    g_msgs = []
    leg_formas = leg.get("formas") or {}
    leg_nuvem = leg.get("nuvem") or {}
    leg_sankey = leg.get("sankey") or {}
    rodape = str(leg.get("rodape") or "")
    # Gráficos usam *todas* as nucleares (N_hits), não o conjunto deduplicado
    # (N_dedup). Os testes χ²/BF usam dedup — daí Resumo ter 1171 e 1156.
    nota_n = (
        f"N_hits={n_nuc}"
        + (f" · N_dedup({modo_dedupe})={n_dedup}" if modo_dedupe != "nenhuma"
           else "")
    )
    g1 = base / "_g_freq_token.png"
    try:
        tit_f = str(leg_formas.get("titulo")
                    or "Frequência por termo canónico")
        if f"N={n_nuc}" not in tit_f and "N_hits=" not in tit_f:
            tit_f = f"{tit_f}  ({nota_n})"
        sub_f = str(leg_formas.get("subtitulo") or "")
        if "N_hits" not in sub_f and "N=" not in sub_f:
            sub_f = (f"{sub_f} · {nota_n}".strip(" ·")
                     if sub_f else nota_n)
        tplot.barras_horizontais(
            list(freq_tok["canonical_term"]),
            list(freq_tok["ocorrencias_token"]),
            g1,
            titulo=tit_f,
            subtitulo=sub_f,
            xlabel=str(leg_formas.get("xlabel") or "Ocorrencias (N_hits)"),
            rodape=rodape,
            max_n=None,
        )
        g_msgs.append(f"OK freq token -> {g1.name}")
    except Exception as exc:
        g_msgs.append(f"FALHA freq: {exc}")

    g_nuvem = base / "_g_nuvem.png"
    try:
        # variante ponderada por obras (matched_form), ainda sobre nucleares
        freq_ob = (nuc.groupby("matched_form")[col_doc].nunique().to_dict())
        tit_n = str(leg_nuvem.get("titulo") or "Nuvem de palavras")
        sub_n = str(leg_nuvem.get("subtitulo")
                    or "Tamanho ∝ nº de obras por forma de superfície")
        if "N_hits" not in sub_n and "N=" not in sub_n:
            sub_n = f"{sub_n} · {nota_n}".strip(" ·")
        tplot.nuvem_palavras(
            freq_ob, g_nuvem,
            titulo=tit_n,
            subtitulo=sub_n,
            rodape=rodape)
        g_msgs.append(f"OK nuvem -> {g_nuvem.name}")
    except Exception as exc:
        g_msgs.append(f"FALHA nuvem: {exc}")

    g_sankey1 = base / "_g_sankey_forma_obra.html"
    g_sankey2 = base / "_g_sankey_termo_rel.html"
    try:
        p1 = tplot.pares_forma_obra(nuc)
        tit_s1 = str(leg_sankey.get("titulo")
                     or "Sankey: forma → documento")
        if "N=" not in tit_s1 and "N_hits=" not in tit_s1:
            tit_s1 = f"{tit_s1}  ({nota_n})"
        tplot.sankey_html(p1, g_sankey1, titulo=tit_s1)
        p2 = tplot.pares_termo_rel_pol(nuc)
        tplot.sankey_html(
            p2, g_sankey2,
            titulo=f"termo → relação → polaridade  ({nota_n})",
        )
        g_msgs.append("OK sankey HTML")
    except Exception as exc:
        g_msgs.append(f"FALHA sankey: {exc}")

    # --- Resumo ----------------------------------------------------------
    resumo = pd.DataFrame({
        "indicador": [
            "Fase", "Data fase 2", "Comando fase 1 (meta)",
            "Schema near (meta)",
            "Linhas brutas (hits)", "Linhas nucleares (N_hits)",
            f"Linhas apos desduplicacao ({modo_dedupe})",
            "Ocorrencias nucleares (N_ocorrencias)",
            "Janelas KWIC do no (denominador logDice)",
            "Ficheiros", "Obras (doc_id)",
            "Tipos (S)", "H (ocorrencias_token)", "J", "1/D",
            "Revisao: linhas marcadas", "Revisao: taxa marcacao",
            "Avisos", "Nulo polaridade", "Unidade co-ocorrencia",
            "Modo desduplicacao", "Analise sem revisao humana",
        ],
        "valor": [
            "2 - analise",
            datetime.now().isoformat(timespec="seconds"),
            meta.get("comando", "—"),
            meta.get("schema_near", "—"),
            n_bruto, n_nuc, n_dedup, n_occ_nuc, n_janelas_no,
            brutas["caminho_ficheiro"].nunique() if "caminho_ficheiro" in brutas else "—",
            brutas[col_doc].nunique() if col_doc in brutas.columns else "—",
            S, round(H, 4) if H == H else "—",
            round(J, 4) if J == J else "—",
            round(invD, 4) if invD == invD else "—",
            rev["marcadas_revisto_por_humano"], rev["taxa_marcacao"],
            " | ".join(avisos) if avisos else "—",
            nulo_polaridade, cooc_unidade, modo_dedupe,
            meta.get("sem_revisao", "nao"),
        ],
    })

    # --- Export ----------------------------------------------------------
    print(f"A escrever {saida.name} ...", flush=True)
    with pd.ExcelWriter(saida, engine="openpyxl") as xw:
        # preservar 0_Instrucoes e concordancia
        try:
            pd.read_excel(xlsx, sheet_name="0_Instrucoes").to_excel(
                xw, sheet_name="0_Instrucoes", index=False)
        except Exception:
            pass
        resumo.to_excel(xw, sheet_name="1_Resumo", index=False)
        _escrever_folha(xw, "2_Frequencias", freq_tok,
                        "linhas nucleares (ocorrencias_token) / doc_id",
                        n_nuc)
        if len(tst_df):
            out_t = tst_df.copy()
            for c in ("p", "p_ajustado_BH"):
                if c in out_t.columns:
                    out_t[c] = out_t[c].map(
                        lambda v: f"{v:.5g}" if v == v else "")
            _escrever_folha(xw, "3_Testes", out_t,
                            f"nucleares apos desduplicacao ({modo_dedupe})",
                            n_dedup)
        if len(sintaxe):
            _escrever_folha(xw, "4_Sintaxe", sintaxe.reset_index(),
                            f"nucleares apos desduplicacao ({modo_dedupe})",
                            n_dedup)
        if len(cooc):
            _escrever_folha(xw, "5_Coocorrencia", cooc.reset_index(),
                            f"obras (co-presenca; unidade={cooc_unidade})",
                            int(nuc[col_doc].nunique()))
        # Concordância sem banner meta — reutilizável / legível por pandas
        _escrever_folha(
            xw, "8_Concordancia", brutas,
            "todas as linhas (brutas=hits)", n_bruto, com_meta=False,
        )
        # Preservar / regenerar folhas de identidade (schema ≥ 2)
        try:
            hits_alias = brutas
            _escrever_folha(
                xw, "8_Concordancia_Hits", hits_alias,
                "hits NEAR (= 8_Concordancia)", n_bruto, com_meta=False,
            )
        except Exception:
            pass
        try:
            if "8_Concordancia_Ocorrencias" in info["sheets"]:
                pd.read_excel(
                    xlsx, sheet_name="8_Concordancia_Ocorrencias"
                ).to_excel(
                    xw, sheet_name="8_Concordancia_Ocorrencias", index=False)
            elif "texture_occurrence_id" in brutas.columns:
                tn.agregar_ocorrencias(brutas).to_excel(
                    xw, sheet_name="8_Concordancia_Ocorrencias", index=False)
        except Exception:
            pass
        for folha_extra in ("Config_lexico", "Manifesto_corpus", "Duplicados",
                            "9_Excluidas"):
            if folha_extra in info["sheets"]:
                try:
                    if folha_extra == "9_Excluidas":
                        excl = ler_folha_concordancia(xlsx, "9_Excluidas")
                        excl.to_excel(
                            xw, sheet_name="9_Excluidas", index=False)
                    else:
                        pd.read_excel(xlsx, sheet_name=folha_extra).to_excel(
                            xw, sheet_name=folha_extra, index=False)
                except Exception:
                    pass
        if len(assoc):
            cols_a = [c for c in (
                "canonical_term", "O11_janela", "O12_banda_ref", "obras",
                "razao_possib", "IC95_inf", "IC95_sup", "DeltaP",
                "logDice", "log_likelihood_G2", "MI", "MI3_heuristica",
                "p_fisher", "leitura", "r1_tokens_near", "r2_tokens_banda",
                "n_janelas_no", "formula_logDice", "DP_Gries", "D_Juilland",
                "partes_nao_nulas",
            ) if c in assoc.columns]
            _escrever_folha(xw, "9_Associacao", assoc[cols_a],
                            "linhas nucleares (O11); logDice usa janelas do no",
                            n_nuc)
        _escrever_folha(xw, "12_Formas", formas,
                        "linhas nucleares; inventario fechado pela consulta",
                        n_nuc)
        if len(reg_tab):
            _escrever_folha(xw, "13_Regressao", reg_tab,
                            "linhas brutas; DV=nuclear vs incidental",
                            n_bruto)
            if reg_aj:
                pd.DataFrame({"indicador": list(reg_aj),
                              "valor": list(reg_aj.values())}).to_excel(
                    xw, sheet_name="13_Ajuste_modelo", index=False)
        if len(perfil):
            _escrever_folha(xw, "15_Perfis", perfil,
                            f"nucleares apos desduplicacao ({modo_dedupe})",
                            n_dedup)
        pd.DataFrame({"aviso": avisos or ["—"]}).to_excel(
            xw, sheet_name="0_Avisos", index=False)
        # Dominios se existir
        if "Dominios_por_rever" in info["sheets"]:
            pd.read_excel(xlsx, sheet_name="Dominios_por_rever").to_excel(
                xw, sheet_name="Dominios_por_rever", index=False)

    # embutir graficos
    wb = load_workbook(saida)
    if "6_Graficos" in wb.sheetnames:
        del wb["6_Graficos"]
    ws = wb.create_sheet("6_Graficos")
    ws["A1"] = f"Graficos | N_nuclear={n_nuc} | {datetime.now().date()}"
    ws["A2"] = " | ".join(g_msgs)
    row = 4
    for p, label in ((g1, "Frequencia token"), (g_nuvem, "Nuvem")):
        ws.cell(row=row, column=1, value=label)
        if p.exists():
            try:
                ws.add_image(XLImage(str(p)), f"A{row + 1}")
            except Exception as exc:
                ws.cell(row=row + 1, column=1, value=f"FALHA embutir: {exc}")
        else:
            ws.cell(row=row + 1, column=1, value="figura nao gerada")
        row += 35
    ws.cell(row=row, column=1,
            value=f"Sankey HTML: {g_sankey1.name} ; {g_sankey2.name}")
    for nome in wb.sheetnames:
        s = wb[nome]
        for c in s[1]:
            if c.value is not None:
                c.font = Font(name="Arial", bold=True)
        s.freeze_panes = "A2"
    wb.save(saida)

    print(resumo.to_string(index=False))
    print()
    print("=== Análise concluída com sucesso ===")
    print(f"Ficheiro de saída (fase Analisar): {saida}")
    print(
        "Próximo passo na GUI: «4. Apêndice DOCX» — pode usar este "
        "*_analise.xlsx ou o Excel revisto (*_v2.xlsx)."
    )
    if avisos:
        print()
        print(
            "Notas estatísticas (salvaguardas — a análise NÃO falhou):"
        )
        for a in avisos:
            txt = a.strip()
            if txt.lower().startswith("aviso:"):
                txt = txt.split(":", 1)[1].strip()
            print(f"  · {txt}")
    return 0


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--xlsx", type=Path, required=True)
    ap.add_argument("--saida", type=Path, default=None)
    ap.add_argument("--nulo-polaridade", choices=["banda", "lexico"],
                    default="banda")
    ap.add_argument("--cooc-unidade", choices=["obra", "frase"],
                    default="obra")
    ap.add_argument("--desduplicacao", choices=list(MODOS_DEDUPE),
                    default="nenhuma",
                    help="nenhuma=usar nucleares tal como revistas; "
                         "contexto=1 linha por snippet; "
                         "candidatos=respeita candidato_duplicado; "
                         "obra_termo=sensibilidade doc×termo; "
                         "ocorrencia=1 linha por texture_occurrence_id; "
                         "ocorrencia_termo=1 linha por ocorrência×termo")
    ap.add_argument("--legendas", type=Path, default=None,
                    help="JSON com titulos/subtitulos editaveis dos graficos")
    ap.add_argument("--relacao", default="",
                    help="subset de relacoes (virgulas); vazio=todas nucleares")
    ap.add_argument("--plano-a-priori", type=Path, default=None,
                    help="reservado (ainda nao aplicado nesta build)")
    ap.add_argument("--kappa-cego", type=Path, default=None,
                    help="reservado (ainda nao aplicado nesta build)")
    ap.add_argument(
        "--estrito",
        action="store_true",
        help="falhar se o checklist de revisão tiver erros "
             "(nuclear≠relação, 0 nucleares, …)",
    )
    args = ap.parse_args()
    if args.plano_a_priori:
        print(f"AVISO: --plano-a-priori ignorado nesta build: "
              f"{args.plano_a_priori}", flush=True)
    if args.kappa_cego:
        print(f"AVISO: --kappa-cego ignorado nesta build: "
              f"{args.kappa_cego}", flush=True)
    rels = [p.strip() for p in (args.relacao or "").split(",") if p.strip()]
    return analisar(
        args.xlsx, args.saida,
        nulo_polaridade=args.nulo_polaridade,
        cooc_unidade=args.cooc_unidade,
        desduplicacao=args.desduplicacao,
        legendas=args.legendas,
        relacoes=rels or None,
        estrito=args.estrito,
    )


if __name__ == "__main__":
    raise SystemExit(main())
