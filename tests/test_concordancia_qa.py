# -*- coding: utf-8 -*-
"""QA pass: normalisation, relations, domain lexicon, duplicate groups."""

from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook

import textura_concordancia_qa as qa


class TestNorm(unittest.TestCase):
    def test_ligature_and_case(self):
        self.assertEqual(qa.norm("Texture ﬁeld"), "texture field")

    def test_fused_hyphen(self):
        self.assertEqual(qa.fused("ever-changing"), qa.fused("everchanging"))


class TestRelations(unittest.TestCase):
    def test_direct_of_texture(self):
        cat, ev = qa.classify_relation(
            "the complexity of the orchestral texture is striking",
            "complexity",
        )
        self.assertEqual(cat, "direct")
        self.assertIn("complexity", ev)

    def test_coord_heterog(self):
        cat, _ = qa.classify_relation(
            "a combination of texture and melody in the coda",
            "combination",
        )
        self.assertEqual(cat, "coord_heterog")

    def test_into_texture(self):
        cat, _ = qa.classify_relation(
            "lines are blended into a dense texture",
            "blended",
        )
        self.assertEqual(cat, "into_texture")

    def test_absent_no_form(self):
        cat, motive = qa.classify_relation(
            "a dense texture of strings",
            "uniformity",
        )
        self.assertEqual(cat, "absent")
        self.assertIn("matched form", motive)

    def test_unresolved_has_texture(self):
        # Both tokens present, but no licensed syntactic bridge → review.
        cat, ev = qa.classify_relation(
            "homogeneity is discussed; later chapters treat texture alone",
            "homogeneity",
        )
        self.assertEqual(cat, "unresolved")
        self.assertEqual(ev, "")


class TestDomain(unittest.TestCase):
    def test_geology(self):
        dom, cue = qa.classify_domain(
            "dolomite facies with bioclast texture"
        )
        self.assertEqual(dom, "geologia")
        self.assertEqual(cue, "dolomite")

    def test_musical_no_hit(self):
        dom, cue = qa.classify_domain(
            "a continuous polyphonic texture of strings"
        )
        self.assertIsNone(dom)
        self.assertIsNone(cue)


class TestDuplicates(unittest.TestCase):
    def test_near_duplicate_group(self):
        ctx_a = (
            "the music gained a continuous non-strophic texture it had not "
            "had either in the frottola or in the school that finished"
        )
        ctx_b = (
            "the music gained a continuous non-strophic texture it had not "
            "had either in the frottola or in the school that ﬁnished"
        )
        df = pd.DataFrame({
            "contexto": [ctx_a, ctx_b, "totally unrelated short window here"],
            "doc_id": ["a", "b", "c"],
            "grupo_passagem_id": [pd.NA, pd.NA, pd.NA],
            "caminho_ficheiro": [
                r"E:\todos os textos\(1933)_A.pdf",
                r"E:\outros\(1933)_A__copy.pdf",
                r"E:\x.pdf",
            ],
            "nuclear": [True, True, True],
            "matched_form": ["continuous", "continuous", "short"],
        })
        new, tracked = qa.find_duplicate_groups(df)
        self.assertEqual(tracked, [])
        self.assertTrue(any(set(g) >= {0, 1} for g in new))

    def test_keeper_prefers_corpus_folder(self):
        df = pd.DataFrame({
            "caminho_ficheiro": [
                r"E:\outros\file__hash.pdf",
                r"E:\todos os textos\(1944)_Music.pdf",
            ],
            "contexto": ["aaa", "bbbb"],
        })
        self.assertGreater(
            qa.keeper_score(df, 1)[0],
            qa.keeper_score(df, 0)[0],
        )


class TestApplyIdempotent(unittest.TestCase):
    def test_columns_not_duplicated_on_rerun(self):
        wb = Workbook()
        ws = wb.active
        ws.title = qa.SHEET
        headers = [
            "contexto", "matched_form", "nuclear", "candidato_duplicado",
            "doc_id", "grupo_passagem_id", "caminho_ficheiro", "dominio",
        ]
        for c, h in enumerate(headers, 1):
            ws.cell(1, c).value = h
        ws.cell(2, 1).value = "the complexity of the orchestral texture is clear"
        ws.cell(2, 2).value = "complexity"
        ws.cell(2, 3).value = True
        ws.cell(2, 5).value = "d1"
        ws.cell(2, 7).value = r"E:\todos os textos\(2000)_X.pdf"
        ws.cell(2, 8).value = "musicologia"

        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "t.xlsx"
            out = Path(tmp) / "t_qa.xlsx"
            wb.save(path)
            r1 = qa.run_qa(path, apply=True, outfile=out)
            self.assertIn("written", r1)
            r2 = qa.run_qa(out, apply=True, outfile=out)
            self.assertEqual(r2.get("duplicates_demoted"), 0)
            ws2 = load_workbook(out)[qa.SHEET]
            names = [ws2.cell(1, c).value for c in range(1, ws2.max_column + 1)]
            for col in qa.QA_COLS:
                self.assertEqual(names.count(col), 1, col)
            self.assertEqual(
                ws2.cell(2, names.index("qa_relacao") + 1).value,
                "direct",
            )


if __name__ == "__main__":
    unittest.main()
