#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Orquestração CLI da extracção NEAR (corpo de textura_near.main)."""

from __future__ import annotations

import argparse
import hashlib
import sys
from collections import Counter
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

try:
    import textura_triagem as ttri
except ImportError:
    ttri = None

import textura_lexico as tlex

from textura.config import CAMPO, NOS, RELACOES_NUCLEARES, SCHEMA_NEAR
from textura.duplicados import (
    agregar_ocorrencias, atribuir_match_ids, deduplicar_hits_exactos,
    fundir_janelas_e_marcar_duplicados, occurrence_id_de,
)
from textura.exportacao import reordenar_colunas_hits
from textura.lexico import dominio_janela
from textura.relacoes import anota_com_heuristica, anota_com_spacy
from textura.tokenizacao import (
    _Consulta, anota_sintaxe, compila_campo, emparelha_contexto, fronteiras_frase, indices_no, normaliza,
    procura_near, tokeniza,
)

# conta_tokens lives in estatistica — import under real name below
from textura.estatistica import (
    POLO_ESTABILIDADE,
    POLO_VARIABILIDADE,
    conta_tokens,
)


def _configurar_consola() -> None:
    """Evita UnicodeEncodeError na consola Windows (cp1252)."""
    for stream in (sys.stdout, sys.stderr):
        try:
            stream.reconfigure(encoding="utf-8", errors="replace")
        except Exception:
            pass


def main() -> int:
    _configurar_consola()
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--xlsx", required=True, type=Path)
    ap.add_argument("--folha", default="Neighbor Contexts")
    ap.add_argument("--near", type=int, default=4)
    ap.add_argument("--lingua", default="en", choices=list(NOS) + ["todas"])
    ap.add_argument("--limite", type=int, default=None,
                    help="processar apenas as N primeiras linhas (teste)")
    ap.add_argument("--sem-fronteira", action="store_true",
                    help="não aplicar exclusão por fronteira de frase")
    ap.add_argument("--termos", type=Path, default=None,
                    help="ficheiro de texto com o campo lexical, uma linha por "
                         "tipo: 'etiqueta = padrao1, padrao2, ...'. Substitui CAMPO.")
    ap.add_argument("--consulta", default=None,
                    help="expressão booleana sobre etiquetas, avaliada por janela. "
                         "Ex.: \"(uniform OR constant) AND NOT varied\"")
    ap.add_argument("--banda", type=int, default=12,
                    help="limite da banda de referência para as medidas de "
                         "associação (tokens); deve exceder --near")
    ap.add_argument("--col-no", type=int, default=6,
                    help="nº da coluna do nó (1 = primeira)")
    ap.add_argument("--col-ctx", type=int, default=15,
                    help="nº da coluna do contexto integral")
    ap.add_argument("--col-src", type=int, default=12,
                    help="nº da coluna da fonte/ficheiro")
    ap.add_argument("--com-cabecalho", action="store_true",
                    help="a folha tem linha de cabeçalho")
    ap.add_argument("--col-url", type=int, default=13,
                    help="nº da coluna com a hiperligação (0 = nenhuma)")
    ap.add_argument("--sintaxe", default="spacy",
                    choices=["heuristica", "spacy"],
                    help="método de identificação da relação sintáctica "
                         "(omissão: spacy; use heuristica para comparação)")
    ap.add_argument("--modelo", default="en_core_web_sm",
                    help="modelo spaCy (en_core_web_sm, pt_core_news_sm, ...)")
    ap.add_argument("--incluir-nao-nucleares", action="store_true",
                    help="não filtrar a estatística por nuclear=True")
    ap.add_argument("--inverter-polaridade-negada", action="store_true",
                    help="inverter polaridade quando negado=True (omissão: não)")
    ap.add_argument("--dominios", type=Path, default=None,
                    help="TSV padrao_ficheiro\\tdominio para triagem documental "
                         "(omissão: dominios.tsv junto do projecto, se existir)")
    ap.add_argument("--incluir-dominio", action="append", default=None,
                    help="domínio a readmitir (repetível); omissão: só musicologia")
    ap.add_argument(
        "--dominio-omissao",
        default="musicologia",
        help="domínio quando o TSV não casa (omissão: musicologia; "
             "vazio '' = manter por_rever)",
    )
    ap.add_argument(
        "--sem-falsos-amigos",
        action="store_true",
        help="não aplicar exclusão automática de falsos amigos "
             "(continuo/continuum/continue…)",
    )
    ap.add_argument("--saida", type=Path, default=Path("resultado_near.xlsx"))
    ap.add_argument("--so-extrair", action="store_true", default=True,
                    help="(omissao) so extrai para revisao; sem estatistica")
    ap.add_argument("--sem-revisao", action="store_true",
                    help="extraccao + analise numa so passagem "
                         "(escreve aviso «sem revisao humana»)")
    ap.add_argument("--fases", type=int, choices=[1, 2], default=None,
                    help="alias: 1 = so extrair (fase 1); "
                         "2 = extrair+analisar sem revisao")
    args = ap.parse_args()
    if args.fases == 1:
        args.so_extrair = True
        args.sem_revisao = False
    elif args.fases == 2:
        args.sem_revisao = True
        args.so_extrair = False
    if args.sem_revisao:
        args.so_extrair = False

    campo = dict(CAMPO)
    if args.termos:
        campo = tlex.carregar_campo_termos(args.termos, campo_ref=CAMPO)
        # sincronizar pólos mutáveis
        POLO_ESTABILIDADE.update(tlex.POLO_ESTABILIDADE)
        POLO_VARIABILIDADE.update(tlex.POLO_VARIABILIDADE)
        print(f"      campo lexical externo: {len(campo)} tipos "
              f"({', '.join(sorted(campo))})", flush=True)
    else:
        tlex.registar_campo(campo)
        tlex.assert_campo_sem_no(campo)

    consulta = _Consulta(args.consulta, set(campo)) if args.consulta else None

    print(f"[1/5] A ler {args.xlsx.name} ...", flush=True)
    bruto = pd.read_excel(args.xlsx, sheet_name=args.folha,
                          header=0 if args.com_cabecalho else None,
                          nrows=args.limite)
    total_bruto = len(bruto)
    ncols = bruto.shape[1]
    for rot, k in (("no", args.col_no), ("contexto", args.col_ctx)):
        if not 1 <= k <= ncols:
            print(f"Coluna de {rot} ({k}) fora do intervalo 1-{ncols}.",
                  file=sys.stderr)
            return 2
    col_src = tlex.escolher_coluna_fonte(bruto, args.col_src)
    # Nº de linha Excel 1-based (com cabeçalho: dados começam na linha 2).
    excel_row0 = 2 if args.com_cabecalho else 1
    df = pd.DataFrame({
        "NODE": bruto.iloc[:, args.col_no - 1],
        "contexto": bruto.iloc[:, args.col_ctx - 1],
        "caminho": bruto.iloc[:, col_src - 1],
        "url": (bruto.iloc[:, args.col_url - 1] if 1 <= args.col_url <= ncols
                else ""),
        "source_matrix_row": np.arange(
            excel_row0, excel_row0 + len(bruto), dtype=np.int64),
    })
    print(f"      {total_bruto} linhas, {ncols} colunas | no=col{args.col_no} "
          f"contexto=col{args.col_ctx} fonte=col{col_src} | schema={SCHEMA_NEAR}",
          flush=True)

    # --- limpeza declarada -------------------------------------------------
    df = df.dropna(subset=["contexto"])
    nos_validos = (set().union(*NOS.values()) if args.lingua == "todas"
                   else NOS[args.lingua])
    df["NODE"] = df["NODE"].astype(str).str.lower()
    df = df[df["NODE"].isin(nos_validos)]
    print(f"      {total_bruto} linhas -> {len(df)} após filtro de língua "
          f"'{args.lingua}' e contexto não vazio", flush=True)

    campo_c = compila_campo(campo)
    registos, censura_esq, censura_dir, sem_no = 0, 0, 0, 0
    excluidos_tot = Counter()
    linhas, janelas = [], []
    tot_near = tot_banda = 0
    hits_near, hits_banda = Counter(), Counter()
    partes_termo: dict[str, Counter] = {e: Counter() for e in campo}
    tam_parte: Counter = Counter()

    print(f"[2/5] A extrair co-ocorrências NEAR/{args.near} ...", flush=True)
    for t in df.itertuples(index=False):
        ctx = normaliza(t.contexto)
        toks = tokeniza(ctx)
        if not toks:
            continue
        idxs = indices_no(toks, nos_validos)
        if not idxs:
            sem_no += 1
            continue
        limites = [] if args.sem_fronteira else fronteiras_frase(ctx)
        mesma_frase = not args.sem_fronteira

        # Emparelhamento: produto cartesiano nó×termo → par mínimo (T1).
        achados, excl, n_nos = emparelha_contexto(
            toks, nos_validos, campo_c, args.near, limites,
            mesma_frase=mesma_frase)
        for k, v in excl.items():
            excluidos_tot[k] += v

        # Nó de referência para censura / banda: o mais frequente entre os
        # pares vencedores; se vazio, o mais à esquerda na janela.
        if achados:
            cont_nos = Counter(a["idx_no"] for a in achados)
            i = cont_nos.most_common(1)[0][0]
        else:
            i = min(idxs)

        c_esq = i < args.near
        c_dir = (len(toks) - i - 1) < args.near
        censura_esq += c_esq
        censura_dir += c_dir

        # Banda de referência: termos além de NEAR, ancorados no mesmo nó.
        achados_banda = procura_near(
            toks, i, campo_c, args.banda, limites, mesma_frase=mesma_frase)
        distantes = [a for a in achados_banda
                     if a["distancia"] > args.near]
        tipos_near = {a["termo_tipo"] for a in achados}

        tot_near += conta_tokens(toks, i, limites, 0, args.near, mesma_frase)
        tot_banda += conta_tokens(toks, i, limites, args.near, args.banda,
                                  mesma_frase)
        for a in distantes:
            if a["termo_tipo"] not in tipos_near:
                hits_banda[a["termo_tipo"]] += 1
        for a in achados:
            hits_near[a["termo_tipo"]] += 1
            partes_termo[a["termo_tipo"]][t.caminho] += 1
        tam_parte[t.caminho] += 1

        caminho_f = str(t.caminho)
        doc_id = tlex.doc_id_de_caminho(caminho_f)
        source_row = int(t.source_matrix_row)
        occ_id = occurrence_id_de(doc_id, source_row)
        for a in achados:
            neg, grad, mod, rel = anota_sintaxe(
                toks, a["idx_no"], a["idx_termo"], ctx)
            can = tlex.canonical_de_forma(a["matched_form"], campo)
            # query_pattern: primeiro padrão da etiqueta canónica
            qpat = (campo.get(can) or [a["matched_form"]])[0]
            pol_base = tlex.polaridade(can, False, inverter_negada=False)
            pol = tlex.polaridade(can, neg,
                                 inverter_negada=args.inverter_polaridade_negada)
            linhas.append({
                "source_matrix_row": source_row,
                "texture_occurrence_id": occ_id,
                "match_id": "",  # preenchido após extracção
                "hit_key": "",
                "grupo_passagem_id": "",
                "candidato_duplicado": "",
                "no": a["no"],
                "termo_tipo": can,
                "canonical_term": can,
                "query_pattern": qpat,
                "termo_forma": a["termo_forma"],
                "matched_form": a["matched_form"],
                "n_palavras": a["n_palavras"],
                "distancia": a["distancia"],
                "lado": a["lado"],
                "negado": neg if neg is None else ("directo" if neg else "nao"),
                "graduado": grad,
                "modalizado": mod,
                "relacao_sintactica": rel,
                "polaridade_base": pol_base or "",
                "polaridade": pol or "",
                "eixo": tlex.eixo_semantico(can),
                "censurado_esq": c_esq,
                "censurado_dir": c_dir,
                "idx_no": int(a["idx_no"]),
                "idx_termo": int(a["idx_termo"]),
                "off_no": a["off_no"],
                "off_termo": a["off_termo"],
                "n_nos_janela": n_nos,
                "forma_em_composto": a["forma_em_composto"],
                "caminho_ficheiro": caminho_f,
                "doc_id": doc_id,
                "url": t.url,
                "contexto": ctx,
                "motivo_exclusao": "",
                "nuclear": rel in RELACOES_NUCLEARES,
                "fonte_classificacao": "heuristica",
                "n_janelas_fundidas": 1,
                "revisao_sugerida": "",
                "nucleo_da_propriedade": "",
            })
        if achados:
            presentes = {a["termo_tipo"] for a in achados}
            janelas.append({
                "no": toks[i][0],
                "n_termos": len(presentes),
                "termos": " + ".join(sorted(presentes)),
                "conjunto": presentes,
                "consulta_satisfeita": (consulta.avalia(presentes)
                                        if consulta else None),
                "caminho": t.caminho,
                "contexto": ctx,
            })
        registos += 1

    res = pd.DataFrame(linhas)
    print(f"      {registos} linhas analisadas | {len(res)} co-ocorrencias | "
          f"{sem_no} sem no localizavel | "
          f"excluidos fronteira_frase={excluidos_tot['fronteira_frase']}",
          flush=True)
    if res.empty:
        print("Nenhuma co-ocorrencia. Verifique o campo lexical.", file=sys.stderr)
        return 1

    # R1 — asserção defensiva sobre o output
    tlex.assert_output_sem_no(res["canonical_term"].unique())

    # R2 — doc_id 1-por-linha + caminhos sem extensao = coluna errada
    frac_ficheiro = float(res["caminho_ficheiro"].map(
        tlex.parece_caminho_ficheiro).mean())
    if frac_ficheiro < 0.2:
        raise SystemExit(
            "Assercao falhou: caminho_ficheiro nao contem ficheiros "
            f"(fracao com extensao={frac_ficheiro:.2f}). "
            "Verifique --col-src (nao use a coluna 'raiz'/directorio).")
    if (len(res) > 5 and res["doc_id"].nunique() == len(res)
            and frac_ficheiro < 0.5):
        raise SystemExit(
            "Assercao falhou: n_doc_id == n_linhas com caminhos duvidosos. "
            "Verifique --col-src.")

    if args.sintaxe == "spacy":
        print("[2b/5] Analise de dependencias (spaCy) ...", flush=True)
        res = anota_com_spacy(res, args.modelo, obrigatorio=True)
    else:
        print("[2b/5] Classificacao heuristica ...", flush=True)
        res = anota_com_heuristica(res)

    n_bruto = len(res)
    por_rever = pd.DataFrame(columns=["caminho", "n_hits", "n_nucleares"])
    if ttri is not None:
        dom_path = args.dominios
        if dom_path is None:
            # Project root (…/Term statistics), not the textura/ package dir
            cand = Path(__file__).resolve().parents[1] / "dominios.tsv"
            if cand.is_file():
                dom_path = cand
        regras = ttri.carregar_dominios(dom_path)
        incluir = set(args.incluir_dominio) if args.incluir_dominio else None
        omis = (args.dominio_omissao or "").strip() or None
        res, _, por_rever = ttri.aplicar_triagem(
            res,
            regras_dominio=regras,
            incluir_dominios=incluir,
            dominio_omissao=omis,
            aplicar_amigos=not args.sem_falsos_amigos,
        )
        for i, row in res.iterrows():
            toks = tokeniza(str(row["contexto"]))
            i_no = next((k for k, (_, o) in enumerate(toks)
                         if o == int(row["off_no"])), None)
            i_te = next((k for k, (_, o) in enumerate(toks)
                         if o == int(row["off_termo"])), None)
            if i_no is not None and i_te is not None and ttri.e_ruido_ocr(
                    toks, i_no, i_te):
                res.at[i, "nuclear"] = False
                if not res.at[i, "motivo_exclusao"]:
                    res.at[i, "motivo_exclusao"] = "ruido_ocr"

    # R7 — mesma assinatura (contexto + percurso) => mesmo nuclear/motivo
    if "percurso_dep" in res.columns:
        for _, grp in res.groupby(
                [res["contexto"].map(lambda c: " ".join(
                    w for w, _ in tokeniza(str(c)))),
                 "percurso_dep"], sort=False):
            if len(grp) < 2:
                continue
            # consenso: se algum nuclear=True sem metatexto residual, alinhar
            nucs = grp["nuclear"].astype(bool)
            if nucs.nunique() > 1:
                # preferir o veredicto maioritário; empate -> False
                ver = bool(nucs.mode().iloc[0])
                mot = grp.loc[grp["nuclear"] == ver, "motivo_exclusao"]
                mot0 = mot.iloc[0] if len(mot) else ""
                for i in grp.index:
                    res.at[i, "nuclear"] = ver
                    res.at[i, "motivo_exclusao"] = (
                        "" if ver else (res.at[i, "motivo_exclusao"] or mot0
                                        or "indeterminada"))

    # IDs de hit + dedupe exacto + fusão de janelas (mesmo termo)
    res = atribuir_match_ids(res)
    n_antes_dedupe = len(res)
    res = deduplicar_hits_exactos(res)
    res = fundir_janelas_e_marcar_duplicados(res, ngrama=8)
    # Rematch IDs após eventual remoção de cópias exactas
    if len(res) != n_antes_dedupe:
        res = atribuir_match_ids(res)
    # Domínio por janela (complementa a triagem documental por caminho)
    if "contexto" in res.columns:
        res["dominio_janela"] = res["contexto"].map(dominio_janela)
        if "revisao_sugerida" in res.columns:
            _mx = res["dominio_janela"].astype(bool)
            res.loc[_mx, "revisao_sugerida"] = res.loc[_mx].apply(
                lambda r: ((str(r["revisao_sugerida"]) + "; ")
                           if r["revisao_sugerida"] else "")
                + "dominio_janela:" + r["dominio_janela"], axis=1)
    res = reordenar_colunas_hits(res)

    res_excluidas = res.loc[~res["nuclear"].astype(bool)].copy() if (
        "nuclear" in res.columns) else res.iloc[0:0].copy()
    if not args.incluir_nao_nucleares and "nuclear" in res.columns:
        res_stat = res.loc[res["nuclear"].astype(bool)].copy()
    else:
        res_stat = res.copy()
    n_nuc = len(res_stat)
    n_ficheiros = (res["caminho_ficheiro"].nunique()
                   if "caminho_ficheiro" in res.columns else 0)
    n_obras = res["doc_id"].nunique() if "doc_id" in res.columns else 0
    n_ocorrencias = (res["texture_occurrence_id"].nunique()
                     if "texture_occurrence_id" in res.columns else 0)
    n_ocorrencias_nuc = (res_stat["texture_occurrence_id"].nunique()
                         if "texture_occurrence_id" in res_stat.columns else 0)
    print(f"      cascata: brutas={n_bruto} -> nucleares(hits)={n_nuc} "
          f"({100*n_nuc/max(n_bruto,1):.1f}%) | "
          f"excluidas={len(res_excluidas)} | "
          f"ocorrencias={n_ocorrencias} (nucleares={n_ocorrencias_nuc}) | "
          f"ficheiros={n_ficheiros} obras/doc_id={n_obras}", flush=True)

    col_doc = "doc_id" if "doc_id" in res_stat.columns else "caminho_ficheiro"
    if "texture_occurrence_id" in res_stat.columns:
        res_obra = res_stat.drop_duplicates(
            subset=["texture_occurrence_id", "termo_tipo"])
    else:
        res_obra = res_stat.drop_duplicates(subset=[col_doc, "termo_tipo"])
    ocorrencias_df = agregar_ocorrencias(res)

    # Folha Duplicados: caminhos múltiplos + grupos de passagem/janela
    dup_rows = []
    if "doc_id" in res.columns:
        g = (res.groupby("doc_id")["caminho_ficheiro"]
             .agg(lambda s: sorted(set(map(str, s))))
             .reset_index())
        for row in g.itertuples(index=False):
            if len(row.caminho_ficheiro) > 1:
                dup_rows.append({
                    "tipo": "mesmo_doc_id_varios_caminhos",
                    "doc_id": row.doc_id,
                    "grupo": "",
                    "n": len(row.caminho_ficheiro),
                    "detalhe": " | ".join(row.caminho_ficheiro),
                })
    if "grupo_passagem_id" in res.columns:
        gpass = res.loc[
            res["grupo_passagem_id"].astype(str).str.strip().ne("")
            & ~res["grupo_passagem_id"].astype(str).str.lower().isin(
                {"nan", "none"})
        ]
        for gid, grp in gpass.groupby("grupo_passagem_id", sort=False):
            termos = sorted({str(t) for t in grp["canonical_term"]})
            n_jan = int((grp["motivo_exclusao"].astype(str)
                         == "janela_sobreposta").sum())
            dup_rows.append({
                "tipo": "passagem_sobreposta",
                "doc_id": grp["doc_id"].iloc[0],
                "grupo": str(gid),
                "n": int(len(grp)),
                "detalhe": (
                    f"hits={len(grp)} janela_sobreposta={n_jan} "
                    f"termos={'; '.join(termos)} | "
                    f"{str(grp['contexto'].iloc[0])[:100]}"
                ),
            })
    duplicados = pd.DataFrame(dup_rows)

    # Colunas de revisao humana (fase 1)
    if "revisto_por_humano" not in res.columns:
        res["revisto_por_humano"] = ""
    if "nota_revisao" not in res.columns:
        res["nota_revisao"] = ""

    # --- FASE 1: sempre escrever Excel de revisao -------------------------
    from datetime import datetime as _dt
    comando = " ".join(sys.argv)
    manifesto = hashlib.sha256(
        pd.Series(df["caminho"].astype(str).unique()).sort_values()
        .str.cat(sep="\n").encode("utf-8", errors="replace")
    ).hexdigest()
    instr = pd.DataFrame({
        "chave": [
            "fase", "schema_near", "data_fase1", "comando", "sem_revisao",
            "janelas_kwic_processadas", "tot_near", "tot_banda",
            "hits_banda", "campo_tipos", "manifesto_sha256",
            "n_hits", "n_ocorrencias", "n_hits_nucleares",
            "n_ocorrencias_nucleares",
            "como_rever", "comando_fase2", "unidade_contagem",
        ],
        "valor": [
            "1 - extracao",
            SCHEMA_NEAR,
            _dt.now().isoformat(timespec="seconds"),
            comando,
            "sim" if args.sem_revisao else "nao",
            registos, tot_near, tot_banda,
            repr(dict(hits_banda)),
            ",".join(sorted(campo)),
            manifesto,
            len(res), n_ocorrencias, n_nuc, n_ocorrencias_nuc,
            "Edite as colunas a amarelo em 8_Concordancia (= hits NEAR). "
            "Nao altere source_matrix_row, texture_occurrence_id, match_id, "
            "hit_key, canonical_term nem matched_form. "
            "8_Concordancia_Ocorrencias e so leitura (1 linha = 1 linha da matriz). "
            "N_hits = linhas nucleares em 8_Concordancia; "
            "N_ocorrencias = texture_occurrence_id unicos com hit nuclear.",
            f'python textura_analise.py --xlsx "{args.saida}"',
            "hit=8_Concordancia; ocorrencia=texture_occurrence_id "
            "(linha da matriz)",
        ],
    })
    cfg_lex = pd.DataFrame([
        {"etiqueta": k, "padroes": ", ".join(v),
         "polaridade": (tlex.polaridade(k) or ""),
         "eixo": tlex.eixo_semantico(k),
         "nota_eixo": (
             "PROPOSTA: varied->invariancia_diacronica (ou ambos); "
             "nao alterar sem adjudicação"
             if k == "varied" else "")}
        for k, v in campo.items()
    ])
    drop_cols = [c for c in ("relacao", "caminho_dep", "atribuicao", "caminho")
                 if c in res.columns]
    conc_out = res.drop(columns=drop_cols, errors="ignore")
    # Alias explícito do nível hits (mesma folha de trabalho)
    hits_out = conc_out
    if len(ocorrencias_df) == 0:
        ocorrencias_df = agregar_ocorrencias(res)

    print(f"[3/5] A escrever extraccao (fase 1) -> {args.saida.name} ...",
          flush=True)
    with pd.ExcelWriter(args.saida, engine="openpyxl") as xw:
        instr.to_excel(xw, sheet_name="0_Instrucoes", index=False)
        cfg_lex.to_excel(xw, sheet_name="Config_lexico", index=False)
        pd.DataFrame({
            "caminho": sorted(df["caminho"].astype(str).unique()),
        }).to_excel(xw, sheet_name="Manifesto_corpus", index=False)
        conc_out.to_excel(xw, sheet_name="8_Concordancia", index=False)
        hits_out.to_excel(xw, sheet_name="8_Concordancia_Hits", index=False)
        if len(ocorrencias_df):
            ocorrencias_df.to_excel(
                xw, sheet_name="8_Concordancia_Ocorrencias", index=False)
        if len(res_excluidas):
            res_excluidas.drop(columns=drop_cols, errors="ignore").to_excel(
                xw, sheet_name="9_Excluidas", index=False)
        if len(por_rever):
            por_rever.to_excel(xw, sheet_name="Dominios_por_rever", index=False)
        if len(duplicados):
            duplicados.to_excel(xw, sheet_name="Duplicados", index=False)

    # Validacao de dados + destaque amarelo nas colunas editaveis
    try:
        from openpyxl.worksheet.datavalidation import DataValidation
        wb = load_workbook(args.saida)
        wsc = wb["8_Concordancia"]
        cab = [c.value for c in wsc[1]]
        amarelo = PatternFill("solid", fgColor="FFF2CC")
        editaveis = {
            "relacao_sintactica": ",".join(sorted(
                RELACOES_NUCLEARES | {
                    "incidental", "adverbial_verbal", "adverbial_de_grau",
                    "coordenada", "indeterminada"})),
            "nuclear": "TRUE,FALSE",
            "polaridade": "estabilidade,variabilidade,",
            "eixo": "homogeneidade_sincronica,invariancia_diacronica,ambos,",
            "negado": "nao,directo,indirecto",
        }
        # Sombrear apenas linhas ainda nao revistas: a marca amarela e um
        # convite a edicao, pelo que desaparece quando revisto_por_humano
        # esta preenchido — evita ruido visual apos a revisao.
        j_rev = (cab.index("revisto_por_humano") + 1
                 if "revisto_por_humano" in cab else None)

        def _linha_por_rever(r: int) -> bool:
            if j_rev is None:
                return True
            v = wsc.cell(row=r, column=j_rev).value
            return v is None or str(v).strip() == ""

        for col_nome, lista in editaveis.items():
            if col_nome not in cab:
                continue
            j = cab.index(col_nome) + 1
            letra = get_column_letter(j)
            for r in range(2, wsc.max_row + 1):
                if _linha_por_rever(r):
                    wsc.cell(row=r, column=j).fill = amarelo
            dv = DataValidation(type="list", formula1=f'"{lista}"',
                                allow_blank=True)
            dv.error = "Valor fora da taxonomia"
            dv.errorTitle = "Invalido"
            wsc.add_data_validation(dv)
            dv.add(f"{letra}2:{letra}{wsc.max_row}")
        # tambem amarelo em dominio / motivo / revisao / candidato
        for col_nome in ("dominio", "motivo_exclusao", "revisto_por_humano",
                         "nota_revisao", "candidato_duplicado"):
            if col_nome in cab:
                j = cab.index(col_nome) + 1
                for r in range(2, wsc.max_row + 1):
                    if _linha_por_rever(r):
                        wsc.cell(row=r, column=j).fill = amarelo
        wb.save(args.saida)
    except Exception as exc:
        print(f"      AVISO: validacao Excel nao aplicada ({exc})", flush=True)

    if not args.sem_revisao:
        print("\n=== FASE 1 concluida ===")
        print(f"Reveja a folha 8_Concordancia (hits) em:\n  {args.saida}")
        print("Folha 8_Concordancia_Ocorrencias = 1 linha por linha da matriz.")
        print("Depois corra:")
        print(f'  python textura_analise.py --xlsx "{args.saida}"')
        print(f"Cascata: hits={n_bruto} nucleares(hits)={n_nuc} "
              f"ocorrencias={n_ocorrencias} (nuc={n_ocorrencias_nuc}) "
              f"ficheiros={n_ficheiros} doc_id={n_obras}")
        return 0

    # --- FASE integrada (--sem-revisao): analisa de seguida ---------------
    print("[4/5] --sem-revisao: a correr fase 2 (AVISO: sem revisao humana) ...",
          flush=True)
    import textura_analise as tanal
    return tanal.analisar(args.saida, args.saida, nulo_polaridade="banda",
                          cooc_unidade="obra")
