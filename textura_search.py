#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
textura_search.py — pesquisa booleana + exportação de resultados
===============================================================

Ponto de entrada principal (CLI e GUI). Lê a matriz KWIC, avalia a
consulta (AND OR NOR NOT NEAR/x * ?) e escreve um livro Excel com:

  Results          — File, Doc, Match, Snippet, Query, Hyperlink
  Frequencias      — documentos e ocorrências por forma casada
  Resumo           — indicadores
  Graficos         — figuras embutidas

Opcionalmente (--analise-near) corre também a mineração de
co-ocorrências de textura_near.py.

Uso:
    python textura_search.py
    python textura_search.py --consulta "music* NEAR/4 texture*"
    python textura_search.py --consulta "uniform* OR constant*" --limite 5000
"""

from __future__ import annotations

import argparse
import math
import subprocess
import sys
from collections import Counter
from pathlib import Path

_AQUI = Path(__file__).resolve().parent
if str(_AQUI) not in sys.path:
    sys.path.insert(0, str(_AQUI))

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

import textura_legendas as tleg
import textura_lexico as tlex
import textura_near as tn
import textura_plots as tplot
import textura_query as tq

# ---------------------------------------------------------------------------
# Folha / saída por omissão — a matriz KWIC NÃO tem default (GUI ou --xlsx)
# ---------------------------------------------------------------------------
DEFAULT_FOLHA = "Neighbor Contexts"
# saídas fora da pasta do código (evitar poluir o projecto)
DEFAULT_SAIDA = Path(r"C:\Users\lmr20\Desktop\EXCEL_list") / "resultado_pesquisa.xlsx"

COL_NO, COL_CTX, COL_SRC, COL_URL = 6, 15, 12, 13

_MAP_REL_PESQUISA = {
    "caracterizacao_atributiva": "atributiva",
    "caracterizacao_predicativa": "predicativa",
    "especificacao_preposicional": "atributiva",
    "definicao_metalinguistica": "atributiva",
    "atributiva": "atributiva",
    "predicativa": "predicativa",
    "indeterminada": "indeterminada",
}


def _canonical_de_forma(forma: str, campo_lex: dict[str, list[str]]) -> str:
    for etq, pads in campo_lex.items():
        if any(tq.forma_casa_padrao(forma, p) for p in pads):
            return etq
    return forma


def _relacao_para_analise(hit_rel: str | None, tokens, no_idx: int,
                          idx_termo: int) -> str:
    """Normaliza para atributiva | predicativa | indeterminada."""
    if hit_rel:
        for parte in str(hit_rel).split(","):
            m = _MAP_REL_PESQUISA.get(parte.strip().lower())
            if m:
                return m
    *_, rel = tn.anota_sintaxe(tokens, no_idx, idx_termo)
    return rel


def _linhas_concordancia_analise(
        ctx: str, toks, hit, formas_campo: list[str],
        campo_lex: dict[str, list[str]], *,
        caminho: str, url: str, no: str,
) -> list[dict]:
    """Uma linha por forma do campo — compatível com textura_analise.py."""
    formas_tok = [t[0] for t in toks]
    idxs_no = [i for i, w in enumerate(formas_tok)
               if tq.forma_e_no(w)]
    if not idxs_no:
        # fallback: token igual ao nó da matriz
        idxs_no = [i for i, w in enumerate(formas_tok)
                   if w == str(no).lower()]
    if not idxs_no:
        idxs_no = [len(toks) // 2]
    i_no = idxs_no[len(idxs_no) // 2]

    dist_por_forma: dict[str, int] = {}
    if hit.distancias_near and hit.spans:
        # associa distância NEAR à forma do span do campo
        for d, (a, b) in zip(hit.distancias_near, hit.spans):
            forma = " ".join(formas_tok[k] for k in range(a, b))
            if forma in formas_campo:
                dist_por_forma[forma] = int(d)

    rel_hit = hit.relacoes[0] if hit.relacoes else None
    out = []
    for forma in formas_campo:
        # índice do termo: 1.º token da forma no contexto
        idx_termo = None
        partes = forma.split()
        L = len(partes)
        for j in range(len(formas_tok) - L + 1):
            if formas_tok[j:j + L] == partes:
                idx_termo = j
                break
        if idx_termo is None:
            idx_termo = i_no
        rel = _relacao_para_analise(rel_hit, toks, i_no, idx_termo)
        neg, graduado, modalizado, _rel_h = tn.anota_sintaxe(
            toks, i_no, idx_termo)
        can = _canonical_de_forma(forma, campo_lex)
        pol = _inferir_polaridade(can, campo_lex.get(can, [forma]))
        dist = dist_por_forma.get(forma)
        if dist is None:
            dist = abs(idx_termo - i_no)
        lado = "esq" if idx_termo < i_no else ("dir" if idx_termo > i_no else "—")
        # padrão da consulta
        qpat = next((p for p in campo_lex.get(can, []) if p), can)
        out.append({
            "canonical_term": can,
            "query_pattern": qpat,
            "matched_form": forma,
            "polaridade": pol or "",
            "dominio": "campo_lexical",
            "motivo_exclusao": "",
            "relacao_sintactica": rel,
            "negado": bool(neg),
            "graduado": bool(graduado),
            "modalizado": bool(modalizado),
            "caminho": caminho,
            "contexto": ctx,
            "distancia": dist,
            "lado": lado,
            "no": str(no).lower(),
            "url": url,
        })
    return out


def _grafico_docs(freq_docs: pd.DataFrame, destino: Path, topo: int | None = None,
                  leg: dict | None = None, html_dir: Path | None = None):
    leg = leg or tleg.PADRAO
    sub = freq_docs if topo is None else freq_docs.head(topo)
    rotulos = [str(d)[:40] + ("..." if len(str(d)) > 40 else "")
               for d in sub["documento"]]
    n_show, n_tot = len(sub), len(freq_docs)
    d = leg["docs"]
    titulo = f"{d['titulo']}  ·  {n_show}/{n_tot}"
    subtitulo = d.get("subtitulo", "")
    xlabel = d.get("xlabel", "Ocorrências")
    rodape = leg.get("rodape", "")
    tplot.barras_horizontais(
        rotulos, sub["ocorrencias"].tolist(), destino,
        titulo=titulo, subtitulo=subtitulo, xlabel=xlabel,
        cmap="docs", rodape=rodape, max_n=topo,
    )
    if html_dir is not None:
        import textura_html as thtml
        thtml.barras_horizontais(
            rotulos, sub["ocorrencias"].tolist(),
            html_dir / "03_docs.html",
            titulo=titulo, subtitulo=subtitulo, xlabel=xlabel, rodape=rodape,
        )


def _grafico_formas(freq_formas: pd.DataFrame, destino: Path,
                    topo: int | None = None, leg: dict | None = None,
                    html_dir: Path | None = None):
    leg = leg or tleg.PADRAO
    sub = freq_formas if topo is None else freq_formas.head(topo)
    n_show, n_tot = len(sub), len(freq_formas)
    d = leg["formas"]
    titulo = f"{d['titulo']}  ·  {n_show}/{n_tot}"
    subtitulo = d.get("subtitulo", "")
    xlabel = d.get("xlabel", "Ocorrências")
    rodape = leg.get("rodape", "")
    tplot.barras_horizontais(
        sub["forma"].astype(str).tolist(),
        sub["ocorrencias"].tolist(), destino,
        titulo=titulo, subtitulo=subtitulo, xlabel=xlabel,
        cmap="formas", rodape=rodape, max_n=topo,
    )
    if html_dir is not None:
        import textura_html as thtml
        thtml.barras_horizontais(
            sub["forma"].astype(str).tolist(),
            sub["ocorrencias"].tolist(),
            html_dir / "04_formas.html",
            titulo=titulo, subtitulo=subtitulo, xlabel=xlabel, rodape=rodape,
        )


def _grafico_near(dists: list[int], destino: Path, leg: dict | None = None,
                  html_dir: Path | None = None):
    leg = leg or tleg.PADRAO
    d = leg["near"]
    kwargs = dict(
        titulo=d["titulo"],
        subtitulo=d.get("subtitulo", ""),
        rodape=leg.get("rodape", ""),
        xlabel=d.get("xlabel", "Distância (tokens)"),
        ylabel=d.get("ylabel", "Nº de pares"),
        rotulo_mediana=d.get("mediana", "Mediana"),
        rotulo_media=d.get("media", "Média"),
    )
    tplot.histograma_near(dists, destino, **kwargs)
    if html_dir is not None:
        import textura_html as thtml
        thtml.histograma_near(dists, html_dir / "05_near.html", **kwargs)


def _grafico_nuvem(freq_formas: pd.DataFrame, destino: Path, leg: dict | None = None,
                   html_dir: Path | None = None):
    leg = leg or tleg.PADRAO
    d = leg["nuvem"]
    freqs = {str(r.forma): int(r.ocorrencias)
             for r in freq_formas.itertuples(index=False)}
    titulo, subtitulo = d["titulo"], d.get("subtitulo", "")
    rodape = leg.get("rodape", "")
    tplot.nuvem_palavras(
        freqs, destino, titulo=titulo, subtitulo=subtitulo, rodape=rodape,
    )
    if html_dir is not None and destino.exists():
        import textura_html as thtml
        thtml.nuvem_com_png(
            destino, html_dir / "02_nuvem.html",
            titulo=titulo, subtitulo=subtitulo, rodape=rodape,
        )


def _pares_sankey(res: pd.DataFrame,
                  padroes: list[str] | None = None) -> pd.DataFrame:
    rows = []
    for _, row in res.iterrows():
        match = str(row.get("Match", "") or "")
        doc = str(row.get("Doc", "") or "")
        if not match or not doc or match == "nan" or doc == "nan":
            continue
        for f in match.replace(";", ",").split(","):
            f = f.strip()
            if not f:
                continue
            if padroes and not tq.forma_no_lexico(f, padroes):
                continue
            rows.append((f, doc, 1))
    if not rows:
        return pd.DataFrame(columns=["forma", "documento", "peso"])
    lig = pd.DataFrame(rows, columns=["forma", "documento", "peso"])
    return lig.groupby(["forma", "documento"], as_index=False)["peso"].sum()


def _grafico_sankey(res: pd.DataFrame, destinos: list[Path],
                    leg: dict | None = None,
                    html_dir: Path | None = None,
                    padroes: list[str] | None = None) -> list[Path]:
    """Gera Sankey em TODOS os caminhos dados. Devolve lista dos PNG criados."""
    import shutil

    leg = leg or tleg.PADRAO
    d = leg["sankey"]
    lig = _pares_sankey(res, padroes=padroes)
    if lig.empty:
        print("      [sankey] sem pares forma/documento — PNG nao gerado", flush=True)
        return []

    print(f"      [sankey] {len(lig)} fluxos", flush=True)

    # primeiro destino = ficheiro mestre; os outros sao copias
    destinos = [Path(p) for p in destinos]
    mestre = destinos[0]
    mestre.parent.mkdir(parents=True, exist_ok=True)
    sk_kw = dict(
        titulo=d["titulo"],
        subtitulo=d.get("subtitulo", ""),
        eixo_esq=d.get("eixo_esq", "Formas casadas"),
        eixo_dir=d.get("eixo_dir", "Documentos"),
        rodape=leg.get("rodape", ""),
    )
    try:
        tplot.sankey_formas_docs(lig, mestre, **sk_kw)
    except Exception as exc:  # noqa: BLE001
        print(f"      [sankey ERRO] {exc}", flush=True)
        return []

    if not mestre.exists() or mestre.stat().st_size == 0:
        print("      [sankey] FALHOU: ficheiro vazio", flush=True)
        return []

    criados = [mestre]
    for extra in destinos[1:]:
        try:
            extra.parent.mkdir(parents=True, exist_ok=True)
            shutil.copy2(mestre, extra)
            criados.append(extra)
        except OSError as exc:
            print(f"      [sankey] copia falhou -> {extra}: {exc}", flush=True)

    for p in criados:
        print(f"      [sankey] GUARDADO: {p.resolve()}", flush=True)

    if html_dir is not None:
        import textura_html as thtml
        thtml.sankey(lig, html_dir / "01_sankey.html", **sk_kw)

    return criados


def _ler_termos_adjudicados(caminho: Path | None) -> dict[str, list[str]] | None:
    """Ficheiro 'etiqueta = padrao1, padrao2' (mesmo formato que textura_near)."""
    if caminho is None or not Path(caminho).exists():
        return None
    campo: dict[str, list[str]] = {}
    for linha in Path(caminho).read_text(encoding="utf-8").splitlines():
        linha = linha.split("#", 1)[0].strip()
        if not linha or "=" not in linha:
            continue
        etq, pads = linha.split("=", 1)
        etq = etq.split(":", 1)[0].strip()
        campo[etq] = [p.strip() for p in pads.split(",") if p.strip()]
    return campo or None


def _inferir_polaridade(etiqueta: str, padroes: list[str]) -> str | None:
    """Liga etiqueta/padrões da consulta ao léxico canónico (E/V)."""
    pol = tlex.polaridade(etiqueta)
    if pol:
        return pol
    # tentar stems dos padroes (uniform* -> uniform)
    for p in padroes:
        stem = tlex.stem_de_padrao(p) if hasattr(tlex, "stem_de_padrao") else (
            p.strip("*").lower())
        pol = tlex.polaridade(stem)
        if pol:
            return pol
        # prefixo curto contra polos
        s = (p or "").strip("*").lower()
        if any(s.startswith(x[: max(4, len(s))]) or x.startswith(s)
               for x in tlex.POLO_ESTABILIDADE):
            return "estabilidade"
        if any(s.startswith(x[: max(4, len(s))]) or x.startswith(s)
               for x in tlex.POLO_VARIABILIDADE):
            return "variabilidade"
    return None


def _escrever_termos(caminho: Path, campo: dict[str, list[str]]) -> Path:
    """Escreve etiqueta[:E|V] = padroes — exclui node_patterns."""
    linhas = []
    for etq, pads in campo.items():
        pads_ok = [p for p in pads if not tq.e_padrao_no(p)]
        if not pads_ok:
            continue
        pol = _inferir_polaridade(etq, pads_ok)
        tag = ""
        if pol == "estabilidade":
            tag = ":E"
        elif pol == "variabilidade":
            tag = ":V"
        linhas.append(f"{etq}{tag} = {', '.join(pads_ok)}")
    Path(caminho).write_text("\n".join(linhas) + "\n", encoding="utf-8")
    return Path(caminho)


def pesquisar(xlsx: Path, consulta: str, folha: str = DEFAULT_FOLHA,
              limite: int | None = None, com_cabecalho: bool = False,
              col_no: int = COL_NO, col_ctx: int = COL_CTX,
              col_src: int = COL_SRC, col_url: int = COL_URL,
              saida: Path = DEFAULT_SAIDA,
              mesma_frase: bool = True,
              exigir_sintaxe: bool = True,
              legendas: Path | dict | None = None,
              termos: Path | None = None,
              com_graficos: bool = False) -> Path:
    """Fase 1: pesquisa e escreve Excel de resultados (sem estatística/gráficos).

    Por omissão NÃO gera gráficos nem índices estatísticos — use
    --com-graficos ou textura_analise.py após rever o Excel.
    Duplicados da coluna de contexto (col. O) são removidos só em memória.
    """
    q = tq.ConsultaBooleana(consulta, mesma_frase=mesma_frase,
                            exigir_sintaxe=exigir_sintaxe)
    campo_ext = _ler_termos_adjudicados(termos)
    if campo_ext:
        padroes_brutos = list(dict.fromkeys(
            p for pads in campo_ext.values() for p in pads))
        if q.padroes:
            padroes_q = set(q.padroes)
            inter = [p for p in padroes_brutos if p in padroes_q]
            padroes_brutos = inter or padroes_brutos
        campo_lex = {etq: [p for p in pads if p in padroes_brutos]
                     for etq, pads in campo_ext.items()}
        campo_lex = {k: v for k, v in campo_lex.items() if v}
    else:
        padroes_brutos = list(q.padroes)
        campo_lex = tq.campo_desde_padroes(padroes_brutos)

    # req. 1–2: node_patterns fora do campo lexical / Match / Formas / NEAR
    padroes_no, padroes = tq.separar_node_e_campo(padroes_brutos)
    campo_lex = {
        etq: [p for p in pads if p in padroes]
        for etq, pads in campo_lex.items()
    }
    campo_lex = {k: v for k, v in campo_lex.items() if v}
    # avaliação da consulta mantém todos os padrões; estatísticas só o campo
    q.padroes = list(padroes_brutos)

    print(f"[1/4] A ler {xlsx.name} ...", flush=True)
    print(f"      node_patterns: {padroes_no or tlex.NODE_PATTERNS}", flush=True)
    print(f"      campo lexical: {len(padroes)} padroes / "
          f"{len(campo_lex)} tipos (textur* excluido do campo)", flush=True)
    bruto = pd.read_excel(xlsx, sheet_name=folha,
                          header=0 if com_cabecalho else None,
                          nrows=limite)
    n_cols = bruto.shape[1]
    for rot, k in (("nó", col_no), ("contexto", col_ctx), ("fonte", col_src)):
        if not 1 <= k <= n_cols:
            raise SystemExit(f"Coluna de {rot} ({k}) fora de 1-{n_cols}")

    df = pd.DataFrame({
        "no": bruto.iloc[:, col_no - 1].astype(str),
        "contexto": bruto.iloc[:, col_ctx - 1],
        "caminho": bruto.iloc[:, col_src - 1].astype(str),
        "url": (bruto.iloc[:, col_url - 1].astype(str)
                if 1 <= col_url <= n_cols else ""),
    }).dropna(subset=["contexto"])
    n_bruto = len(df)
    # desduplicação exclusiva pela coluna O (contexto / snippet)
    df["_ctx_chave"] = df["contexto"].map(
        lambda v: tq.normaliza(v) if pd.notna(v) else "")
    df = df[df["_ctx_chave"] != ""].drop_duplicates(
        subset=["_ctx_chave"], keep="first").drop(columns=["_ctx_chave"])
    n_dup = n_bruto - len(df)
    print(f"      {n_bruto} linhas -> {len(df)} apos remover {n_dup} "
          f"duplicados (col. O / contexto)", flush=True)
    print(f"      filtros: mesma_frase={mesma_frase}  "
          f"sintaxe={exigir_sintaxe}", flush=True)

    print(f"[2/4] A avaliar '{consulta}' ...", flush=True)
    resultados, dists_near = [], []
    formas_c, docs_c = Counter(), Counter()

    for i, row in enumerate(df.itertuples(index=False), 1):
        ctx = tq.normaliza(row.contexto)
        toks = tq.tokeniza(ctx)
        if not toks:
            continue
        hit = q.avalia(toks, texto=ctx)
        if not hit.ok:
            continue
        # formas da consulta, depois só o CAMPO (nunca node/textur*)
        formas_hit = q.filtra_formas(hit.formas)
        if not formas_hit and hit.spans:
            formas_hit = q.filtra_formas([
                " ".join(toks[k][0] for k in range(a, b))
                for a, b in hit.spans
            ])
        formas_campo = [
            f for f in formas_hit
            if tq.forma_no_lexico(f, padroes) and not tq.forma_e_no(f)
        ]
        if not formas_campo:
            continue
        snip = tq.snippet_destacado(ctx, toks, hit.spans)
        doc = tq.nome_documento(row.caminho)
        link = tq.hiperligacao(row.url, row.caminho)
        match = ", ".join(formas_campo)
        rel = ", ".join(hit.relacoes) if hit.relacoes else ""
        resultados.append({
            "File": str(row.caminho),
            "Doc": doc,
            "Match": match,
            "Relation": rel,
            "Snippet": snip,
            "Query": consulta,
            "Node": row.no,
            "Hyperlink": link,
            "Context_full": ctx,
        })
        # F1: 8_Concordancia de análise NÃO é construída aqui —
        # delega-se a textura_near.py (--extrair-near, omissão).
        for f in formas_campo:
            formas_c[f] += 1
        docs_c[doc] += 1
        dists_near.extend(hit.distancias_near)
        if i % 25000 == 0:
            print(f"      ... {i}/{len(df)}  hits={len(resultados)}", flush=True)

    res = pd.DataFrame(resultados)
    print(f"      {len(res)} resultados em {res['Doc'].nunique() if len(res) else 0} documentos",
          flush=True)
    if res.empty:
        print("Nenhum resultado. Experimente alargar a consulta.", file=sys.stderr)
        # ainda assim escreve um resumo vazio útil
        with pd.ExcelWriter(saida, engine="openpyxl") as xw:
            pd.DataFrame({"indicador": ["Consulta", "Resultados"],
                          "valor": [consulta, 0]}).to_excel(
                xw, sheet_name="Resumo", index=False)
            pd.DataFrame(columns=["File", "Doc", "Match", "Snippet",
                                  "Query", "Hyperlink"]).to_excel(
                xw, sheet_name="Results", index=False)
        return saida

    saida = Path(saida)
    freq_docs = (pd.DataFrame({"documento": list(docs_c),
                               "ocorrencias": list(docs_c.values())})
                 .sort_values("ocorrencias", ascending=False)
                 .reset_index(drop=True))
    freq_formas = (pd.DataFrame({"forma": list(formas_c),
                                 "ocorrencias": list(formas_c.values())})
                   .sort_values("ocorrencias", ascending=False)
                   .reset_index(drop=True))
    if len(freq_formas):
        freq_formas["proporcao"] = (
            freq_formas["ocorrencias"] / freq_formas["ocorrencias"].sum()
        ).round(4)
    N = int(freq_formas["ocorrencias"].sum()) if len(freq_formas) else 0
    V = len(freq_formas)

    # inventário do léxico × formas atestadas (para revisão)
    tab_termos = []
    for etq, pads in campo_lex.items():
        formas_tipo = [
            f for f in freq_formas["forma"].astype(str)
            if any(tq.forma_casa_padrao(f, p) for p in pads)
        ] if len(freq_formas) else []
        n_occ = int(freq_formas.loc[
            freq_formas["forma"].isin(formas_tipo), "ocorrencias"
        ].sum()) if formas_tipo else 0
        tab_termos.append({
            "tipo": etq,
            "padroes": ", ".join(pads),
            "formas_atestadas": ", ".join(formas_tipo) if formas_tipo else "—",
            "n_formas": len(formas_tipo),
            "ocorrencias": n_occ,
            "atestado": "sim" if n_occ else "não",
        })
    df_termos = pd.DataFrame(tab_termos)

    pares = Counter()
    for formas in res["Match"].str.split(", "):
        formas = [f for f in (formas or []) if f and tq.forma_no_lexico(f, padroes)]
        u = sorted(set(formas))
        for a_i in range(len(u)):
            for b_i in range(a_i + 1, len(u)):
                pares[(u[a_i], u[b_i])] += 1
    cooc = (pd.DataFrame(
        [{"forma_A": a, "forma_B": b, "hits_juntos": n}
         for (a, b), n in pares.most_common(200)]
    ) if pares else pd.DataFrame(columns=["forma_A", "forma_B", "hits_juntos"]))

    termos_path = saida.parent / f"{saida.stem}_termos_adjudicados.txt"
    _escrever_termos(termos_path, campo_lex)

    resumo_inds = [
        "Consulta", "Linhas brutas lidas", "Duplicados removidos (col. O)",
        "Linhas KWIC examinadas", "Recusar termos em frases diferentes",
        "Exigir relação sintáctica", "Resultados (hits)", "Documentos únicos",
        "Formas distintas casadas (V)", "Ocorrências de formas (N)",
        "Padrões adjudicados (léxico)", "Tipos adjudicados",
        "Fase", "Proximo passo",
    ]
    resumo_vals = [
        consulta, n_bruto, n_dup, len(df),
        "sim" if mesma_frase else "não",
        "sim" if exigir_sintaxe else "não",
        len(res), res["Doc"].nunique(), V, N,
        len(padroes), len(campo_lex),
        "1 — pesquisa / Excel para revisão",
        "Rever folha 8_Concordancia (relacao_sintactica); "
        "depois Analisar Excel… na GUI",
    ]

    figs_dir = saida.parent / f"{saida.stem}_figs"
    g_sankey_lado = saida.parent / f"{saida.stem}_SANKEY.png"
    g_sankey = figs_dir / "01_sankey.png"
    g_nuvem = figs_dir / "02_nuvem.png"
    g_docs = figs_dir / "03_docs.png"
    g_formas = figs_dir / "04_formas.png"
    g_near = figs_dir / "05_near.png"
    leg = tleg.PADRAO

    if com_graficos:
        print("[3/4] Estatistica e graficos (opcional) ...", flush=True)
        cont = freq_formas["ocorrencias"].values if len(freq_formas) else np.array([])
        p = cont / cont.sum() if len(cont) else cont
        H = float(-(p * np.log(p)).sum()) if len(p) else float("nan")
        J = H / math.log(V) if V > 1 else float("nan")
        resumo_inds += [
            "Entropia de Shannon (H)", "Equitabilidade de Pielou (J)",
            "Distância NEAR mediana", "Distância NEAR média",
        ]
        resumo_vals += [
            round(H, 4) if H == H else "-",
            round(J, 4) if J == J else "-",
            (float(np.median(dists_near)) if dists_near else "-"),
            (round(float(np.mean(dists_near)), 3) if dists_near else "-"),
        ]
        if isinstance(legendas, dict):
            leg = tleg.carregar(None, consulta=consulta)
            for k, v in legendas.items():
                if isinstance(v, dict) and isinstance(leg.get(k), dict):
                    leg[k].update({kk: vv for kk, vv in v.items()
                                   if vv is not None})
                elif v is not None:
                    leg[k] = v
        else:
            leg = tleg.carregar(legendas, consulta=consulta)
        figs_dir.mkdir(parents=True, exist_ok=True)
        print(f"      pasta resultados: {saida.parent.resolve()}", flush=True)
        _grafico_sankey(res, [g_sankey_lado, g_sankey], leg=leg,
                        html_dir=figs_dir, padroes=padroes)
        _grafico_nuvem(freq_formas, g_nuvem, leg=leg, html_dir=figs_dir)
        _grafico_docs(freq_docs, g_docs, leg=leg, html_dir=figs_dir)
        _grafico_formas(freq_formas, g_formas, leg=leg, html_dir=figs_dir)
        if dists_near:
            _grafico_near(dists_near, g_near, leg=leg, html_dir=figs_dir)
        try:
            import textura_html as thtml
            thtml.indice(figs_dir, [
                ("Sankey (formas → documentos)", figs_dir / "01_sankey.html"),
                ("Nuvem de formas", figs_dir / "02_nuvem.html"),
                ("Dispersão por documentos", figs_dir / "03_docs.html"),
                ("Formas casadas", figs_dir / "04_formas.html"),
                ("Distâncias NEAR", figs_dir / "05_near.html"),
            ], titulo="Gráficos TEXTURA", consulta=consulta)
        except Exception as exc:  # noqa: BLE001
            print(f"      [html indice] {exc}", flush=True)
    else:
        print("[3/4] Sem gráficos/estatística (fase 1 — só Excel para revisão)",
              flush=True)

    resumo = pd.DataFrame({"indicador": resumo_inds, "valor": resumo_vals})

    print(f"[4/4] A escrever {saida.name} ...", flush=True)
    results_out = res[["File", "Doc", "Match", "Relation", "Snippet", "Query",
                       "Node", "Hyperlink"]].copy()

    cfg_export = pd.DataFrame([
        {"papel": "node_pattern", "valor": p,
         "canonical_term": "", "polaridade": "", "dominio": "no"}
        for p in tlex.NODE_PATTERNS
    ] + [
        {"papel": "search_pattern", "valor": p,
         "canonical_term": etq,
         "polaridade": (_inferir_polaridade(etq, pads) or ""),
         "dominio": "campo_lexical"}
        for etq, pads in campo_lex.items() for p in pads
    ])

    with pd.ExcelWriter(saida, engine="openpyxl") as xw:
        results_out.to_excel(xw, sheet_name="Results", index=False)
        resumo.to_excel(xw, sheet_name="Resumo", index=False)
        df_termos.to_excel(xw, sheet_name="Termos", index=False)
        freq_docs.to_excel(xw, sheet_name="Docs", index=False)
        freq_formas.to_excel(xw, sheet_name="Formas", index=False)
        if not cooc.empty:
            cooc.to_excel(xw, sheet_name="Coocorrencia", index=False)
        res[["Doc", "Match", "Snippet", "Context_full", "File", "Hyperlink"]].to_excel(
            xw, sheet_name="Concordancia", index=False)
        # F1: sem 8_Concordancia própria — só textura_near.py a produz
        cfg_export.to_excel(xw, sheet_name="Config_lexico", index=False)

    wb = load_workbook(saida)
    ws = wb["Results"]
    cab = [c.value for c in ws[1]]
    i_doc = cab.index("Doc") + 1 if "Doc" in cab else cab.index("File") + 1
    i_link = cab.index("Hyperlink") + 1
    for lin in range(2, ws.max_row + 1):
        alvo = ws.cell(row=lin, column=i_link).value
        if isinstance(alvo, str) and alvo.strip():
            cel = ws.cell(row=lin, column=i_doc)
            cel.hyperlink = alvo.strip()
            cel.style = "Hyperlink"

    if com_graficos:
        png_sankey = g_sankey_lado if g_sankey_lado.exists() else (
            g_sankey if g_sankey.exists() else None)
        if png_sankey is not None:
            try:
                ws_sk = wb.create_sheet("Sankey", 1)
                ws_sk["A1"] = leg["sankey"]["titulo"]
                ws_sk["A1"].font = Font(name="Calibri", bold=True, size=14,
                                        color="1F2A24")
                tplot.embeber_imagens(ws_sk, [png_sankey], linha0=5,
                                      largura_px=900)
            except Exception as exc:  # noqa: BLE001
                print(f"      Excel Sankey falhou: {exc}", flush=True)
        wg = wb.create_sheet("Graficos")
        outros = [p for p in (g_nuvem, g_docs, g_formas, g_near) if p.exists()]
        if png_sankey is not None:
            outros = [png_sankey] + outros
        try:
            tplot.embeber_imagens(wg, outros, linha0=6)
        except Exception as exc:  # noqa: BLE001
            print(f"      Excel Graficos falhou: {exc}", flush=True)

    fill = PatternFill("solid", fgColor="E8EEE9")
    for nome in wb.sheetnames:
        s = wb[nome]
        if s.max_row < 1:
            continue
        for c in s[1]:
            if c.value is None:
                continue
            c.font = Font(name="Calibri", bold=True, color="1F2A24")
            c.fill = fill
            c.alignment = Alignment(vertical="center", wrap_text=True)
        if nome not in ("Graficos", "Sankey"):
            s.freeze_panes = "A2"
        for col in range(1, min(s.max_column or 1, 10) + 1):
            letra = get_column_letter(col)
            s.column_dimensions[letra].width = 18 if col < 4 else 42
    wb.save(saida)

    print("\n" + resumo.to_string(index=False))
    print(f"\nConcluido: {saida.resolve()}")
    print(f"Termos:    {termos_path.resolve()}")
    if com_graficos:
        print(f"Figuras:   {figs_dir.resolve()}")
    else:
        print("Nota: gráficos/estatística = passo seguinte "
              "(textura_analise.py / Analisar Excel na GUI)")
    return saida


def main() -> int:
    ap = argparse.ArgumentParser(
        description="Pesquisa booleana na matriz TEXTURA (AND OR NOR NOT NEAR/x * ?)")
    ap.add_argument("--xlsx", type=Path, required=True,
                    help="matriz KWIC (.xlsx) — obrigatório, sem default")
    ap.add_argument("--folha", default=DEFAULT_FOLHA)
    ap.add_argument("--consulta", required=True,
                    help='ex.: "music* NEAR/4 texture*"  ou  '
                         '"(uniform* OR constant*) AND NOT varied*"')
    ap.add_argument("--saida", type=Path, default=DEFAULT_SAIDA)
    ap.add_argument("--limite", type=int, default=None)
    ap.add_argument("--com-cabecalho", action="store_true")
    ap.add_argument("--col-no", type=int, default=COL_NO)
    ap.add_argument("--col-ctx", type=int, default=COL_CTX)
    ap.add_argument("--col-src", type=int, default=COL_SRC)
    ap.add_argument("--col-url", type=int, default=COL_URL)
    ap.add_argument("--extrair-near", action="store_true", default=True,
                    help="após a pesquisa, extrai 8_Concordancia via "
                         "textura_near.py (omissão: sim)")
    ap.add_argument("--sem-extrair-near", action="store_true",
                    help="nao delegar extraccao a textura_near "
                         "(so Results — sem analise estatistica)")
    ap.add_argument("--analise-near", action="store_true",
                    help=argparse.SUPPRESS)  # alias legado → --extrair-near
    ap.add_argument("--com-graficos", action="store_true",
                    help="(opcional) gerar também gráficos na fase de pesquisa")
    ap.add_argument("--permitir-outra-frase", action="store_true",
                    help="aceitar termos separados por ponto / fronteira de frase")
    ap.add_argument("--sem-sintaxe", action="store_true",
                    help="não exigir relação sintáctica entre os termos")
    ap.add_argument("--legendas", type=Path, default=None,
                    help="JSON com titulos/subtitulos editaveis dos graficos")
    ap.add_argument("--termos", type=Path, default=None,
                    help="lista adjudicada (etiqueta = padrao1, padrao2). "
                         "Fecha o léxico de Results/gráficos/NEAR")
    args = ap.parse_args()

    if not args.xlsx.exists():
        print(f"Matriz não encontrada:\n  {args.xlsx}", file=sys.stderr)
        return 2

    saida = pesquisar(
        xlsx=args.xlsx, consulta=args.consulta, folha=args.folha,
        limite=args.limite, com_cabecalho=args.com_cabecalho,
        col_no=args.col_no, col_ctx=args.col_ctx,
        col_src=args.col_src, col_url=args.col_url,
        saida=args.saida,
        mesma_frase=not args.permitir_outra_frase,
        exigir_sintaxe=not args.sem_sintaxe,
        legendas=args.legendas,
        termos=args.termos,
        com_graficos=args.com_graficos,
    )

    extrair = ((args.extrair_near or args.analise_near)
               and not args.sem_extrair_near)
    if extrair:
        print("\n-- Extracção NEAR (unica via para 8_Concordancia / analise) --",
              flush=True)
        near_out = Path(saida).with_name(Path(saida).stem + "_near.xlsx")
        termos_auto = Path(saida).with_name(
            Path(saida).stem + "_termos_adjudicados.txt")
        termos_usar = args.termos if args.termos else termos_auto
        qtmp = tq.ConsultaBooleana(args.consulta, mesma_frase=True,
                                   exigir_sintaxe=False)
        cmd = [sys.executable, str(Path(__file__).with_name("textura_near.py")),
               "--xlsx", str(args.xlsx),
               "--folha", args.folha,
               "--saida", str(near_out),
               "--so-extrair",
               "--termos", str(termos_usar),
               "--col-no", str(args.col_no),
               "--col-ctx", str(args.col_ctx),
               "--col-src", str(args.col_src)]
        if args.com_cabecalho:
            cmd.append("--com-cabecalho")
        if qtmp.near_n:
            cmd += ["--near", str(qtmp.near_n)]
        if args.limite:
            cmd += ["--limite", str(args.limite)]
        print(" ".join(cmd), flush=True)
        rc = subprocess.call(cmd)
        if rc == 0:
            print(f"Concordancia para revisao/analise: {near_out}", flush=True)
            print("NAO use o Excel de Results para textura_analise.py.",
                  flush=True)
        return rc
    print("AVISO: --sem-extrair-near — sem 8_Concordancia defensavel; "
          "textura_analise.py recusara este livro.", flush=True)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
