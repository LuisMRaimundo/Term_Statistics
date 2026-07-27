#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
textura_near.py — mineração de co-ocorrências NEAR/x sobre a matriz KWIC
=======================================================================

Opera sobre TEXTURA_TUDO_MATRIZ_v*.xlsx (folha 'Neighbor Contexts', sem
linha de cabeçalho).

ONTOLOGIA DE TRÊS NÍVEIS (schema ≥ 2)
------------------------------------
1. Ocorrência mestra de textura = uma linha da matriz
   (`source_matrix_row` → `texture_occurrence_id`).
   Unidade primária para contar «quantas ocorrências têm a propriedade P».
2. Hit NEAR = um termo do campo lexical perto do nó nessa ocorrência
   (`match_id`, `hit_key`). Uma ocorrência pode ter vários hits.
3. Janela de contexto exportada = objecto de exibição. Janelas KWIC
   deslocadas do mesmo hit são fundidas (`janela_sobreposta`); não
   constituem observações independentes.

Folhas: `8_Concordancia` (= hits, revisão humana) e
`8_Concordancia_Ocorrencias` (agregado por ocorrência mestra).

DECISÃO METODOLÓGICA CENTRAL
----------------------------
As colunas de vizinhança (1-5 e 7-11) NÃO são usadas para o cálculo da
distância. Razão: contêm palavras de conteúdo com remoção de palavras
funcionais e com janelas assimétricas — o que torna a contagem de
posições não comparável com uma janela NEAR/x em palavras. Além disso,
não conservam pontuação, pelo que não permitem detectar fronteira de
frase.

Toda a análise assenta na coluna 15 (contexto integral), que conserva
pontuação e ordem. As colunas de vizinhança são conservadas apenas como
metadados.

LIMITAÇÃO A DECLARAR
--------------------
O campo de contexto está truncado: mediana de 11 tokens à esquerda e 10
à direita. Em ~12,5% dos casos há menos de 5 tokens à direita. Essas
linhas são marcadas como censuradas (campo 'censurado_dir' /
'censurado_esq'): a ausência de co-ocorrência nelas não é evidência de
ausência, e deve ser tratada como dado em falta, não como zero.

Os offsets (`off_no`, `off_termo`, `idx_no`, `idx_termo`) são absolutos
*dentro do contexto da linha da matriz*, não no documento integral.

Uso:
    python textura_near.py --xlsx CAMINHO.xlsx --near 4 --lingua en
    python textura_near.py --xlsx CAMINHO.xlsx --near 4 --limite 20000
"""

from __future__ import annotations

import argparse
import hashlib
import math
import re
import sys
import unicodedata
from collections import Counter, defaultdict
from pathlib import Path

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from scipy import stats

try:
    import textura_stats as tst_avancada
except ImportError:
    tst_avancada = None

try:
    import textura_triagem as ttri
except ImportError:
    ttri = None

import textura_lexico as tlex

# ---------------------------------------------------------------------------
# 1. CONFIGURAÇÃO — editar aqui
# ---------------------------------------------------------------------------

COLUNAS = ["L5", "L4", "L3", "L2", "L1", "NODE",
           "R1", "R2", "R3", "R4", "R5",
           "caminho", "url", "raiz", "contexto", "n"]

# Formas do nó por língua (col. 6). Ajustar conforme necessário.
NOS = {
    "en": {"texture", "textures", "textural", "textured", "texturally", "texturing"},
    "pt": {"textura", "texturas", "textural", "texturais"},
    "de": {"textur", "texturen"},
}

# Campo lexical. Chave = etiqueta do tipo lexical; valor = truncaturas.
# O asterisco final funciona como truncatura à direita (prefixo).
CAMPO = {
    # --- núcleo da invariância -------------------------------------------
    "uniform":      ["uniform*"],
    "invariable":   ["invariab*", "invarian*"],
    "unvarying":    ["unvarying"],
    "immutable":    ["immutab*"],
    "unchanging":   ["unchanging", "unchanged"],
    # --- sinónimos e relacionados ----------------------------------------
    "constant":     ["constant*"],
    "consistent":   ["consisten*"],
    "regular":      ["regular*"],
    "stable":       ["stab*"],
    "steady":       ["stead*"],
    "sustained":    ["sustain*"],
    "static":       ["static", "stasis"],
    "monotonous":   ["monoton*"],
    # --- campo contrastante ----------------------------------------------
    "varied":       ["varied", "variety", "variet*"],
    "varying":      ["varying", "variab*"],
    "changing":     ["changing", "changeab*"],
    "irregular":    ["irregular*"],
    "unequal":      ["unequal", "uneven*"],
    "diverse":      ["divers*"],
    "mutable":      ["mutab*"],
    "multiform":    ["multiform*"],
    # --- eixo da homogeneidade (controlo) --------------------------------
    "homogeneous":  ["homogene*", "homogenous"],
    "heterogeneous": ["heterogene*"],
    # --- controlo negativo -----------------------------------------------
    "loud":         ["loud*"],
    "orchestral":   ["orchestral"],
}

# Negação: "no" NÃO entra aqui (colisões com «no. 1» — ver _nega_no_opus).
NEGACAO = {"not", "never", "nor", "neither", "without", "lacking",
           "hardly", "scarcely", "barely", "rarely", "seldom",
           "n't", "cannot", "isn", "aren", "wasn", "weren", "doesn", "don",
           "absence", "lack", "devoid"}

# Graduação (antes chamada incorrectamente «modalização»).
GRADUACAO = {"less", "more", "quite", "rather", "fairly", "somewhat",
             "relatively", "certain", "some", "largely", "broadly",
             "mostly", "generally", "almost", "nearly", "increasingly",
             "essentially", "virtually", "comparatively", "slightly",
             "very", "highly", "fully", "completely", "entirely"}
MODALIZACAO = GRADUACAO  # alias de leitura legado

# Operadores modais / evidenciais (coluna modalizado).
MODALIDADE = {
    "may", "might", "could", "would", "can", "should", "must",
    "seems", "seem", "appears", "appear", "apparently", "perhaps",
    "possibly", "presumably", "arguably", "suggests", "tends",
    "likely", "unlikely", "pode", "poderia", "parece", "talvez",
    "seemingly",
}

RELACOES_NUCLEARES = frozenset({
    "atributiva", "predicativa", "predicativa_secundaria",
    "nominal_composto", "nominal_genitiva", "adverbial",
})

# Cópulas: presença entre nó e termo indicia construção predicativa.
COPULAS = {"is", "are", "was", "were", "be", "been", "being",
           "remains", "remain", "remained", "becomes", "become", "became",
           "seems", "seem", "seemed", "appears", "appear", "appeared",
           "stays", "stayed"}

# Abreviaturas que terminam em ponto sem fechar frase.
ABREVIATURAS = {
    "p", "pp", "vol", "vols", "no", "nos", "ed", "eds", "cf", "ibid", "op",
    "cit", "et", "al", "e.g", "i.e", "fig", "figs", "ex", "exx", "ms", "mss",
    "mr", "mrs", "ms", "dr", "prof", "st", "ca", "c", "n", "trans", "rev",
    "repr", "diss", "univ", "publ", "chap", "chaps", "sec", "secs", "bk",
    "jan", "feb", "mar", "apr", "jun", "jul", "aug", "sep", "sept", "oct",
    "nov", "dec", "esp", "viz", "etc", "bwv", "kv", "hob", "d", "k",
}

# ---------------------------------------------------------------------------
# 2. SEGMENTAÇÃO E TOKENIZAÇÃO
# ---------------------------------------------------------------------------

# Hífen interno preservado (di-uniform, rotation-invariant); hífen de
# fim de linha continua a ser colado por RE_HIFEN_QUEBRA em normaliza().
RE_TOKEN = re.compile(
    r"[A-Za-zÀ-ÖØ-öø-ÿ](?:[A-Za-zÀ-ÖØ-öø-ÿ'’]|-(?=[A-Za-zÀ-ÖØ-öø-ÿ]))*"
)
RE_FIM_FRASE = re.compile(r"[.!?;]+[\"'”’\)\]]*\s")
RE_HIFEN_QUEBRA = re.compile(r"(\w)-\s+(\w)")


def normaliza(texto: str) -> str:
    """Repara quebras de linha hifenizadas ('propor- tions' -> 'proportions')."""
    texto = str(texto).replace("\n", " ").replace("\r", " ")
    texto = RE_HIFEN_QUEBRA.sub(r"\1\2", texto)
    return re.sub(r"\s+", " ", texto)


def fronteiras_frase(texto: str) -> list[int]:
    """Devolve os offsets de carácter onde termina uma frase.

    Um ponto não fecha frase quando: (i) a palavra que o precede é uma
    abreviatura conhecida; (ii) está entre dígitos (decimal); (iii) é
    parte de reticências. O ponto-e-vírgula é tratado como fronteira,
    por separar predicações autónomas — decisão a declarar no protocolo.
    """
    limites = []
    for m in RE_FIM_FRASE.finditer(texto + " "):
        i = m.start()
        anterior = texto[max(0, i - 30):i]
        palavra = re.search(r"([A-Za-zÀ-ÿ.]+)$", anterior)
        if palavra:
            cand = palavra.group(1).lower().rstrip(".")
            if cand in ABREVIATURAS:
                continue
        if i > 0 and texto[i - 1].isdigit() and i + 1 < len(texto) and texto[i + 1:i + 2].isdigit():
            continue
        if texto[i:i + 3] == "...":
            continue
        limites.append(i)
    return limites


def indice_frase(pos: int, limites: list[int]) -> int:
    """Número da frase a que pertence um offset de carácter."""
    return sum(1 for L in limites if L < pos)


def tokeniza(texto: str) -> list[tuple[str, int]]:
    """Lista de (token minúsculo, offset de carácter)."""
    return [(m.group(0).lower(), m.start()) for m in RE_TOKEN.finditer(texto)]


# ---------------------------------------------------------------------------
# 3. OPERADOR NEAR/x
# ---------------------------------------------------------------------------

def _rx_palavra(p: str) -> re.Pattern:
    """Compila uma palavra, com truncatura à direita (*) e/ou à esquerda.

    Truncatura só à direita («uniform*», «static*»): ancora no início do
    token e NÃO atravessa hífen — assim «static*» não casa «static-chordal».
    Truncatura à esquerda («*uniform»): permite hífen e casa o composto
    pelo núcleo final («di-uniform»).
    """
    esq = p.startswith("*")
    dir_ = p.endswith("*")
    nucleo = re.escape(p.strip("*"))
    if esq and dir_:
        star = r"[\w\-]{0,20}"
        return re.compile("^" + star + nucleo + star + "$")
    if esq:
        return re.compile(r"^(?:[\w\-]*-)?" + nucleo + r"$")
    if dir_:
        # sem hífen no wildcard — A15 / T2
        return re.compile("^" + nucleo + r"\w{0,20}$")
    return re.compile("^" + nucleo + "$")


def compila_campo(campo: dict[str, list[str]]):
    """Converte truncaturas em sequências de padrões de token."""
    saida = []
    for etiqueta, padroes in campo.items():
        seqs = [[_rx_palavra(w) for w in p.split()] for p in padroes]
        seqs.sort(key=len, reverse=True)
        saida.append((etiqueta, seqs))
    return saida


def _casa_em(tokens, j, seq) -> int:
    """Devolve o comprimento da sequência se casar em j; senão 0."""
    if j + len(seq) > len(tokens):
        return 0
    for k, rx in enumerate(seq):
        if not rx.match(tokens[j + k][0]):
            return 0
    return len(seq)


def indices_no(tokens, nos_validos: set[str]) -> list[int]:
    """Índices de todas as ocorrências do nó no contexto tokenizado."""
    return [i for i, (w, _) in enumerate(tokens) if w in nos_validos]


def melhor_par_tokens(tokens, idxs_no: list[int], idx_termo: int,
                      n_palavras: int, limites: list[int],
                      mesma_frase: bool = True):
    """Como melhor_par, com verificação de frase sobre offsets reais."""
    span = list(range(idx_termo, idx_termo + n_palavras))
    melhor = None  # (dist, pos_esq, i_no)
    for i_no in idxs_no:
        if i_no in span:
            continue
        if mesma_frase and limites:
            f_no = indice_frase(tokens[i_no][1], limites)
            if any(indice_frase(tokens[k][1], limites) != f_no for k in span):
                continue
        dist = min(abs(k - i_no) for k in span)
        pos_esq = min(i_no, idx_termo)
        cand = (dist, pos_esq, i_no)
        if melhor is None or cand < melhor:
            melhor = cand
    if melhor is None:
        return None
    dist, _, i_no = melhor
    lado = "esq" if idx_termo < i_no else ("dir" if idx_termo > i_no else "—")
    return i_no, idx_termo, dist, lado


def encontra_termos(tokens, campo_compilado) -> list[dict]:
    """Todas as ocorrências do campo lexical no contexto (sem fixar nó)."""
    achados = []
    vistos = set()  # (etiqueta, idx_termo, n_palavras)
    for j in range(len(tokens)):
        for etiqueta, seqs in campo_compilado:
            casou = 0
            for seq in seqs:
                casou = _casa_em(tokens, j, seq)
                if casou:
                    break
            if not casou:
                continue
            chave = (etiqueta, j, casou)
            if chave in vistos:
                continue
            vistos.add(chave)
            forma = " ".join(tokens[k][0] for k in range(j, j + casou))
            achados.append({
                "termo_tipo": etiqueta,
                "termo_forma": forma,
                "idx_termo": j,
                "n_palavras": casou,
                "forma_em_composto": "-" in forma,
            })
            break  # uma etiqueta por posição (padrão mais longo já ordenado)
    return achados


def procura_near(tokens, no_idx, campo_compilado, n, limites, mesma_frase=True):
    """Co-ocorrências dentro de NEAR/n de um nó fixo (legado / banda)."""
    achados = []
    for t in encontra_termos(tokens, campo_compilado):
        par = melhor_par_tokens(
            tokens, [no_idx], t["idx_termo"], t["n_palavras"],
            limites, mesma_frase=mesma_frase)
        if par is None:
            continue
        i_no, idx_te, dist, lado = par
        if dist > n:
            continue
        achados.append({
            "termo_tipo": t["termo_tipo"],
            "termo_forma": t["termo_forma"],
            "distancia": dist,
            "lado": lado,
            "idx_termo": idx_te,
            "n_palavras": t["n_palavras"],
            "forma_em_composto": t["forma_em_composto"],
            "idx_no": i_no,
        })
    return achados


def emparelha_contexto(tokens, nos_validos, campo_compilado, near: int,
                       limites, mesma_frase: bool = True):
    """Emparelha cada termo com o nó mais próximo na mesma frase.

    Uma linha por (termo_tipo, idx_termo): o par de menor distância;
    empate → o mais à esquerda. Pares que atravessam fronteira de frase
    são descartados (contam em excluidos['fronteira_frase']).
    """
    idxs_no = indices_no(tokens, nos_validos)
    excluidos = {"fronteira_frase": 0}
    if not idxs_no:
        return [], excluidos, 0

    linhas = []
    # Uma ocorrência de termo = uma candidatura; retenção do melhor nó.
    por_chave = {}  # (termo_tipo, idx_termo) -> dict
    for t in encontra_termos(tokens, campo_compilado):
        par = melhor_par_tokens(
            tokens, idxs_no, t["idx_termo"], t["n_palavras"],
            limites, mesma_frase=mesma_frase)
        if par is None:
            excluidos["fronteira_frase"] += 1
            continue
        i_no, idx_te, dist, lado = par
        if dist > near:
            continue
        chave = (t["termo_tipo"], idx_te)
        cand = {
            "termo_tipo": t["termo_tipo"],
            "termo_forma": t["termo_forma"],
            "matched_form": t["termo_forma"],
            "canonical_term": t["termo_tipo"],
            "distancia": dist,
            "lado": lado,
            "idx_termo": idx_te,
            "idx_no": i_no,
            "n_palavras": t["n_palavras"],
            "forma_em_composto": t["forma_em_composto"],
            "n_nos_janela": len(idxs_no),
            "off_no": tokens[i_no][1],
            "off_termo": tokens[idx_te][1],
            "no": tokens[i_no][0],
        }
        ant = por_chave.get(chave)
        if ant is None or (dist, min(i_no, idx_te)) < (
                ant["distancia"], min(ant["idx_no"], ant["idx_termo"])):
            por_chave[chave] = cand

    # Colapsar várias ocorrências do mesmo tipo na janela: uma linha —
    # a do par globalmente mais próximo (empate → esquerda).
    por_tipo: dict[str, dict] = {}
    for cand in por_chave.values():
        etq = cand["termo_tipo"]
        ant = por_tipo.get(etq)
        chave_ord = (cand["distancia"], min(cand["idx_no"], cand["idx_termo"]))
        if ant is None or chave_ord < (
                ant["distancia"], min(ant["idx_no"], ant["idx_termo"])):
            por_tipo[etq] = cand

    return list(por_tipo.values()), excluidos, len(idxs_no)


def recalcular_distancia_lado(contexto: str, no_forma: str, matched_form: str,
                              *, mesma_frase: bool = True,
                              nos_validos: set[str] | None = None):
    """Recalcula distancia/lado a partir do contexto (teste de invariante)."""
    ctx = normaliza(contexto)
    toks = tokeniza(ctx)
    if nos_validos is None:
        no_l = no_forma.lower()
        nos_validos = {no_l}
        for conj in NOS.values():
            if no_l in conj:
                nos_validos = set(conj)
                break
    idxs_no = indices_no(toks, nos_validos)
    partes = matched_form.lower().split()
    L = len(partes)
    idxs_te = [j for j in range(len(toks) - L + 1)
               if [toks[k][0] for k in range(j, j + L)] == partes]
    if not idxs_no or not idxs_te:
        return None
    limites = fronteiras_frase(ctx) if mesma_frase else []
    melhor = None
    for j in idxs_te:
        par = melhor_par_tokens(toks, idxs_no, j, L, limites, mesma_frase)
        if par is None:
            continue
        i_no, idx_te, dist, lado = par
        chave = (dist, min(i_no, idx_te))
        if melhor is None or chave < melhor[0]:
            melhor = (chave, dist, lado)
    if melhor is None:
        return None
    return {"distancia": melhor[1], "lado": melhor[2]}


# ---------------------------------------------------------------------------
# 3b. CONSULTA BOOLEANA SOBRE OS TERMOS ENCONTRADOS NA MESMA JANELA
# ---------------------------------------------------------------------------
#
# Gramática:
#     consulta := termo ( ('AND'|'OR') termo )*
#     termo    := 'NOT'? ( '(' consulta ')' | ETIQUETA )
#
# ETIQUETA é uma chave de CAMPO. AND tem precedência sobre OR.
# Exemplos:
#     "uniform AND NOT varied"
#     "(uniform OR constant OR consistent) AND NOT homogeneous"
#     "homogeneous AND uniform"

class _Consulta:
    def __init__(self, expr: str, etiquetas: set[str]):
        self.toks = re.findall(r"\(|\)|[^\s()]+", expr)
        self.i = 0
        self.etiquetas = etiquetas
        self.arvore = self._ou()
        if self.i < len(self.toks):
            raise ValueError(f"consulta mal formada junto de {self.toks[self.i]!r}")

    def _olha(self):
        return self.toks[self.i] if self.i < len(self.toks) else None

    def _ou(self):
        no = self._e()
        while (self._olha() or "").upper() == "OR":
            self.i += 1
            no = ("or", no, self._e())
        return no

    def _e(self):
        no = self._unario()
        while (self._olha() or "").upper() == "AND":
            self.i += 1
            no = ("and", no, self._unario())
        return no

    def _unario(self):
        t = self._olha()
        if t is None:
            raise ValueError("consulta incompleta")
        if t.upper() == "NOT":
            self.i += 1
            return ("not", self._unario())
        if t == "(":
            self.i += 1
            no = self._ou()
            if self._olha() != ")":
                raise ValueError("parêntese por fechar")
            self.i += 1
            return no
        self.i += 1
        if t not in self.etiquetas:
            raise ValueError(f"etiqueta desconhecida: {t!r}. "
                             f"Disponíveis: {sorted(self.etiquetas)}")
        return ("lit", t)

    def avalia(self, presentes: set[str]) -> bool:
        return self._av(self.arvore, presentes)

    def _av(self, no, p):
        op = no[0]
        if op == "lit":
            return no[1] in p
        if op == "not":
            return not self._av(no[1], p)
        if op == "and":
            return self._av(no[1], p) and self._av(no[2], p)
        return self._av(no[1], p) or self._av(no[2], p)


def anota_polaridade_linear(tokens, idx_termo, texto: str | None = None):
    """Graduação e modalidade por proximidade; negação sem «no.» de opus.

    Graduação: 3 tokens à esquerda. Modalidade: ±4 tokens (ex.: «texture can»).
    """
    jan_idx = range(max(0, idx_termo - 3), idx_termo)
    jan = [tokens[k][0] for k in jan_idx]
    graduado = any(w in GRADUACAO for w in jan)
    jan_mod = range(max(0, idx_termo - 4),
                    min(len(tokens), idx_termo + 5))
    modalizado = any(tokens[k][0] in MODALIDADE for k in jan_mod)

    negado = None
    for k in jan_idx:
        w = tokens[k][0]
        if w in {"no", "nos"}:
            # «no. 1» / «nos. 3-4»: não é negação
            fim = tokens[k][1] + len(w)
            if texto is not None:
                seq = texto[fim:fim + 4]
                if re.match(r"\.?\s*\d", seq):
                    continue
            continue  # sem texto, não afirmar negação por «no»
        if w in NEGACAO:
            negado = True
            break
    if negado is None:
        # só marcar False se virmos a janela e não houver negador claro
        negado = False if not any(
            tokens[k][0] in (NEGACAO - {"absence", "lack", "devoid", "lacking",
                                        "without"})
            for k in jan_idx) else True

    return negado, graduado, modalizado


def anota_sintaxe(tokens, no_idx, idx_termo, texto: str | None = None):
    """Heurística de função sintáctica (+ polaridade linear).

    Devolve (negado, graduado, modalizado, relacao).
    Mantém compatibilidade: os primeiros dois slots históricos eram
    (negado, modalizado≈graduação); o 3.º era a relação.
    """
    negado, graduado, modalizado = anota_polaridade_linear(
        tokens, idx_termo, texto)

    a, b = sorted((no_idx, idx_termo))
    entre = [tokens[k][0] for k in range(a + 1, b)]
    predicativa = any(w in COPULAS for w in entre)

    if predicativa:
        relacao = "predicativa"
    elif idx_termo == no_idx - 1 or idx_termo == no_idx + 1:
        relacao = "atributiva"
    else:
        relacao = "indeterminada"

    return negado, graduado, modalizado, relacao


# ---------------------------------------------------------------------------
# 4. ESTATÍSTICA DE REFERÊNCIA
# ---------------------------------------------------------------------------

def shannon(contagens) -> float:
    total = sum(contagens)
    if total == 0:
        return float("nan")
    p = np.array([c / total for c in contagens if c > 0])
    return float(-(p * np.log(p)).sum())


def pielou(contagens) -> float:
    S = sum(1 for c in contagens if c > 0)
    return float(shannon(contagens) / math.log(S)) if S > 1 else float("nan")


def simpson_inverso(contagens) -> float:
    total = sum(contagens)
    if total == 0:
        return float("nan")
    p = np.array([c / total for c in contagens if c > 0])
    return float(1.0 / (p ** 2).sum())


def benjamini_hochberg(pvals) -> np.ndarray:
    """Correcção BH; devolve p ajustados na ordem de entrada."""
    p = np.asarray(pvals, dtype=float)
    m = len(p)
    ordem = np.argsort(p)
    ajust = np.empty(m, dtype=float)
    anterior = 1.0
    for rank in range(m - 1, -1, -1):
        i = ordem[rank]
        val = min(anterior, p[i] * m / (rank + 1))
        ajust[i] = val
        anterior = val
    return np.clip(ajust, 0, 1)


def _token_em(doc, offset: int):
    """Token cujo intervalo de caracteres contém o offset dado."""
    for t in doc:
        if t.idx <= offset < t.idx + len(t.text):
            return t
    return None


def _resultado_rel(rel: str, governante: str, percurso: str,
                   orientacao: str = "", *,
                   nucleo_prop: str = "", revisao: str = "",
                   matched: str = "") -> dict:
    nuclear = rel in RELACOES_NUCLEARES
    # R5: governante nunca é o próprio termo
    gov = governante
    if matched and gov and gov.lower() == matched.lower():
        gov = nucleo_prop or ""
    if not orientacao:
        if nuclear and rel in ("nominal_composto", "nominal_genitiva",
                               "adverbial"):
            orientacao = "no_sobre_termo"
        elif nuclear:
            orientacao = "termo_sobre_no"
        else:
            orientacao = "termo_sobre_outro"
    return {
        "relacao_sintactica": rel,
        "orientacao": orientacao,
        "governante": gov,
        "percurso_dep": percurso,
        "nuclear": nuclear,
        "motivo_exclusao": "" if nuclear else (rel or "indeterminada"),
        "nucleo_da_propriedade": nucleo_prop or ("" if not nuclear else ""),
        "revisao_sugerida": revisao,
    }


def _mesmo_token(a, b) -> bool:
    """Comparação estável entre tokens spaCy (evitar `is`, frágil)."""
    return a is not None and b is not None and a.i == b.i and a.doc is b.doc


def _no_no_subtree(tok, t_no) -> bool:
    return _mesmo_token(tok, t_no) or any(_mesmo_token(x, t_no)
                                          for x in tok.subtree)


def _gov_efectivo(tok) -> str:
    """Núcleo sintáctico do termo (head após conj), nunca o próprio texto."""
    base = tok
    while base.dep_ == "conj" and base.head.i != base.i:
        base = base.head
    h = base.head
    if h.i == base.i:
        return ""
    return h.text


def _tem_complemento_genitivo(termo, t_no) -> bool:
    """True se T rege prep cujo objecto é N (qualquer papel de T)."""
    base = termo
    while base.dep_ == "conj" and base.head.i != base.i:
        base = base.head
    for node in (termo, base):
        for child in node.children:
            if child.dep_ == "prep" and child.text.lower() in {
                    "of", "de", "in", "on"}:
                for gc in child.children:
                    if gc.dep_ in ("pobj", "nmod") and _no_no_subtree(
                            gc, t_no):
                        return True
    # N sob prep regida por T
    if t_no.dep_ in ("pobj", "nmod"):
        prep = t_no.head
        regente = prep.head if prep.dep_ == "prep" else prep
        if _mesmo_token(regente, termo) or _mesmo_token(regente, base):
            return True
    return False


def _amod_coordenado_do_no(t_te, t_no) -> bool:
    """T é adjectivo coordenado com outro amod cujo núcleo é N."""
    if t_no.pos_ not in ("NOUN", "PROPN"):
        return False
    # irmãos amod de N
    for am in t_no.children:
        if am.dep_ != "amod":
            continue
        if _mesmo_token(am, t_te):
            return True
        # cadeia de conjunção
        conj = list(am.conjuncts) if hasattr(am, "conjuncts") else []
        if any(_mesmo_token(c, t_te) for c in conj):
            return True
        # T.conj aponta para am, ou vice-versa
        b = t_te
        while b.dep_ == "conj" and b.head.i != b.i:
            if _mesmo_token(b.head, am):
                return True
            b = b.head
    # heurística linear: ADJ ... and/or ADJ NOUN=N
    if t_te.pos_ == "ADJ" and t_te.i < t_no.i:
        entre = list(t_te.doc[t_te.i + 1:t_no.i])
        if entre and all(
                t.text.lower() in {"and", "or", "but", ",", "the", "a", "an"}
                or t.dep_ in ("cc", "punct", "det", "amod", "conj")
                for t in entre):
            return True
    return False


def relacao_dependencia(doc, off_no: int, off_termo: int) -> dict:
    """Classifica a relação sintáctica (taxonomia nuclear / não nuclear)."""
    t_no, t_te = _token_em(doc, off_no), _token_em(doc, off_termo)
    if t_no is None or t_te is None:
        return _resultado_rel("indeterminada", "", "", "")

    matched = t_te.text
    percurso = f"{t_te.text}/{t_te.dep_}->{t_te.head.text}"
    era_conj = t_te.dep_ == "conj"
    base = t_te
    while base.dep_ == "conj" and base.head.i != base.i:
        base = base.head
    gov0 = _gov_efectivo(t_te)

    # R4.2 — genitiva ANTES de qualquer teste de governante directo
    if _tem_complemento_genitivo(t_te, t_no):
        rev = "genitiva_por_complemento" if t_te.dep_ not in (
            "pobj", "nmod", "") else ""
        return _resultado_rel(
            "nominal_genitiva", gov0, percurso, "no_sobre_termo",
            nucleo_prop=t_no.text, revisao=rev, matched=matched)

    # --- adverbial: «texturally uniform» ----------------------------------
    if t_no.dep_ == "advmod" and _mesmo_token(t_no.head, t_te):
        return _resultado_rel(
            "adverbial", gov0 or t_te.head.text, percurso, "no_sobre_termo",
            nucleo_prop=t_no.text, matched=matched)

    # --- adverbial de grau / verbal ---------------------------------------
    if base.dep_ == "advmod" and base.head.pos_ in ("ADJ", "ADV"):
        if not _mesmo_token(base.head, t_no):
            return _resultado_rel(
                "adverbial_de_grau", base.head.text, percurso,
                "termo_sobre_outro", nucleo_prop=base.head.text,
                matched=matched)
    if base.dep_ == "advmod" and base.head.pos_ == "VERB":
        return _resultado_rel(
            "adverbial_verbal", base.head.text, percurso,
            "termo_sobre_outro", nucleo_prop=base.head.text, matched=matched)

    # --- predicação secundária --------------------------------------------
    if base.dep_ in ("oprd", "xcomp"):
        pred = base.head
        objs = [c for c in pred.children
                if c.dep_ in ("dobj", "obj", "oprd", "attr")]
        if any(_no_no_subtree(o, t_no) for o in objs) or any(
                _no_no_subtree(s, t_no) for s in pred.children
                if s.dep_ in ("nsubj", "nsubjpass")):
            return _resultado_rel(
                "predicativa_secundaria", gov0 or pred.text, percurso,
                "termo_sobre_no", nucleo_prop=t_no.text, matched=matched)

    # --- predicação: acomp/attr — reclassificar após conj -----------------
    if base.dep_ in ("acomp", "attr"):
        pred = base.head
        sujeitos = [c for c in pred.children
                    if c.dep_ in ("nsubj", "nsubjpass")]
        if sujeitos:
            s = sujeitos[0]
            if _no_no_subtree(s, t_no):
                return _resultado_rel(
                    "predicativa", gov0 or pred.text, percurso,
                    "termo_sobre_no", nucleo_prop=t_no.text, matched=matched)
            # attr de factor ... is uniformity of texture — já coberto por genitiva
            return _resultado_rel(
                "incidental", s.text, percurso, "termo_sobre_outro",
                nucleo_prop=s.text, matched=matched)

    # --- atributiva directa (também após normalização de conj) ------------
    if base.dep_ in ("amod", "appos") and _mesmo_token(base.head, t_no):
        rev = "atributiva_via_conj" if era_conj else ""
        return _resultado_rel(
            "atributiva", gov0 or t_no.text, percurso, "termo_sobre_no",
            nucleo_prop=t_no.text, revisao=rev, matched=matched)

    # R4.3 — modificação partilhada / coordenação adjectival
    if _amod_coordenado_do_no(t_te, t_no):
        return _resultado_rel(
            "atributiva", gov0 or t_no.text, percurso, "termo_sobre_no",
            nucleo_prop=t_no.text, revisao="atributiva_coordenada",
            matched=matched)

    # --- nominal composto: «textural diversity» ---------------------------
    if t_no.dep_ in ("amod", "compound") and _mesmo_token(t_no.head, t_te):
        return _resultado_rel(
            "nominal_composto", gov0 or t_te.head.text, percurso,
            "no_sobre_termo", nucleo_prop=t_no.text, matched=matched)

    if base.dep_ == "compound" and _mesmo_token(base.head, t_no):
        return _resultado_rel(
            "atributiva", gov0 or t_no.text, percurso, "termo_sobre_no",
            nucleo_prop=t_no.text, matched=matched)

    # --- coordenação entre constituintes distintos ------------------------
    if era_conj:
        gov = base.head if base.head.i != base.i else base
        if not _mesmo_token(gov, t_no) and not _no_no_subtree(gov, t_no):
            return _resultado_rel(
                "coordenada", gov.text, percurso, "termo_sobre_outro",
                nucleo_prop=gov.text, matched=matched)

    gov = base.head
    return _resultado_rel(
        "incidental", gov.text if gov is not None else "",
        percurso, "termo_sobre_outro",
        nucleo_prop=gov.text if gov is not None else "", matched=matched)


def _escopo_negacao(doc, off_termo: int, off_no: int) -> str:
    """Negação: 'directo' | 'indirecto' | 'nao'."""
    t_te = _token_em(doc, off_termo)
    if t_te is None:
        return "nao"
    base = t_te
    while base.dep_ == "conj" and base.head.i != base.i:
        base = base.head
    # predicado em que o termo participa
    pred = base.head if base.dep_ in (
        "acomp", "attr", "oprd", "xcomp", "amod", "advmod") else t_te.head

    negadores = []
    for tok in doc:
        w = tok.text.lower().replace("'", "'")
        if tok.dep_ == "neg" or w in {"not", "never", "n't"} or \
                tok.lemma_.lower() in {"not", "never"}:
            negadores.append(tok)

    if not negadores:
        return "nao"

    for neg in negadores:
        # neg domina o predicado do termo?
        cabeças = {neg.head}
        if pred in list(neg.head.subtree) or _mesmo_token(neg.head, pred) \
                or _mesmo_token(neg.head, t_te) or t_te in list(neg.head.subtree):
            # directo: o head do neg é o mesmo predicado do termo
            if _mesmo_token(neg.head, pred) or _mesmo_token(neg.head, base.head):
                # e o termo é complemento/modificador desse predicado
                if base.dep_ in ("acomp", "attr", "oprd", "xcomp") or \
                        (base.dep_ == "amod" and _mesmo_token(base.head,
                                                              _token_em(doc, off_no) or base.head)):
                    # amod sob N: neg no verbo superior = nao/indirecto
                    if base.dep_ == "amod":
                        # «is not uniform» (acomp) vs «does not create uniform texture»
                        if neg.head.pos_ == "AUX" or neg.head.lemma_ in {
                                "be", "remain", "become", "seem", "appear"}:
                            if _mesmo_token(base.head, _token_em(doc, off_no)):
                                # texture is not uniform — termo é acomp, não amod
                                pass
                        if not _mesmo_token(neg.head, base.head):
                            # neg noutro verbo: «does not ... uniform texture»
                            return "indirecto"
                    else:
                        return "directo"
            # neg em cláusula superior que contém o NP
            if t_te in list(neg.head.subtree):
                # se há verbo interveniente entre neg.head e o termo
                if neg.head.pos_ == "VERB" and not _mesmo_token(neg.head, pred):
                    return "indirecto"
                if _mesmo_token(neg.head, pred):
                    return "directo"
    # negação na frase mas fora do predicado do termo
    for neg in negadores:
        if t_te in list(neg.head.subtree):
            return "nao"
    return "nao"


def anota_com_spacy(res: pd.DataFrame, modelo: str, *,
                    obrigatorio: bool = True) -> pd.DataFrame:
    """Classifica cada linha pela árvore de dependências (spaCy)."""
    try:
        import spacy
    except ImportError:
        msg = ("spaCy nao instalado. Execute: pip install 'spacy>=3.7' "
               f"&& python -m spacy download {modelo}")
        if obrigatorio:
            raise SystemExit(msg) from None
        print("AVISO: " + msg + " - a usar heuristica.", file=sys.stderr)
        return res
    try:
        nlp = spacy.load(modelo, disable=["ner", "lemmatizer", "textcat"])
    except OSError:
        msg = (f"modelo spaCy '{modelo}' indisponivel. Execute:\n"
               f"  python -m spacy download {modelo}")
        if obrigatorio:
            raise SystemExit(msg) from None
        print("AVISO: " + msg + " - a usar heuristica.", file=sys.stderr)
        return res

    contextos = res["contexto"].astype(str).unique().tolist()
    print(f"      a analisar sintaxe de {len(contextos)} contextos unicos ...",
          flush=True)
    docs = {c: d for c, d in zip(contextos, nlp.pipe(contextos, batch_size=64))}

    keys = ("relacao_sintactica", "orientacao", "governante", "percurso_dep",
            "nuclear", "motivo_exclusao", "nucleo_da_propriedade",
            "revisao_sugerida", "negado", "modalizado")
    cols = {k: [] for k in keys}
    for t in res.itertuples(index=False):
        doc = docs[str(t.contexto)]
        r = relacao_dependencia(doc, int(t.off_no), int(t.off_termo))
        for k in ("relacao_sintactica", "orientacao", "governante",
                  "percurso_dep", "nuclear", "motivo_exclusao",
                  "nucleo_da_propriedade", "revisao_sugerida"):
            cols[k].append(r[k])
        cols["negado"].append(_escopo_negacao(
            doc, int(t.off_termo), int(t.off_no)))
        # R8: modal no escopo local do termo (±5 tokens ou ancestral)
        t_te = _token_em(doc, int(t.off_termo))
        mod = False
        if t_te is not None:
            for tok in doc:
                w = tok.text.lower()
                if w in MODALIDADE or tok.lemma_.lower() in MODALIDADE:
                    if abs(tok.i - t_te.i) <= 5:
                        mod = True
                        break
                    if t_te in list(tok.subtree) or tok in list(t_te.ancestors):
                        mod = True
                        break
        cols["modalizado"].append(mod)

    res = res.copy()
    for k, vals in cols.items():
        res[k] = vals
    res["fonte_classificacao"] = "dependencias"
    # R10: nao duplicar relacao / caminho_dep / atribuicao
    for drop in ("relacao", "caminho_dep", "atribuicao", "caminho"):
        if drop in res.columns:
            res = res.drop(columns=[drop])
    return res


def anota_com_heuristica(res: pd.DataFrame) -> pd.DataFrame:
    """Preenche taxonomia reduzida sem spaCy; fonte_classificacao=heuristica."""
    res = res.copy()
    rels, nucs, mots, oris, govs = [], [], [], [], []
    grads, mods, negs = [], [], []
    for t in res.itertuples(index=False):
        toks = tokeniza(str(t.contexto))
        # localizar índices por offset
        i_no = next((i for i, (_, o) in enumerate(toks)
                     if o == int(t.off_no)), None)
        i_te = next((i for i, (_, o) in enumerate(toks)
                     if o == int(t.off_termo)), None)
        if i_no is None or i_te is None:
            rels.append("indeterminada"); nucs.append(False)
            mots.append("indeterminada"); oris.append(""); govs.append("")
            grads.append(False); mods.append(False); negs.append(None)
            continue
        neg, grad, mod, rel = anota_sintaxe(
            toks, i_no, i_te, str(t.contexto))
        nuclear = rel in RELACOES_NUCLEARES
        rels.append(rel); nucs.append(nuclear)
        mots.append("" if nuclear else rel)
        oris.append("termo_sobre_no" if nuclear else "")
        govs.append(toks[i_no][0] if nuclear else "")
        grads.append(grad); mods.append(mod); negs.append(neg)
    res["relacao_sintactica"] = rels
    res["relacao"] = rels
    res["nuclear"] = nucs
    res["motivo_exclusao"] = mots
    res["orientacao"] = oris
    res["governante"] = govs
    res["percurso_dep"] = ""
    res["caminho_dep"] = ""
    res["fonte_classificacao"] = "heuristica"
    res["graduado"] = grads
    res["modalizado"] = mods
    res["negado"] = negs
    res["atribuicao"] = np.where(res["nuclear"], "genuína", "incidental")
    return res


def conta_tokens(tokens, no_idx, limites, lo, hi, mesma_frase=True) -> int:
    """Nº de tokens cuja distância ao nó está em ]lo, hi], na mesma frase."""
    frase_no = indice_frase(tokens[no_idx][1], limites)
    n = 0
    for k in range(max(0, no_idx - hi), min(len(tokens), no_idx + hi + 1)):
        d = abs(k - no_idx)
        if d <= lo or d > hi:
            continue
        if mesma_frase and indice_frase(tokens[k][1], limites) != frase_no:
            continue
        n += 1
    return n


def medidas_associacao(o11, o12, r1, r2, n_janelas):
    """Medidas de associação sobre a tabela 2x2 'janela vs. banda de referência'.

        o11 = ocorrências do termo dentro de NEAR/x
        o12 = ocorrências do termo na banda de referência (mais distante)
        r1  = total de tokens dentro de NEAR/x
        r2  = total de tokens na banda de referência

    O termo de comparação é a banda distante dos MESMOS contextos, e não o
    corpus integral — o qual não está disponível na matriz KWIC. Trata-se,
    portanto, de uma associação posicional: mede se o termo se concentra
    junto do nó mais do que na sua vizinhança alargada. É esta a leitura
    que deve constar do texto.
    """
    o21, o22 = r1 - o11, r2 - o12
    n = r1 + r2
    if n == 0 or (o11 + o12) == 0:
        return {}
    e11 = r1 * (o11 + o12) / n
    e12 = r2 * (o11 + o12) / n
    e21 = r1 * (o21 + o22) / n
    e22 = r2 * (o21 + o22) / n

    def _g(o, e):
        return 2 * o * math.log(o / e) if o > 0 and e > 0 else 0.0

    ll = _g(o11, e11) + _g(o12, e12) + _g(o21, e21) + _g(o22, e22)
    ll = ll if o11 >= e11 else -ll          # sinal: + atracção, - repulsão

    mi = math.log2(o11 / e11) if o11 > 0 and e11 > 0 else float("nan")
    mi3 = math.log2(o11 ** 3 / e11) if o11 > 0 and e11 > 0 else float("nan")
    tscore = (o11 - e11) / math.sqrt(o11) if o11 > 0 else float("nan")
    zscore = (o11 - e11) / math.sqrt(e11) if e11 > 0 else float("nan")
    dice = 2 * o11 / (n_janelas + o11 + o12) if (n_janelas + o11 + o12) else float("nan")
    logdice = 14 + math.log2(dice) if dice > 0 else float("nan")
    dp_termo = (o11 / r1 - o12 / r2) if r1 and r2 else float("nan")

    # razão de possibilidades com correcção de Haldane e IC de Woolf
    a, b, c, d = o11 + 0.5, o12 + 0.5, o21 + 0.5, o22 + 0.5
    orr = (a * d) / (b * c)
    se = math.sqrt(1 / a + 1 / b + 1 / c + 1 / d)
    ic_inf, ic_sup = math.exp(math.log(orr) - 1.96 * se), math.exp(math.log(orr) + 1.96 * se)

    p_fisher = stats.fisher_exact([[o11, o12], [o21, o22]]).pvalue

    return {
        "O11_janela": o11, "E11_esperado": round(e11, 3),
        "O12_banda_ref": o12,
        "log_likelihood_G2": round(ll, 3),
        "MI": round(mi, 3), "MI3": round(mi3, 3),
        "t_score": round(tscore, 3), "z_score": round(zscore, 3),
        "logDice": round(logdice, 3),
        "DeltaP": round(dp_termo, 5),
        "razao_possib": round(orr, 3),
        "IC95_inf": round(ic_inf, 3), "IC95_sup": round(ic_sup, 3),
        "p_fisher": p_fisher,
    }


def dispersao_gries_dp(cont_por_parte: dict, tam_por_parte: dict) -> float:
    """Desvio de proporções (DP) de Gries. 0 = uniforme; 1 = concentrado."""
    total = sum(cont_por_parte.values())
    tam_tot = sum(tam_por_parte.values())
    if total == 0 or tam_tot == 0:
        return float("nan")
    s = 0.0
    for parte, tam in tam_por_parte.items():
        esperado = tam / tam_tot
        observado = cont_por_parte.get(parte, 0) / total
        s += abs(observado - esperado)
    return round(s / 2, 4)


def juilland_d(cont_por_parte: dict, n_partes: int) -> float:
    """D de Juilland. 1 = distribuição uniforme pelas partes."""
    v = list(cont_por_parte.values()) + [0] * (n_partes - len(cont_por_parte))
    m = float(np.mean(v))
    if m == 0 or n_partes < 2:
        return float("nan")
    cv = float(np.std(v, ddof=0)) / m
    return round(1 - cv / math.sqrt(n_partes - 1), 4)


def bootstrap_proporcao(v, n_rep=5000, semente=20260724):
    """IC95% percentílico para uma proporção binária, por reamostragem."""
    rng = np.random.default_rng(semente)
    v = np.asarray(v, dtype=float)
    if v.size == 0:
        return (float("nan"), float("nan"))
    amostras = rng.choice(v, size=(n_rep, v.size), replace=True).mean(axis=1)
    return (round(float(np.percentile(amostras, 2.5)), 4),
            round(float(np.percentile(amostras, 97.5)), 4))


POLO_ESTABILIDADE = {
    "uniform", "invariable", "unvarying", "immutable", "unchanging",
    "constant", "consistent", "regular", "stable", "steady", "sustained",
    "static", "monotonous", "homogeneous",
}
POLO_VARIABILIDADE = {
    "varied", "varying", "changing", "irregular", "unequal", "diverse",
    "mutable", "multiform", "heterogeneous",
}


def polaridade(tipo: str, negado: bool | None = False,
               *, inverter_negada: bool = False) -> str | None:
    """Polaridade de base; inversão sob negação só se inverter_negada=True."""
    if tipo in POLO_ESTABILIDADE:
        base = "estabilidade"
    elif tipo in POLO_VARIABILIDADE:
        base = "variabilidade"
    else:
        return None
    if inverter_negada and negado:
        base = "variabilidade" if base == "estabilidade" else "estabilidade"
    return base


def eixo_semantico(tipo: str) -> str:
    """Eixo: homogeneidade_sincronica | invariancia_diacronica | ambos."""
    sinc = {"uniform", "homogeneous", "heterogeneous", "diverse",
            "varied", "unequal", "irregular", "multiform", "mutable"}
    diac = {"static", "invariable", "unvarying", "immutable", "unchanging",
            "constant", "varying", "changing", "stable", "steady", "sustained",
            "monotonous", "consistent", "regular"}
    if tipo in sinc and tipo in diac:
        return "ambos"
    if tipo in sinc:
        return "homogeneidade_sincronica"
    if tipo in diac:
        return "invariancia_diacronica"
    return "ambos"


# ---------------------------------------------------------------------------
# 5. GRÁFICOS
# ---------------------------------------------------------------------------

def grafico_frequencias(df, destino: Path):
    col = "doc_id" if "doc_id" in df.columns else "caminho"
    cont = df.groupby("termo_tipo")[col].nunique().sort_values(ascending=False)
    fig, ax = plt.subplots(figsize=(9, max(4, 0.32 * len(cont))))
    ax.barh(cont.index[::-1], cont.values[::-1], color="#4a5a6a")
    ax.set_xlabel("Obras únicas com co-ocorrência")
    ax.set_ylabel("")
    ax.set_title("Dispersão do campo lexical (obras únicas)")
    ax.grid(axis="x", linewidth=0.4, alpha=0.5)
    fig.tight_layout()
    fig.savefig(destino, dpi=150)
    plt.close(fig)


def grafico_distancias(df, destino: Path):
    fig, ax = plt.subplots(figsize=(7, 4))
    for lado, cor in (("esq", "#4a5a6a"), ("dir", "#a8763e")):
        sub = df[df["lado"] == lado]["distancia"]
        ax.hist(sub, bins=np.arange(0.5, df["distancia"].max() + 1.5),
                alpha=0.65, label=f"{lado}erda" if lado == "esq" else "direita",
                color=cor)
    ax.set_xlabel("Distância em tokens ao nó")
    ax.set_ylabel("Co-ocorrências")
    ax.set_title("Distribuição da distância por lado")
    ax.legend()
    ax.grid(axis="y", linewidth=0.4, alpha=0.5)
    fig.tight_layout()
    fig.savefig(destino, dpi=150)
    plt.close(fig)


def grafico_polaridade(df, destino: Path):
    sub = df.replace({"polaridade": {"": np.nan}}).dropna(subset=["polaridade"])
    col = ("relacao_sintactica" if "relacao_sintactica" in sub.columns
           else "relacao")
    tab = pd.crosstab(sub[col], sub["polaridade"])
    fig, ax = plt.subplots(figsize=(7, 4))
    tab.plot(kind="bar", stacked=True, ax=ax,
             color=["#4a5a6a", "#a8763e"])
    ax.set_xlabel("Relação sintáctica")
    ax.set_ylabel("Co-ocorrências")
    ax.set_title("Polaridade por relação sintáctica")
    ax.grid(axis="y", linewidth=0.4, alpha=0.5)
    fig.tight_layout()
    fig.savefig(destino, dpi=150)
    plt.close(fig)


# ---------------------------------------------------------------------------
# 6. FLUXO PRINCIPAL
# ---------------------------------------------------------------------------

def _configurar_consola() -> None:
    """Evita UnicodeEncodeError na consola Windows (cp1252)."""
    for stream in (sys.stdout, sys.stderr):
        try:
            stream.reconfigure(encoding="utf-8", errors="replace")
        except Exception:
            pass


def _partilha_ngrama(a: list[str], b: list[str], n: int = 8) -> bool:
    """True se a e b partilham pelo menos n tokens contiguos."""
    if len(a) < n or len(b) < n:
        return False
    grams = {" ".join(a[i:i + n]) for i in range(len(a) - n + 1)}
    return any(" ".join(b[j:j + n]) in grams for j in range(len(b) - n + 1))


SCHEMA_NEAR = 2


def occurrence_id_de(doc_id: str, source_matrix_row: int) -> str:
    """Identificador imutável da ocorrência mestra (= linha da matriz)."""
    return f"{doc_id}::ROW_{int(source_matrix_row)}"


def hit_key_de(occurrence_id: str, canonical_term: str, matched_form: str,
               off_no: int, off_termo: int) -> str:
    """Chave de hit distinta dentro de uma ocorrência mestra."""
    return (
        f"{occurrence_id}|{canonical_term}|{matched_form}|"
        f"{int(off_no)}|{int(off_termo)}"
    )


def atribuir_match_ids(res: pd.DataFrame) -> pd.DataFrame:
    """Numera M001… por ocorrência e preenche hit_key."""
    if res.empty:
        res = res.copy()
        res["match_id"] = pd.Series(dtype=str)
        res["hit_key"] = pd.Series(dtype=str)
        return res
    out = res.copy()
    out["match_id"] = ""
    out["hit_key"] = ""
    cols_ord = [c for c in ("off_termo", "canonical_term", "matched_form",
                            "idx_termo") if c in out.columns]
    for _occ, grp in out.groupby("texture_occurrence_id", sort=False):
        ordem = grp.sort_values(cols_ord, kind="mergesort").index
        for n, i in enumerate(ordem, 1):
            mid = f"M{n:03d}"
            out.at[i, "match_id"] = mid
            out.at[i, "hit_key"] = hit_key_de(
                str(out.at[i, "texture_occurrence_id"]),
                str(out.at[i, "canonical_term"]),
                str(out.at[i, "matched_form"]),
                int(out.at[i, "off_no"]),
                int(out.at[i, "off_termo"]),
            )
    return out


def deduplicar_hits_exactos(res: pd.DataFrame) -> pd.DataFrame:
    """Remove cópias exactas do mesmo hit_key (mantém a primeira)."""
    if res.empty or "hit_key" not in res.columns:
        return res
    return res.drop_duplicates(subset=["hit_key"], keep="first").copy()


def _score_sobrevivente_janela(row) -> tuple:
    """Maior = melhor candidato a representar um grupo de janelas."""
    nuc = 1 if bool(row.get("nuclear")) else 0
    ctx = len(str(row.get("contexto") or ""))
    # preferir a linha de matriz mais antiga (menor número)
    smr = -int(row.get("source_matrix_row") or 10**12)
    return (nuc, ctx, smr)


def fundir_janelas_e_marcar_duplicados(res: pd.DataFrame,
                                       ngrama: int = 8) -> pd.DataFrame:
    """Funde janelas KWIC deslocadas do *mesmo* hit; assinala passagens.

    Regras:
    - Mesmo ``doc_id`` + mesmo ``(canonical_term, matched_form)`` +
      partilha de n-grama → um sobrevivente; restantes
      ``nuclear=False``, ``motivo_exclusao=janela_sobreposta``.
    - Mesmo ``doc_id`` + n-grama partilhado com termos *diferentes* →
      apenas ``candidato_duplicado`` / ``grupo_passagem_id`` (não excluir:
      podem ser hits legítimos na mesma passagem / ocorrência mestra).
    - Citações exactas sob ``doc_id`` distintos → ``citacao_repetida``.
    """
    if res.empty or "doc_id" not in res.columns:
        return res
    out = res.copy()
    if "candidato_duplicado" not in out.columns:
        out["candidato_duplicado"] = ""
    if "grupo_passagem_id" not in out.columns:
        out["grupo_passagem_id"] = ""
    if "n_janelas_fundidas" not in out.columns:
        out["n_janelas_fundidas"] = 1
    out["n_janelas_fundidas"] = out["n_janelas_fundidas"].fillna(1).astype(int)

    # --- citações repetidas (mesmo texto, doc_ids distintos) -------------
    grupos_ctx: dict[str, list] = defaultdict(list)
    for i, row in out.iterrows():
        toks_n = [w for w, _ in tokeniza(str(row["contexto"]))]
        if len(toks_n) >= 8:
            grupos_ctx[" ".join(toks_n)].append(i)
    for idxs in grupos_ctx.values():
        docs = {out.at[i, "doc_id"] for i in idxs}
        if len(docs) <= 1:
            continue
        # manter o primeiro; marcar restantes
        for i in idxs[1:]:
            out.at[i, "nuclear"] = False
            if not out.at[i, "motivo_exclusao"]:
                out.at[i, "motivo_exclusao"] = "citacao_repetida"
            prev = str(out.at[i, "candidato_duplicado"] or "")
            tag = "citacao_entre_doc_ids"
            out.at[i, "candidato_duplicado"] = (
                f"{prev}; {tag}" if prev else tag)

    # --- fusão / passagem no mesmo documento -----------------------------
    por_doc: dict = defaultdict(list)
    for i, row in out.iterrows():
        por_doc[row["doc_id"]].append(i)

    gid_pass = 0
    gid_jan = 0
    for idxs in por_doc.values():
        if len(idxs) < 2:
            continue
        assin = {
            i: [w for w, _ in tokeniza(str(out.at[i, "contexto"]))]
            for i in idxs
        }
        # Union-find leve por partilha de n-grama (passagem)
        parent = {i: i for i in idxs}

        def find(x):
            while parent[x] != x:
                parent[x] = parent[parent[x]]
                x = parent[x]
            return x

        def union(a, b):
            ra, rb = find(a), find(b)
            if ra != rb:
                parent[rb] = ra

        for a_i, a in enumerate(idxs):
            for b in idxs[a_i + 1:]:
                # Hits da mesma ocorrência mestra partilham o contexto por
                # definição — não são janelas KWIC deslocadas.
                if (str(out.at[a, "texture_occurrence_id"])
                        == str(out.at[b, "texture_occurrence_id"])):
                    continue
                if _partilha_ngrama(assin[a], assin[b], ngrama):
                    union(a, b)

        clusters: dict = defaultdict(list)
        for i in idxs:
            clusters[find(i)].append(i)

        for membros in clusters.values():
            if len(membros) < 2:
                continue
            gid_pass += 1
            gpass = f"P{gid_pass:04d}"
            for i in membros:
                out.at[i, "grupo_passagem_id"] = gpass
                prev = str(out.at[i, "candidato_duplicado"] or "")
                tag = f"passagem_sobreposta:{gpass}"
                if tag not in prev:
                    out.at[i, "candidato_duplicado"] = (
                        f"{prev}; {tag}" if prev else tag)

            # Dentro da passagem: fundir só hits lexicais idênticos
            por_termo: dict = defaultdict(list)
            for i in membros:
                chave = (
                    str(out.at[i, "canonical_term"]),
                    str(out.at[i, "matched_form"]),
                )
                por_termo[chave].append(i)
            for grupo in por_termo.values():
                if len(grupo) < 2:
                    continue
                gid_jan += 1
                gjan = f"J{gid_jan:04d}"
                ranked = sorted(
                    grupo,
                    key=lambda i: _score_sobrevivente_janela(out.loc[i]),
                    reverse=True,
                )
                keep = ranked[0]
                out.at[keep, "n_janelas_fundidas"] = len(ranked)
                tag_keep = f"janela_sobreposta:{gjan}"
                prev_k = str(out.at[keep, "candidato_duplicado"] or "")
                if tag_keep not in prev_k:
                    out.at[keep, "candidato_duplicado"] = (
                        f"{prev_k}; {tag_keep}" if prev_k else tag_keep)
                for i in ranked[1:]:
                    out.at[i, "nuclear"] = False
                    if not out.at[i, "motivo_exclusao"]:
                        out.at[i, "motivo_exclusao"] = "janela_sobreposta"
                    out.at[i, "n_janelas_fundidas"] = 1
                    prev = str(out.at[i, "candidato_duplicado"] or "")
                    tag = f"janela_sobreposta:{gjan}"
                    if tag not in prev:
                        out.at[i, "candidato_duplicado"] = (
                            f"{prev}; {tag}" if prev else tag)
    return out


def agregar_ocorrencias(res: pd.DataFrame) -> pd.DataFrame:
    """Uma linha por ``texture_occurrence_id`` (nível 1)."""
    if res.empty or "texture_occurrence_id" not in res.columns:
        return pd.DataFrame()
    linhas = []
    for occ_id, grp in res.groupby("texture_occurrence_id", sort=False):
        # contexto canónico: o mais longo (melhor exibição)
        ctxs = grp["contexto"].astype(str)
        i_ctx = ctxs.str.len().idxmax()
        termos = sorted({str(t) for t in grp["canonical_term"].tolist() if t})
        formas = sorted({str(t) for t in grp["matched_form"].tolist() if t})
        nucs = grp["nuclear"].map(
            lambda v: v is True or str(v).lower() in {"true", "1", "sim"})
        motivos = [str(m) for m in grp["motivo_exclusao"].tolist()
                   if m and str(m).strip() and str(m).lower() != "nan"]
        linhas.append({
            "texture_occurrence_id": occ_id,
            "source_matrix_row": int(grp["source_matrix_row"].iloc[0]),
            "doc_id": grp["doc_id"].iloc[0],
            "caminho_ficheiro": grp["caminho_ficheiro"].iloc[0],
            "url": grp["url"].iloc[0] if "url" in grp.columns else "",
            "no": grp["no"].iloc[0] if "no" in grp.columns else "",
            "matched_terms": "; ".join(termos),
            "matched_forms": "; ".join(formas),
            "n_matches": int(len(grp)),
            "n_matches_nucleares": int(nucs.sum()),
            "nuclear": bool(nucs.any()),
            "canonical_context": ctxs.loc[i_ctx],
            "grupo_passagem_id": (
                str(grp["grupo_passagem_id"].iloc[0])
                if "grupo_passagem_id" in grp.columns else ""),
            "match_ids": "; ".join(
                str(x) for x in grp.sort_values("match_id")["match_id"]),
            "motivos_exclusao": "; ".join(sorted(set(motivos))),
            "dominio": (
                grp["dominio"].iloc[0] if "dominio" in grp.columns else ""),
        })
    cols = [
        "texture_occurrence_id", "source_matrix_row", "doc_id",
        "caminho_ficheiro", "url", "no", "matched_terms", "matched_forms",
        "n_matches", "n_matches_nucleares", "nuclear", "canonical_context",
        "grupo_passagem_id", "match_ids", "motivos_exclusao", "dominio",
    ]
    return pd.DataFrame(linhas)[cols]


def reordenar_colunas_hits(res: pd.DataFrame) -> pd.DataFrame:
    """Coloca identificadores à frente sem perder colunas extra."""
    prioridade = [
        "source_matrix_row", "texture_occurrence_id", "match_id", "hit_key",
        "grupo_passagem_id", "candidato_duplicado",
        "no", "termo_tipo", "canonical_term", "query_pattern",
        "termo_forma", "matched_form", "n_palavras", "distancia", "lado",
        "negado", "graduado", "modalizado", "relacao_sintactica",
        "polaridade_base", "polaridade", "eixo",
        "censurado_esq", "censurado_dir",
        "idx_no", "idx_termo", "off_no", "off_termo",
        "n_nos_janela", "forma_em_composto",
        "caminho_ficheiro", "doc_id", "url", "contexto",
        "motivo_exclusao", "nuclear", "fonte_classificacao",
        "n_janelas_fundidas", "revisao_sugerida",
        "nucleo_da_propriedade", "orientacao", "governante", "percurso_dep",
        "dominio", "revisto_por_humano", "nota_revisao",
    ]
    frente = [c for c in prioridade if c in res.columns]
    resto = [c for c in res.columns if c not in frente]
    return res[frente + resto]


def main() -> int:
    _configurar_consola()
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--xlsx", required=True, type=Path)
    ap.add_argument("--folha", default="Neighbor Contexts")
    ap.add_argument("--near", type=int, default=4)
    ap.add_argument("--lingua", default="en", choices=list(NOS) + ["todas"])
    ap.add_argument("--limite", type=int, default=None,
                    help="processar apenas as N primeiras linhas (teste)")
    ap.add_argument("--sem-fronteira", action="store_true",
                    help="não aplicar exclusão por fronteira de frase")
    ap.add_argument("--termos", type=Path, default=None,
                    help="ficheiro de texto com o campo lexical, uma linha por "
                         "tipo: 'etiqueta = padrao1, padrao2, ...'. Substitui CAMPO.")
    ap.add_argument("--consulta", default=None,
                    help="expressão booleana sobre etiquetas, avaliada por janela. "
                         "Ex.: \"(uniform OR constant) AND NOT varied\"")
    ap.add_argument("--banda", type=int, default=12,
                    help="limite da banda de referência para as medidas de "
                         "associação (tokens); deve exceder --near")
    ap.add_argument("--col-no", type=int, default=6,
                    help="nº da coluna do nó (1 = primeira)")
    ap.add_argument("--col-ctx", type=int, default=15,
                    help="nº da coluna do contexto integral")
    ap.add_argument("--col-src", type=int, default=12,
                    help="nº da coluna da fonte/ficheiro")
    ap.add_argument("--com-cabecalho", action="store_true",
                    help="a folha tem linha de cabeçalho")
    ap.add_argument("--col-url", type=int, default=13,
                    help="nº da coluna com a hiperligação (0 = nenhuma)")
    ap.add_argument("--sintaxe", default="spacy",
                    choices=["heuristica", "spacy"],
                    help="método de identificação da relação sintáctica "
                         "(omissão: spacy; use heuristica para comparação)")
    ap.add_argument("--modelo", default="en_core_web_sm",
                    help="modelo spaCy (en_core_web_sm, pt_core_news_sm, ...)")
    ap.add_argument("--incluir-nao-nucleares", action="store_true",
                    help="não filtrar a estatística por nuclear=True")
    ap.add_argument("--inverter-polaridade-negada", action="store_true",
                    help="inverter polaridade quando negado=True (omissão: não)")
    ap.add_argument("--dominios", type=Path, default=None,
                    help="TSV padrao_ficheiro\\tdominio para triagem documental "
                         "(omissão: dominios.tsv junto do projecto, se existir)")
    ap.add_argument("--incluir-dominio", action="append", default=None,
                    help="domínio a readmitir (repetível); omissão: só musicologia")
    ap.add_argument(
        "--dominio-omissao",
        default="musicologia",
        help="domínio quando o TSV não casa (omissão: musicologia; "
             "vazio '' = manter por_rever)",
    )
    ap.add_argument(
        "--sem-falsos-amigos",
        action="store_true",
        help="não aplicar exclusão automática de falsos amigos "
             "(continuo/continuum/continue…)",
    )
    ap.add_argument("--saida", type=Path, default=Path("resultado_near.xlsx"))
    ap.add_argument("--so-extrair", action="store_true", default=True,
                    help="(omissao) so extrai para revisao; sem estatistica")
    ap.add_argument("--sem-revisao", action="store_true",
                    help="extraccao + analise numa so passagem "
                         "(escreve aviso «sem revisao humana»)")
    ap.add_argument("--fases", type=int, choices=[1, 2], default=None,
                    help="alias: 1 = so extrair (fase 1); "
                         "2 = extrair+analisar sem revisao")
    args = ap.parse_args()
    if args.fases == 1:
        args.so_extrair = True
        args.sem_revisao = False
    elif args.fases == 2:
        args.sem_revisao = True
        args.so_extrair = False
    if args.sem_revisao:
        args.so_extrair = False

    campo = dict(CAMPO)
    if args.termos:
        campo = tlex.carregar_campo_termos(args.termos, campo_ref=CAMPO)
        # sincronizar pólos mutáveis
        POLO_ESTABILIDADE.update(tlex.POLO_ESTABILIDADE)
        POLO_VARIABILIDADE.update(tlex.POLO_VARIABILIDADE)
        print(f"      campo lexical externo: {len(campo)} tipos "
              f"({', '.join(sorted(campo))})", flush=True)
    else:
        tlex.registar_campo(campo)
        tlex.assert_campo_sem_no(campo)

    consulta = _Consulta(args.consulta, set(campo)) if args.consulta else None

    print(f"[1/5] A ler {args.xlsx.name} ...", flush=True)
    bruto = pd.read_excel(args.xlsx, sheet_name=args.folha,
                          header=0 if args.com_cabecalho else None,
                          nrows=args.limite)
    total_bruto = len(bruto)
    ncols = bruto.shape[1]
    for rot, k in (("no", args.col_no), ("contexto", args.col_ctx)):
        if not 1 <= k <= ncols:
            print(f"Coluna de {rot} ({k}) fora do intervalo 1-{ncols}.",
                  file=sys.stderr)
            return 2
    col_src = tlex.escolher_coluna_fonte(bruto, args.col_src)
    # Nº de linha Excel 1-based (com cabeçalho: dados começam na linha 2).
    excel_row0 = 2 if args.com_cabecalho else 1
    df = pd.DataFrame({
        "NODE": bruto.iloc[:, args.col_no - 1],
        "contexto": bruto.iloc[:, args.col_ctx - 1],
        "caminho": bruto.iloc[:, col_src - 1],
        "url": (bruto.iloc[:, args.col_url - 1] if 1 <= args.col_url <= ncols
                else ""),
        "source_matrix_row": np.arange(
            excel_row0, excel_row0 + len(bruto), dtype=np.int64),
    })
    print(f"      {total_bruto} linhas, {ncols} colunas | no=col{args.col_no} "
          f"contexto=col{args.col_ctx} fonte=col{col_src} | schema={SCHEMA_NEAR}",
          flush=True)

    # --- limpeza declarada -------------------------------------------------
    df = df.dropna(subset=["contexto"])
    nos_validos = (set().union(*NOS.values()) if args.lingua == "todas"
                   else NOS[args.lingua])
    df["NODE"] = df["NODE"].astype(str).str.lower()
    df = df[df["NODE"].isin(nos_validos)]
    print(f"      {total_bruto} linhas -> {len(df)} após filtro de língua "
          f"'{args.lingua}' e contexto não vazio", flush=True)

    campo_c = compila_campo(campo)
    registos, censura_esq, censura_dir, sem_no = 0, 0, 0, 0
    excluidos_tot = Counter()
    linhas, janelas = [], []
    tot_near = tot_banda = 0
    hits_near, hits_banda = Counter(), Counter()
    partes_termo: dict[str, Counter] = {e: Counter() for e in campo}
    tam_parte: Counter = Counter()

    print(f"[2/5] A extrair co-ocorrências NEAR/{args.near} ...", flush=True)
    for t in df.itertuples(index=False):
        ctx = normaliza(t.contexto)
        toks = tokeniza(ctx)
        if not toks:
            continue
        idxs = indices_no(toks, nos_validos)
        if not idxs:
            sem_no += 1
            continue
        limites = [] if args.sem_fronteira else fronteiras_frase(ctx)
        mesma_frase = not args.sem_fronteira

        # Emparelhamento: produto cartesiano nó×termo → par mínimo (T1).
        achados, excl, n_nos = emparelha_contexto(
            toks, nos_validos, campo_c, args.near, limites,
            mesma_frase=mesma_frase)
        for k, v in excl.items():
            excluidos_tot[k] += v

        # Nó de referência para censura / banda: o mais frequente entre os
        # pares vencedores; se vazio, o mais à esquerda na janela.
        if achados:
            cont_nos = Counter(a["idx_no"] for a in achados)
            i = cont_nos.most_common(1)[0][0]
        else:
            i = min(idxs)

        c_esq = i < args.near
        c_dir = (len(toks) - i - 1) < args.near
        censura_esq += c_esq
        censura_dir += c_dir

        # Banda de referência: termos além de NEAR, ancorados no mesmo nó.
        achados_banda = procura_near(
            toks, i, campo_c, args.banda, limites, mesma_frase=mesma_frase)
        distantes = [a for a in achados_banda
                     if a["distancia"] > args.near]
        tipos_near = {a["termo_tipo"] for a in achados}

        tot_near += conta_tokens(toks, i, limites, 0, args.near, mesma_frase)
        tot_banda += conta_tokens(toks, i, limites, args.near, args.banda,
                                  mesma_frase)
        for a in distantes:
            if a["termo_tipo"] not in tipos_near:
                hits_banda[a["termo_tipo"]] += 1
        for a in achados:
            hits_near[a["termo_tipo"]] += 1
            partes_termo[a["termo_tipo"]][t.caminho] += 1
        tam_parte[t.caminho] += 1

        caminho_f = str(t.caminho)
        doc_id = tlex.doc_id_de_caminho(caminho_f)
        source_row = int(t.source_matrix_row)
        occ_id = occurrence_id_de(doc_id, source_row)
        for a in achados:
            neg, grad, mod, rel = anota_sintaxe(
                toks, a["idx_no"], a["idx_termo"], ctx)
            can = tlex.canonical_de_forma(a["matched_form"], campo)
            # query_pattern: primeiro padrão da etiqueta canónica
            qpat = (campo.get(can) or [a["matched_form"]])[0]
            pol_base = tlex.polaridade(can, False, inverter_negada=False)
            pol = tlex.polaridade(can, neg,
                                 inverter_negada=args.inverter_polaridade_negada)
            linhas.append({
                "source_matrix_row": source_row,
                "texture_occurrence_id": occ_id,
                "match_id": "",  # preenchido após extracção
                "hit_key": "",
                "grupo_passagem_id": "",
                "candidato_duplicado": "",
                "no": a["no"],
                "termo_tipo": can,
                "canonical_term": can,
                "query_pattern": qpat,
                "termo_forma": a["termo_forma"],
                "matched_form": a["matched_form"],
                "n_palavras": a["n_palavras"],
                "distancia": a["distancia"],
                "lado": a["lado"],
                "negado": neg if neg is None else ("directo" if neg else "nao"),
                "graduado": grad,
                "modalizado": mod,
                "relacao_sintactica": rel,
                "polaridade_base": pol_base or "",
                "polaridade": pol or "",
                "eixo": tlex.eixo_semantico(can),
                "censurado_esq": c_esq,
                "censurado_dir": c_dir,
                "idx_no": int(a["idx_no"]),
                "idx_termo": int(a["idx_termo"]),
                "off_no": a["off_no"],
                "off_termo": a["off_termo"],
                "n_nos_janela": n_nos,
                "forma_em_composto": a["forma_em_composto"],
                "caminho_ficheiro": caminho_f,
                "doc_id": doc_id,
                "url": t.url,
                "contexto": ctx,
                "motivo_exclusao": "",
                "nuclear": rel in RELACOES_NUCLEARES,
                "fonte_classificacao": "heuristica",
                "n_janelas_fundidas": 1,
                "revisao_sugerida": "",
                "nucleo_da_propriedade": "",
            })
        if achados:
            presentes = {a["termo_tipo"] for a in achados}
            janelas.append({
                "no": toks[i][0],
                "n_termos": len(presentes),
                "termos": " + ".join(sorted(presentes)),
                "conjunto": presentes,
                "consulta_satisfeita": (consulta.avalia(presentes)
                                        if consulta else None),
                "caminho": t.caminho,
                "contexto": ctx,
            })
        registos += 1

    res = pd.DataFrame(linhas)
    print(f"      {registos} linhas analisadas | {len(res)} co-ocorrencias | "
          f"{sem_no} sem no localizavel | "
          f"excluidos fronteira_frase={excluidos_tot['fronteira_frase']}",
          flush=True)
    if res.empty:
        print("Nenhuma co-ocorrencia. Verifique o campo lexical.", file=sys.stderr)
        return 1

    # R1 — asserção defensiva sobre o output
    tlex.assert_output_sem_no(res["canonical_term"].unique())

    # R2 — doc_id 1-por-linha + caminhos sem extensao = coluna errada
    frac_ficheiro = float(res["caminho_ficheiro"].map(
        tlex.parece_caminho_ficheiro).mean())
    if frac_ficheiro < 0.2:
        raise SystemExit(
            "Assercao falhou: caminho_ficheiro nao contem ficheiros "
            f"(fracao com extensao={frac_ficheiro:.2f}). "
            "Verifique --col-src (nao use a coluna 'raiz'/directorio).")
    if (len(res) > 5 and res["doc_id"].nunique() == len(res)
            and frac_ficheiro < 0.5):
        raise SystemExit(
            "Assercao falhou: n_doc_id == n_linhas com caminhos duvidosos. "
            "Verifique --col-src.")

    if args.sintaxe == "spacy":
        print("[2b/5] Analise de dependencias (spaCy) ...", flush=True)
        res = anota_com_spacy(res, args.modelo, obrigatorio=True)
    else:
        print("[2b/5] Classificacao heuristica ...", flush=True)
        res = anota_com_heuristica(res)

    n_bruto = len(res)
    por_rever = pd.DataFrame(columns=["caminho", "n_hits", "n_nucleares"])
    if ttri is not None:
        dom_path = args.dominios
        if dom_path is None:
            cand = Path(__file__).resolve().parent / "dominios.tsv"
            if cand.is_file():
                dom_path = cand
        regras = ttri.carregar_dominios(dom_path)
        incluir = set(args.incluir_dominio) if args.incluir_dominio else None
        omis = (args.dominio_omissao or "").strip() or None
        res, _, por_rever = ttri.aplicar_triagem(
            res,
            regras_dominio=regras,
            incluir_dominios=incluir,
            dominio_omissao=omis,
            aplicar_amigos=not args.sem_falsos_amigos,
        )
        for i, row in res.iterrows():
            toks = tokeniza(str(row["contexto"]))
            i_no = next((k for k, (_, o) in enumerate(toks)
                         if o == int(row["off_no"])), None)
            i_te = next((k for k, (_, o) in enumerate(toks)
                         if o == int(row["off_termo"])), None)
            if i_no is not None and i_te is not None and ttri.e_ruido_ocr(
                    toks, i_no, i_te):
                res.at[i, "nuclear"] = False
                if not res.at[i, "motivo_exclusao"]:
                    res.at[i, "motivo_exclusao"] = "ruido_ocr"

    # R7 — mesma assinatura (contexto + percurso) => mesmo nuclear/motivo
    if "percurso_dep" in res.columns:
        for _, grp in res.groupby(
                [res["contexto"].map(lambda c: " ".join(
                    w for w, _ in tokeniza(str(c)))),
                 "percurso_dep"], sort=False):
            if len(grp) < 2:
                continue
            # consenso: se algum nuclear=True sem metatexto residual, alinhar
            nucs = grp["nuclear"].astype(bool)
            if nucs.nunique() > 1:
                # preferir o veredicto maioritário; empate -> False
                ver = bool(nucs.mode().iloc[0])
                mot = grp.loc[grp["nuclear"] == ver, "motivo_exclusao"]
                mot0 = mot.iloc[0] if len(mot) else ""
                for i in grp.index:
                    res.at[i, "nuclear"] = ver
                    res.at[i, "motivo_exclusao"] = (
                        "" if ver else (res.at[i, "motivo_exclusao"] or mot0
                                        or "indeterminada"))

    # IDs de hit + dedupe exacto + fusão de janelas (mesmo termo)
    res = atribuir_match_ids(res)
    n_antes_dedupe = len(res)
    res = deduplicar_hits_exactos(res)
    res = fundir_janelas_e_marcar_duplicados(res, ngrama=8)
    # Rematch IDs após eventual remoção de cópias exactas
    if len(res) != n_antes_dedupe:
        res = atribuir_match_ids(res)
    res = reordenar_colunas_hits(res)

    res_excluidas = res.loc[~res["nuclear"].astype(bool)].copy() if (
        "nuclear" in res.columns) else res.iloc[0:0].copy()
    if not args.incluir_nao_nucleares and "nuclear" in res.columns:
        res_stat = res.loc[res["nuclear"].astype(bool)].copy()
    else:
        res_stat = res.copy()
    n_nuc = len(res_stat)
    n_ficheiros = (res["caminho_ficheiro"].nunique()
                   if "caminho_ficheiro" in res.columns else 0)
    n_obras = res["doc_id"].nunique() if "doc_id" in res.columns else 0
    n_ocorrencias = (res["texture_occurrence_id"].nunique()
                     if "texture_occurrence_id" in res.columns else 0)
    n_ocorrencias_nuc = (res_stat["texture_occurrence_id"].nunique()
                         if "texture_occurrence_id" in res_stat.columns else 0)
    print(f"      cascata: brutas={n_bruto} -> nucleares(hits)={n_nuc} "
          f"({100*n_nuc/max(n_bruto,1):.1f}%) | "
          f"excluidas={len(res_excluidas)} | "
          f"ocorrencias={n_ocorrencias} (nucleares={n_ocorrencias_nuc}) | "
          f"ficheiros={n_ficheiros} obras/doc_id={n_obras}", flush=True)

    col_doc = "doc_id" if "doc_id" in res_stat.columns else "caminho_ficheiro"
    if "texture_occurrence_id" in res_stat.columns:
        res_obra = res_stat.drop_duplicates(
            subset=["texture_occurrence_id", "termo_tipo"])
    else:
        res_obra = res_stat.drop_duplicates(subset=[col_doc, "termo_tipo"])
    ocorrencias_df = agregar_ocorrencias(res)

    # Folha Duplicados: caminhos múltiplos + grupos de passagem/janela
    dup_rows = []
    if "doc_id" in res.columns:
        g = (res.groupby("doc_id")["caminho_ficheiro"]
             .agg(lambda s: sorted(set(map(str, s))))
             .reset_index())
        for row in g.itertuples(index=False):
            if len(row.caminho_ficheiro) > 1:
                dup_rows.append({
                    "tipo": "mesmo_doc_id_varios_caminhos",
                    "doc_id": row.doc_id,
                    "grupo": "",
                    "n": len(row.caminho_ficheiro),
                    "detalhe": " | ".join(row.caminho_ficheiro),
                })
    if "grupo_passagem_id" in res.columns:
        gpass = res.loc[
            res["grupo_passagem_id"].astype(str).str.strip().ne("")
            & ~res["grupo_passagem_id"].astype(str).str.lower().isin(
                {"nan", "none"})
        ]
        for gid, grp in gpass.groupby("grupo_passagem_id", sort=False):
            termos = sorted({str(t) for t in grp["canonical_term"]})
            n_jan = int((grp["motivo_exclusao"].astype(str)
                         == "janela_sobreposta").sum())
            dup_rows.append({
                "tipo": "passagem_sobreposta",
                "doc_id": grp["doc_id"].iloc[0],
                "grupo": str(gid),
                "n": int(len(grp)),
                "detalhe": (
                    f"hits={len(grp)} janela_sobreposta={n_jan} "
                    f"termos={'; '.join(termos)} | "
                    f"{str(grp['contexto'].iloc[0])[:100]}"
                ),
            })
    duplicados = pd.DataFrame(dup_rows)

    # Colunas de revisao humana (fase 1)
    if "revisto_por_humano" not in res.columns:
        res["revisto_por_humano"] = ""
    if "nota_revisao" not in res.columns:
        res["nota_revisao"] = ""

    # --- FASE 1: sempre escrever Excel de revisao -------------------------
    from datetime import datetime as _dt
    comando = " ".join(sys.argv)
    manifesto = hashlib.sha256(
        pd.Series(df["caminho"].astype(str).unique()).sort_values()
        .str.cat(sep="\n").encode("utf-8", errors="replace")
    ).hexdigest()
    instr = pd.DataFrame({
        "chave": [
            "fase", "schema_near", "data_fase1", "comando", "sem_revisao",
            "janelas_kwic_processadas", "tot_near", "tot_banda",
            "hits_banda", "campo_tipos", "manifesto_sha256",
            "n_hits", "n_ocorrencias", "n_hits_nucleares",
            "n_ocorrencias_nucleares",
            "como_rever", "comando_fase2", "unidade_contagem",
        ],
        "valor": [
            "1 - extracao",
            SCHEMA_NEAR,
            _dt.now().isoformat(timespec="seconds"),
            comando,
            "sim" if args.sem_revisao else "nao",
            registos, tot_near, tot_banda,
            repr(dict(hits_banda)),
            ",".join(sorted(campo)),
            manifesto,
            len(res), n_ocorrencias, n_nuc, n_ocorrencias_nuc,
            "Edite as colunas a amarelo em 8_Concordancia (= hits NEAR). "
            "Nao altere source_matrix_row, texture_occurrence_id, match_id, "
            "hit_key, canonical_term nem matched_form. "
            "8_Concordancia_Ocorrencias e so leitura (1 linha = 1 linha da matriz). "
            "N_hits = linhas nucleares em 8_Concordancia; "
            "N_ocorrencias = texture_occurrence_id unicos com hit nuclear.",
            f'python textura_analise.py --xlsx "{args.saida}"',
            "hit=8_Concordancia; ocorrencia=texture_occurrence_id "
            "(linha da matriz)",
        ],
    })
    cfg_lex = pd.DataFrame([
        {"etiqueta": k, "padroes": ", ".join(v),
         "polaridade": (tlex.polaridade(k) or ""),
         "eixo": tlex.eixo_semantico(k),
         "nota_eixo": (
             "PROPOSTA: varied->invariancia_diacronica (ou ambos); "
             "nao alterar sem adjudicação"
             if k == "varied" else "")}
        for k, v in campo.items()
    ])
    drop_cols = [c for c in ("relacao", "caminho_dep", "atribuicao", "caminho")
                 if c in res.columns]
    conc_out = res.drop(columns=drop_cols, errors="ignore")
    # Alias explícito do nível hits (mesma folha de trabalho)
    hits_out = conc_out
    if len(ocorrencias_df) == 0:
        ocorrencias_df = agregar_ocorrencias(res)

    print(f"[3/5] A escrever extraccao (fase 1) -> {args.saida.name} ...",
          flush=True)
    with pd.ExcelWriter(args.saida, engine="openpyxl") as xw:
        instr.to_excel(xw, sheet_name="0_Instrucoes", index=False)
        cfg_lex.to_excel(xw, sheet_name="Config_lexico", index=False)
        pd.DataFrame({
            "caminho": sorted(df["caminho"].astype(str).unique()),
        }).to_excel(xw, sheet_name="Manifesto_corpus", index=False)
        conc_out.to_excel(xw, sheet_name="8_Concordancia", index=False)
        hits_out.to_excel(xw, sheet_name="8_Concordancia_Hits", index=False)
        if len(ocorrencias_df):
            ocorrencias_df.to_excel(
                xw, sheet_name="8_Concordancia_Ocorrencias", index=False)
        if len(res_excluidas):
            res_excluidas.drop(columns=drop_cols, errors="ignore").to_excel(
                xw, sheet_name="9_Excluidas", index=False)
        if len(por_rever):
            por_rever.to_excel(xw, sheet_name="Dominios_por_rever", index=False)
        if len(duplicados):
            duplicados.to_excel(xw, sheet_name="Duplicados", index=False)

    # Validacao de dados + destaque amarelo nas colunas editaveis
    try:
        from openpyxl.worksheet.datavalidation import DataValidation
        wb = load_workbook(args.saida)
        wsc = wb["8_Concordancia"]
        cab = [c.value for c in wsc[1]]
        amarelo = PatternFill("solid", fgColor="FFF2CC")
        editaveis = {
            "relacao_sintactica": ",".join(sorted(
                RELACOES_NUCLEARES | {
                    "incidental", "adverbial_verbal", "adverbial_de_grau",
                    "coordenada", "indeterminada"})),
            "nuclear": "TRUE,FALSE",
            "polaridade": "estabilidade,variabilidade,",
            "eixo": "homogeneidade_sincronica,invariancia_diacronica,ambos,",
            "negado": "nao,directo,indirecto",
        }
        for col_nome, lista in editaveis.items():
            if col_nome not in cab:
                continue
            j = cab.index(col_nome) + 1
            letra = get_column_letter(j)
            for r in range(2, wsc.max_row + 1):
                wsc.cell(row=r, column=j).fill = amarelo
            dv = DataValidation(type="list", formula1=f'"{lista}"',
                                allow_blank=True)
            dv.error = "Valor fora da taxonomia"
            dv.errorTitle = "Invalido"
            wsc.add_data_validation(dv)
            dv.add(f"{letra}2:{letra}{wsc.max_row}")
        # tambem amarelo em dominio / motivo / revisao / candidato
        for col_nome in ("dominio", "motivo_exclusao", "revisto_por_humano",
                         "nota_revisao", "candidato_duplicado"):
            if col_nome in cab:
                j = cab.index(col_nome) + 1
                for r in range(2, wsc.max_row + 1):
                    wsc.cell(row=r, column=j).fill = amarelo
        wb.save(args.saida)
    except Exception as exc:
        print(f"      AVISO: validacao Excel nao aplicada ({exc})", flush=True)

    if not args.sem_revisao:
        print("\n=== FASE 1 concluida ===")
        print(f"Reveja a folha 8_Concordancia (hits) em:\n  {args.saida}")
        print("Folha 8_Concordancia_Ocorrencias = 1 linha por linha da matriz.")
        print("Depois corra:")
        print(f'  python textura_analise.py --xlsx "{args.saida}"')
        print(f"Cascata: hits={n_bruto} nucleares(hits)={n_nuc} "
              f"ocorrencias={n_ocorrencias} (nuc={n_ocorrencias_nuc}) "
              f"ficheiros={n_ficheiros} doc_id={n_obras}")
        return 0

    # --- FASE integrada (--sem-revisao): analisa de seguida ---------------
    print("[4/5] --sem-revisao: a correr fase 2 (AVISO: sem revisao humana) ...",
          flush=True)
    import textura_analise as tanal
    return tanal.analisar(args.saida, args.saida, nulo_polaridade="banda",
                          cooc_unidade="obra")

if __name__ == "__main__":
    raise SystemExit(main())
