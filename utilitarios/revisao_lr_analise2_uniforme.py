#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Revisão LR (GUIA_REVISAO_FASE1) — Análise 2 / Textura_Uniforme_near.xlsx.

Preserva formatação amarela e listas do Excel original (openpyxl).
Saída no mesmo pasta: Textura_Uniforme_near_revisto_LR.xlsx
"""

from __future__ import annotations

import re
import shutil
import sys
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

import textura_triagem as tri  # noqa: E402

SRC = Path(
    r"C:\Users\lmr20\Desktop\Tesaurus e Dicionários"
    r"\CLASSES TEXTURAIS\UNIFORM\Análise 2\Textura_Uniforme_near.xlsx"
)
OUT = SRC.with_name("Textura_Uniforme_near_revisto_LR.xlsx")
REVISOR = "LR"

NUCLEARES = {
    "atributiva", "predicativa", "predicativa_secundaria",
    "nominal_composto", "nominal_genitiva", "adverbial",
}
NAO_NUCLEARES = {
    "incidental", "adverbial_verbal", "adverbial_de_grau",
    "coordenada", "indeterminada",
}

# Falsos amigos / fora de classe (matched_form.lower())
FORMAS_EXCLUIR = {
    "continuo": "termo_nao_relacionado_directamente_com_textura",
    "continue": "termo_nao_relacionado_directamente_com_textura",
    "continues": "termo_nao_relacionado_directamente_com_textura",
    "continuation": "termo_nao_relacionado_directamente_com_textura",
    "continuum": "fora_da_classe_uniforme",
    "continua": "fora_da_classe_uniforme",
    "continuums": "fora_da_classe_uniforme",
    "constants": "fora_de_dominio",
    # even* → event / events / eventually… (não são «even/uniforme»)
    "event": "termo_nao_relacionado_directamente_com_textura",
    "events": "termo_nao_relacionado_directamente_com_textura",
    "eventful": "termo_nao_relacionado_directamente_com_textura",
    "eventfulness": "termo_nao_relacionado_directamente_com_textura",
    "eventually": "termo_nao_relacionado_directamente_com_textura",
}

EDIT_COLS = [
    "relacao_sintactica", "nuclear", "polaridade", "eixo", "negado",
    "candidato_duplicado", "dominio", "motivo_exclusao",
    "revisto_por_humano", "nota_revisao",
]

RE_META_EXTRA = re.compile(
    r"(?i)(jstor\.org/terms|all rights reserved|this content downloaded|"
    r"tandfonline\.com|copyright\s*©|terms of use)"
)
RE_EVEN_DISCOURSE = re.compile(
    r"(?i)\beven\s+(a|an|the|if|when|though|more|less|so|as)\b"
)


def _truthy(v) -> bool:
    if isinstance(v, bool):
        return v
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return False
    return str(v).strip().lower() in {"1", "true", "verdadeiro", "yes", "sim", "t", "y"}


def _empty(v) -> bool:
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return True
    s = str(v).strip()
    return s == "" or s.lower() == "nan"


def _note(df: pd.DataFrame, i, text: str) -> None:
    prev = df.at[i, "nota_revisao"]
    base = "" if _empty(prev) else str(prev).rstrip(" |") + " | "
    df.at[i, "nota_revisao"] = base + text


def _excluir(df, i, motivo: str, nota: str, stats: dict, key: str) -> None:
    if _truthy(df.at[i, "nuclear"]):
        stats[key] = stats.get(key, 0) + 1
    df.at[i, "nuclear"] = False
    fortes = {
        "metatexto", "ruido_ocr", "fora_da_classe_uniforme",
        "termo_nao_relacionado_directamente_com_textura", "fora_de_dominio",
        "citacao_repetida", "janela_sobreposta",
    }
    cur = "" if _empty(df.at[i, "motivo_exclusao"]) else str(df.at[i, "motivo_exclusao"])
    if not cur or cur in NAO_NUCLEARES or cur not in fortes:
        if cur not in fortes:
            df.at[i, "motivo_exclusao"] = motivo
    if _empty(df.at[i, "motivo_exclusao"]):
        df.at[i, "motivo_exclusao"] = motivo
    _note(df, i, nota)


def revisar(df: pd.DataFrame, cfg: pd.DataFrame) -> tuple[pd.DataFrame, dict]:
    for col in (
        "nota_revisao", "revisto_por_humano", "motivo_exclusao",
        "polaridade", "eixo", "dominio", "negado",
    ):
        if col in df.columns:
            df[col] = df[col].astype(object)

    pol_map, eixo_map = {}, {}
    if cfg is not None and not cfg.empty:
        for _, r in cfg.iterrows():
            et = str(r.get("etiqueta", "")).strip()
            if not _empty(r.get("polaridade")):
                pol_map[et] = str(r["polaridade"]).strip()
            if not _empty(r.get("eixo")):
                eixo_map[et] = str(r["eixo"]).strip()

    stats: dict[str, int] = {}

    # C — relação ↔ nuclear
    for i, row in df.iterrows():
        rel = str(row.get("relacao_sintactica") or "").strip()
        if rel in NAO_NUCLEARES and _truthy(row.get("nuclear")):
            _excluir(
                df, i, rel,
                f"LR: relação '{rel}' não nuclear",
                stats, "coerencia_relacao",
            )
        elif rel in NAO_NUCLEARES and _empty(row.get("motivo_exclusao")):
            df.at[i, "motivo_exclusao"] = rel

    # B — duplicados mesmo termo_tipo na mesma passagem
    if "grupo_passagem_id" in df.columns:
        gids = df.loc[
            df["grupo_passagem_id"].notna() & df["nuclear"].map(_truthy),
            "grupo_passagem_id",
        ].unique()
        for gid in gids:
            block = df[(df["grupo_passagem_id"] == gid) & (df["nuclear"].map(_truthy))]
            for tipo, sub in block.groupby(block["termo_tipo"].astype(str)):
                if len(sub) < 2:
                    continue
                cand = list(sub.index)

                def score(idx):
                    r = df.loc[idx]
                    return (
                        -len(str(r["matched_form"])),
                        float(r["distancia"]) if pd.notna(r["distancia"]) else 99.0,
                        str(r["hit_key"]),
                    )

                cand.sort(key=score)
                keep = cand[0]
                for idx in cand[1:]:
                    _excluir(
                        df, idx, "janela_sobreposta",
                        f"LR: duplicado tipo '{tipo}' em {gid}; "
                        f"mantida '{df.at[keep, 'matched_form']}'",
                        stats, "duplicado_mesmo_tipo",
                    )

    # Falsos amigos / metatexto / domínio
    for i, row in df.iterrows():
        form = str(row.get("matched_form") or "").strip().lower()
        ctx = str(row.get("contexto") or "")

        if _empty(row.get("dominio")) or str(row.get("dominio")).strip() == "por_rever":
            df.at[i, "dominio"] = tri.DOMINIO_OMISSAO
            stats["dominio_musicologia"] = stats.get("dominio_musicologia", 0) + 1

        if not _truthy(df.at[i, "nuclear"]):
            continue

        if form in FORMAS_EXCLUIR:
            _excluir(
                df, i, FORMAS_EXCLUIR[form],
                f"LR: forma '{form}' fora do critério da classe / falso amigo",
                stats, f"forma:{form}",
            )
            continue

        if form == "even" and RE_EVEN_DISCOURSE.search(ctx):
            if "texture" not in ctx.lower() or not re.search(
                r"(?i)even\s+\w*\s*textures?", ctx
            ):
                _excluir(
                    df, i, "termo_nao_relacionado_directamente_com_textura",
                    "LR: 'even' discursivo, não propriedade textural",
                    stats, "even_discourse",
                )
                continue

        if tri.e_metatexto(ctx) or RE_META_EXTRA.search(ctx):
            _excluir(
                df, i, "metatexto",
                "LR: metatexto / copyright / cabeçalho editorial",
                stats, "metatexto",
            )
            continue

    # D — polaridade / eixo / negado
    for i, row in df.iterrows():
        if not _truthy(row.get("nuclear")):
            continue
        tipo = str(row.get("termo_tipo") or "").strip()
        if _empty(row.get("polaridade")):
            if not _empty(row.get("polaridade_base")):
                df.at[i, "polaridade"] = str(row["polaridade_base"]).strip()
            elif tipo in pol_map:
                df.at[i, "polaridade"] = pol_map[tipo]
            else:
                df.at[i, "polaridade"] = "estabilidade"
            stats["polaridade"] = stats.get("polaridade", 0) + 1
        if _empty(row.get("eixo")):
            df.at[i, "eixo"] = eixo_map.get(tipo, "ambos")
            stats["eixo"] = stats.get("eixo", 0) + 1
        if _empty(row.get("negado")):
            df.at[i, "negado"] = "nao"
            stats["negado"] = stats.get("negado", 0) + 1

    # F — marca de revisão
    for i in df.index:
        df.at[i, "revisto_por_humano"] = REVISOR
    stats["marcadas_LR"] = len(df)
    return df, stats


def _headers(ws) -> dict[str, int]:
    return {
        str(ws.cell(1, c).value): c
        for c in range(1, ws.max_column + 1)
        if ws.cell(1, c).value is not None
    }


def _write_edit_cols(ws, df: pd.DataFrame) -> None:
    hdr = _headers(ws)
    for col in EDIT_COLS:
        if col not in hdr or col not in df.columns:
            continue
        c = hdr[col]
        for r_i, idx in enumerate(df.index, start=2):
            val = df.at[idx, col]
            if col == "nuclear":
                cell_val = bool(_truthy(val))
            elif _empty(val):
                cell_val = None
            else:
                cell_val = val
                if isinstance(cell_val, float) and pd.isna(cell_val):
                    cell_val = None
            ws.cell(r_i, c).value = cell_val


def _replace_sheet_with_df(wb, name: str, df: pd.DataFrame) -> None:
    if name in wb.sheetnames:
        ws_old = wb[name]
        # preserve sheet position
        idx = wb.sheetnames.index(name)
        wb.remove(ws_old)
        ws = wb.create_sheet(name, idx)
    else:
        ws = wb.create_sheet(name)
    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append(row)


def main() -> int:
    if not SRC.is_file():
        raise SystemExit(f"Fonte ausente: {SRC}")

    cfg = pd.read_excel(SRC, sheet_name="Config_lexico")
    df0 = pd.read_excel(SRC, sheet_name="8_Concordancia")
    df, stats = revisar(df0.copy(), cfg)

    # backup original once
    bak = SRC.with_suffix(SRC.suffix + ".bak-antes-LR")
    if not bak.exists():
        shutil.copy2(SRC, bak)

    wb = load_workbook(SRC)
    _write_edit_cols(wb["8_Concordancia"], df)

    # Hits = cópia dos valores editáveis
    if "8_Concordancia_Hits" in wb.sheetnames:
        _write_edit_cols(wb["8_Concordancia_Hits"], df)

    excl = df.loc[~df["nuclear"].map(_truthy)].copy()
    _replace_sheet_with_df(wb, "9_Excluidas", excl)

    still = df[df["dominio"].astype(str).str.strip() == "por_rever"]
    if still.empty:
        dom = pd.DataFrame(columns=["caminho", "n_hits", "n_nucleares"])
    else:
        dom = (
            still.groupby("caminho_ficheiro", dropna=False)
            .agg(
                n_hits=("hit_key", "count"),
                n_nucleares=("nuclear", lambda s: int(s.map(_truthy).sum())),
            )
            .reset_index()
            .rename(columns={"caminho_ficheiro": "caminho"})
        )
    _replace_sheet_with_df(wb, "Dominios_por_rever", dom)

    # nota na folha de instruções
    if "0_Instrucoes" in wb.sheetnames:
        ws = wb["0_Instrucoes"]
        # append meta rows if chave/valor layout
        keys = {str(ws.cell(r, 1).value): r for r in range(1, ws.max_row + 1)}
        n_nuc = int(df["nuclear"].map(_truthy).sum())
        updates = {
            "revisto_por_humano": REVISOR,
            "n_hits_nucleares_apos_revisao": n_nuc,
            "guia_revisao": "GUIA_REVISAO_FASE1.md",
        }
        for k, v in updates.items():
            if k in keys:
                ws.cell(keys[k], 2).value = v
            else:
                ws.append([k, v])

    try:
        wb.save(OUT)
        dest = OUT
    except PermissionError:
        dest = OUT.with_name("Textura_Uniforme_near_revisto_LR_FULL.xlsx")
        wb.save(dest)
        print(f"AVISO: destino bloqueado — gravei em {dest.name}", flush=True)

    n_nuc = int(df["nuclear"].map(_truthy).sum())
    print(f"Fonte: {SRC}")
    print(f"Saída: {dest}")
    print(f"Total={len(df)} | nuclear_TRUE={n_nuc} | excluídas={len(df) - n_nuc}")
    print("Edições:", {k: v for k, v in sorted(stats.items())})
    print(
        "motivo_exclusao (FALSE):\n",
        df.loc[~df["nuclear"].map(_truthy), "motivo_exclusao"]
        .value_counts(dropna=False)
        .head(20)
        .to_string(),
    )
    print(
        "checklist: pol vazia=",
        int(df.loc[df["nuclear"].map(_truthy), "polaridade"].map(_empty).sum()),
        "eixo vazio=",
        int(df.loc[df["nuclear"].map(_truthy), "eixo"].map(_empty).sum()),
        "por_rever=",
        int((df["dominio"].astype(str) == "por_rever").sum()),
        "TRUE má-relação=",
        int(
            (
                df["nuclear"].map(_truthy)
                & df["relacao_sintactica"].isin(NAO_NUCLEARES)
            ).sum()
        ),
        "revisto=",
        int(df["revisto_por_humano"].astype(str).str.strip().eq(REVISOR).sum()),
    )
    # doctor checklist
    chk = tri.checklist_revisao(df)
    print("doctor:", {k: chk[k] for k in ("ok", "score", "n", "n_nuclear", "erros", "avisos")})
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
