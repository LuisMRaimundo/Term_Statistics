#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
textura_concordancia_qa.py — QA pass over ``8_Concordancia``
===========================================================

Post-hoc complement to ``textura_near.py`` (which already demotes
near-duplicate cross-doc citations, flags heterogeneous coordination in
``revisao_sugerida``, and writes ``dominio_janela`` at extraction time).
Use this script on *already extracted / revised* workbooks, or as a second
opinion with different thresholds (Jaccard/containment vs 8-gram share).

Three independent, auditable modules:

  1. DUPLICATES  — exact and near-duplicate context windows (cross- and
     within-document), robust to OCR/edition variants; writes QD tags to
     ``candidato_duplicado``, colours, and (optionally) demotes redundant
     copies keeping one attestation per group.
  2. RELATIONS   — pattern-based classification of the term–texture
     syntactic connection over the KWIC window, as a *second opinion*
     against the dependency-based ``relacao_sintactica``; writes suggestions
     to new columns, never overwriting pipeline or human decisions.
  3. DOMAIN      — lexicon-based detection of extra-musical uses of
     "texture" (geology, visual arts, haptics, speech, etc.) that the
     ``dominio`` column missed.

Design rules: idempotent (re-running does not duplicate tags / columns),
dry-run by default, every automatic decision written to *new* columns with
an explicit motive string, colours used only as visual mirrors of those
columns.

Usage::

    python textura_concordancia_qa.py --xlsx IN.xlsx
    python textura_concordancia_qa.py --xlsx IN.xlsx --saida OUT.xlsx --apply
    python textura_concordancia_qa.py --xlsx IN.xlsx -o OUT.xlsx --apply --demote-duplicates
"""

from __future__ import annotations

import argparse
import json
import re
import sys
import unicodedata
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from textura.lexico import DOMAIN_LEXICON

SHEET = "8_Concordancia"

FILL_DUP = PatternFill("solid", fgColor="00FF00")      # green  — duplicate
FILL_REL = PatternFill("solid", fgColor="FFFF00")      # yellow — no clear link
FILL_DOM = PatternFill("solid", fgColor="FF9999")      # red    — extra-musical
FILL_REV = PatternFill("solid", fgColor="00FFFF")      # turquoise — review

# Post-hoc columns written by this script only (not part of the pipeline
# column dictionary in dados/dicionario_colunas.md):
#   qa_relacao            — second-opinion relation label
#   qa_relacao_evidencia  — short evidence / motive string
#   qa_dominio_extra      — extra-musical domain guess
#   qa_dominio_cue        — cue token that triggered the domain guess
QA_COLS = (
    "qa_relacao",
    "qa_relacao_evidencia",
    "qa_dominio_extra",
    "qa_dominio_cue",
)

# --------------------------------------------------------------------------
# Normalisation
# --------------------------------------------------------------------------

LIGATURES = {
    "ﬁ": "fi", "ﬂ": "fl", "ﬀ": "ff", "ﬃ": "ffi", "ﬄ": "ffl",
    "’": "'", "‘": "'", "“": '"', "”": '"',
}


def norm(s: str) -> str:
    """Word-level key: lowercase ASCII words, punctuation stripped."""
    if not isinstance(s, str):
        return ""
    for k, v in LIGATURES.items():
        s = s.replace(k, v)
    s = unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode()
    s = re.sub(r"[^a-z0-9]+", " ", s.lower())
    return s.strip()


def fused(s: str) -> str:
    """Space-free key: catches hyphenation/joining variants
    ('ever-changing' vs 'everchanging', 'wordpainting' vs 'word painting')."""
    return norm(s).replace(" ", "")


def _truthy(val: Any) -> bool:
    if isinstance(val, bool):
        return val
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return False
    return str(val).strip().lower() in {"1", "true", "yes", "y", "sim"}


# --------------------------------------------------------------------------
# 1. Duplicate detection
# --------------------------------------------------------------------------

def shingle_set(words, k=5):
    if len(words) >= k:
        return {" ".join(words[i:i + k]) for i in range(len(words) - k + 1)}
    return set()


def char_grams(s, k=24):
    s = fused(s)
    if len(s) >= k:
        return {s[i:i + k] for i in range(len(s) - k + 1)}
    return {s} if s else set()


def find_duplicate_groups(df, jaccard=0.45, containment=0.80,
                          min_words=8, max_posting=25):
    """Union-find groups of near-equal ``contexto`` windows.

    Two-tier signature: 5-word shingles for normal windows, 24-char grams of
    the fused string for short windows (< min_words), so that short excerpts
    are compared rather than silently skipped, while empty/whitespace rows
    are excluded outright (they previously produced spurious mega-groups).
    """
    texts = df["contexto"].map(norm)
    sigs, index = {}, defaultdict(set)
    for i in df.index:
        w = texts[i].split()
        if not w:
            continue
        sig = shingle_set(w) if len(w) >= min_words else char_grams(texts[i])
        if not sig:
            continue
        sigs[i] = sig
        for s in sig:
            index[s].add(i)

    pair_hits = Counter()
    for s, ids in index.items():
        if 1 < len(ids) <= max_posting:
            ids = sorted(ids)
            for a in range(len(ids)):
                for b in range(a + 1, len(ids)):
                    pair_hits[(ids[a], ids[b])] += 1

    parent = {i: i for i in sigs}

    def find(x):
        while parent[x] != x:
            parent[x] = parent[parent[x]]
            x = parent[x]
        return x

    for (a, b), c in pair_hits.items():
        if c < 2:
            continue
        A, B = sigs[a], sigs[b]
        inter = len(A & B)
        if inter / len(A | B) >= jaccard or inter / min(len(A), len(B)) >= containment:
            ra, rb = find(a), find(b)
            if ra != rb:
                parent[ra] = rb

    groups = defaultdict(list)
    for i in sigs:
        groups[find(i)].append(i)
    groups = [sorted(v) for v in groups.values() if len(v) > 1]
    groups.sort(key=lambda g: g[0])

    tracked, new = [], []
    if "grupo_passagem_id" in df.columns:
        for g in groups:
            gp = df.loc[g, "grupo_passagem_id"]
            (tracked if gp.notna().all() and gp.nunique() == 1 else new).append(g)
    else:
        new = groups
    return new, tracked


def keeper_score(df, r):
    """Retention preference inside a duplicate group (higher = keep)."""
    path = str(df.loc[r, "caminho_ficheiro"])
    fn = path.replace("\\", "/").rsplit("/", 1)[-1]
    s = 0
    if "todos os textos" in path:
        s += 4                      # catalogued corpus folder
    if re.match(r"^\(\d{4}\)", fn):
        s += 2                      # '(YYYY)_Title' naming convention
    if "__" in fn:
        s -= 1                      # hash-suffixed second copies
    return (s, len(str(df.loc[r, "contexto"])))


# --------------------------------------------------------------------------
# 2. Term–texture relation (pattern-based second opinion)
# --------------------------------------------------------------------------

TEXTURE = r"(?:textures?|textura[ls]?(?:is)?|textural(?:ly)?|texture)"
# NP material allowed between the term and 'texture'
GAP = r"(?:\w+\s+){0,3}"

NON_TEXTURAL_HEADS = set(
    "melody melodies melodic rhythm rhythms rhythmic rhythmical harmony "
    "harmonies harmonic timbre timbres pitch pitches dynamics instrument "
    "instruments voice voices colour colours color colors lyric lyrics text "
    "texts theme themes thematic tempo tempi register registers structure "
    "structures flavor flavors flavour flavours space spaces imagery "
    "notation meter meters metre metres tone tones pattern patterns gesture "
    "gestures effect effects perception perceptions time development "
    "developments".split())


def _has_head(fragment):
    return any(w in NON_TEXTURAL_HEADS for w in fragment.split())


def classify_relation(context: str, matched_form: str):
    """Return (category, evidence) for the term–texture link in the window.

    Categories:
      direct           — term predicates texture(s) itself
      coord_heterog    — texture only one combinand among heterogeneous items
      into_texture     — elements combined/blended INTO a texture
      with_texture     — X combined/blended WITH a texture
      absent           — 'texture' not connected to the term in the window
    The classifier is deliberately conservative: anything not matching a
    pattern falls back to 'unresolved' for human review.
    """
    c = " " + norm(context) + " "
    t = norm(matched_form)
    if not t or t not in c:
        return "absent", "matched form not found in window"

    T = re.escape(t)

    # -- direct patterns --------------------------------------------------
    direct = [
        rf"{T}\s+(?:of|de|d)\s+{GAP}{TEXTURE}",          # complexity of texture
        rf"{TEXTURE}\s+{GAP}{T}",                        # textural complexity / texture combines
        rf"{T}\s+{GAP}{TEXTURE}",                        # complex (polyphonic) texture(s)
        rf"{TEXTURE}\s+(?:is|are|as)\s+(?:\w+\s+){{0,2}}{T}",  # texture is a combination
    ]
    # -- non-direct patterns (checked first: more specific) ---------------
    # texture coordinated after the term:  "combination of texture(s) and X"
    coord = (rf"{T}\s+of\s+(?:\w+\s+){{0,6}}?{TEXTURE}"
             rf"\s+(?:and|e|,)\s+((?:\w+\s*){{1,3}})")
    # texture coordinated later in a list: "combination of A(, B) and texture"
    coord2 = (rf"{T}s?\s+of\s+((?:\w+\s*,?\s*){{1,6}})"
              rf"\band\s+{GAP}{TEXTURE}")
    # texture as SUBJECT of the term, joined with a non-textural element:
    # "textures (are) combined/blended with lyrics"
    subj_with = (rf"{TEXTURE}\s+(?:is|are|was|were)?\s*{T}(?:ed|d|s)?"
                 rf"\s+with\s+((?:\w+\s*){{1,4}})")
    # term(ing) [small gap] with ... texture(s) and NON-TEXTURAL head
    obj_with = (rf"{T}(?:ed|d|s|ing)?\s+(?:\w+\s+){{0,2}}with\s+"
                rf"(?:\w+\s+){{0,2}}{TEXTURE}\s+and\s+((?:\w+\s*){{1,3}})")
    into = rf"{T}(?:ed|s)?\s+(?:in)?to\s+(?:a|an|the|one)?\s*{GAP}{TEXTURE}"
    with_ = rf"{T}(?:ed|s)?\s+with\s+(?:a|an|the)?\s*{GAP}{TEXTURE}"

    for pat in (coord, coord2, subj_with, obj_with):
        m = re.search(pat, c)
        if m and _has_head(m.group(1)):
            return "coord_heterog", m.group(0).strip()
    m = re.search(into, c)
    if m:
        return "into_texture", m.group(0).strip()
    m = re.search(with_, c)
    if m:
        return "with_texture", m.group(0).strip()
    for pat in direct:
        m = re.search(pat, c)
        if m:
            return "direct", m.group(0).strip()
    if not re.search(TEXTURE, c):
        return "absent", "no texture token in window"
    return "unresolved", ""


# --------------------------------------------------------------------------
# 3. Extra-musical domain lexicon
# --------------------------------------------------------------------------

def classify_domain(context: str):
    c = norm(context)
    for dom, cues in DOMAIN_LEXICON.items():
        for cue in cues:
            if cue in c:
                return dom, cue
    return None, None


# --------------------------------------------------------------------------
# Annotate / report / write
# --------------------------------------------------------------------------

def annotate_qa(df: pd.DataFrame) -> pd.DataFrame:
    """Add QA suggestion columns (never overwrites pipeline fields)."""
    out = df.copy()
    if len(out) == 0:
        for col in QA_COLS:
            out[col] = pd.Series(dtype=object)
        return out

    rel = out.apply(
        lambda r: classify_relation(r.get("contexto"), r.get("matched_form")),
        axis=1,
    )
    out["qa_relacao"] = [x[0] for x in rel]
    out["qa_relacao_evidencia"] = [x[1] for x in rel]
    dom = out["contexto"].map(classify_domain)
    out["qa_dominio_extra"] = [x[0] for x in dom]
    out["qa_dominio_cue"] = [x[1] for x in dom]
    return out


def build_report(df: pd.DataFrame, new_groups, tracked,
                 jaccard: float, containment: float) -> dict:
    cross = [g for g in new_groups
             if "doc_id" in df.columns and df.loc[g, "doc_id"].nunique() > 1]
    within = [g for g in new_groups
              if "doc_id" not in df.columns
              or df.loc[g, "doc_id"].nunique() == 1]
    report: dict[str, Any] = {
        "duplicates": {
            "tracked_by_pipeline": len(tracked),
            "new_groups": len(new_groups),
            "cross_document_groups": len(cross),
            "within_document_groups": len(within),
            "rows_involved": sum(map(len, new_groups)),
            "jaccard": jaccard,
            "containment": containment,
        },
        "relations": df["qa_relacao"].value_counts(dropna=False).to_dict(),
        "domain_hits": (
            df["qa_dominio_extra"].dropna().value_counts().to_dict()
        ),
        "nuclear_true_but_suspect": int(len(
            df[(df["qa_relacao"].isin(["absent", "coord_heterog"]))
               & df["nuclear"].map(_truthy)]
        )) if "nuclear" in df.columns else 0,
    }
    return report


def _header_map(ws: Worksheet) -> dict[Any, int]:
    return {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}


def _ensure_qa_columns(ws: Worksheet, header: dict[Any, int]) -> dict[str, int]:
    """Reuse existing QA columns or append them once (idempotent)."""
    cols: dict[str, int] = {}
    base = ws.max_column
    for name in QA_COLS:
        if name in header and header[name]:
            cols[name] = int(header[name])
        else:
            base += 1
            ws.cell(1, base).value = name
            cols[name] = base
            header[name] = base
    return cols


def apply_to_workbook(
    xlsx: Path,
    df: pd.DataFrame,
    new_groups: list[list],
    *,
    demote_duplicates: bool = False,
    outfile: Path | None = None,
) -> tuple[Path, int]:
    """Write tags, QA columns and colour mirrors. Returns (path, demoted)."""
    wb = load_workbook(xlsx)
    if SHEET not in wb.sheetnames:
        raise SystemExit(f"Falta a folha {SHEET} em {xlsx}")
    ws = wb[SHEET]
    header = _header_map(ws)

    for required in ("contexto", "candidato_duplicado", "nuclear"):
        if required not in header:
            raise SystemExit(f"Coluna em falta em {SHEET}: {required}")

    col_ctx = header["contexto"]
    col_nuc = header["nuclear"]
    col_dup = header["candidato_duplicado"]
    qa_cols = _ensure_qa_columns(ws, header)

    # Excel row = pandas positional index + 2 (header on row 1)
    pos = {idx: i for i, idx in enumerate(df.index)}

    for r in df.index:
        xl = pos[r] + 2
        for name in QA_COLS:
            v = df.loc[r, name]
            ws.cell(xl, qa_cols[name]).value = None if pd.isna(v) else v

    demoted = 0
    for gi, g in enumerate(new_groups, 1):
        tag = f"quase_duplicado:QD{gi:04d}"
        for r in g:
            xl = pos[r] + 2
            cur = ws.cell(xl, col_dup).value
            if not cur:
                ws.cell(xl, col_dup).value = tag
            elif "quase_duplicado" not in str(cur):
                ws.cell(xl, col_dup).value = f"{cur}; {tag}"
            ws.cell(xl, col_dup).fill = FILL_DUP
            if ws.cell(xl, col_ctx).fill.patternType != "solid":
                ws.cell(xl, col_ctx).fill = FILL_DUP
        if demote_duplicates and "nuclear" in df.columns:
            true_rows = [r for r in g if _truthy(df.loc[r, "nuclear"])]
            if true_rows:
                keep = max(true_rows, key=lambda r: keeper_score(df, r))
                for r in true_rows:
                    if r != keep:
                        ws.cell(pos[r] + 2, col_nuc).value = False
                        ws.cell(pos[r] + 2, col_nuc).fill = FILL_DUP
                        demoted += 1

    # colour mirrors for suspect relations / domains (suggestions only)
    for r in df.index:
        xl = pos[r] + 2
        if df.loc[r, "qa_dominio_extra"] and (
                ws.cell(xl, col_ctx).fill.patternType != "solid"):
            ws.cell(xl, col_ctx).fill = FILL_DOM
        elif (df.loc[r, "qa_relacao"] in (
                "coord_heterog", "into_texture", "with_texture")
              and _truthy(df.loc[r, "nuclear"])
              and ws.cell(xl, col_ctx).fill.patternType != "solid"):
            ws.cell(xl, col_ctx).fill = FILL_REV

    out = Path(outfile) if outfile else Path(xlsx)
    wb.save(out)
    return out, demoted


def run_qa(
    xlsx: Path,
    *,
    apply: bool = False,
    demote_duplicates: bool = False,
    outfile: Path | None = None,
    jaccard: float = 0.45,
    containment: float = 0.80,
) -> dict:
    """Load ``8_Concordancia``, run the three QA modules, optionally write."""
    df0 = pd.read_excel(xlsx, sheet_name=SHEET)
    for col in ("contexto", "matched_form"):
        if col not in df0.columns:
            raise SystemExit(f"Coluna em falta em {SHEET}: {col}")

    new_groups, tracked = find_duplicate_groups(
        df0, jaccard=jaccard, containment=containment)
    df = annotate_qa(df0)
    report = build_report(df, new_groups, tracked, jaccard, containment)

    if apply:
        out, demoted = apply_to_workbook(
            xlsx, df, new_groups,
            demote_duplicates=demote_duplicates,
            outfile=outfile,
        )
        report["written"] = str(out)
        report["duplicates_demoted"] = demoted
        print(f"written: {out}  (duplicates demoted: {demoted})",
              file=sys.stderr)
    return report


def main(argv=None) -> int:
    ap = argparse.ArgumentParser(
        description="QA pass over 8_Concordancia (duplicates / relations / domain)",
    )
    ap.add_argument("infile", nargs="?", type=Path,
                    help="Excel de entrada (alternativa a --xlsx)")
    ap.add_argument("--xlsx", type=Path, help="Excel com folha 8_Concordancia")
    ap.add_argument("-o", "--outfile", "--saida", dest="outfile", type=Path,
                    help="Excel de saída (omissão: sobrescreve a entrada)")
    ap.add_argument("--apply", action="store_true",
                    help="write tags, columns and colours (default: report only)")
    ap.add_argument("--demote-duplicates", action="store_true",
                    help="set nuclear=FALSE on redundant copies (one keeper per group)")
    ap.add_argument("--jaccard", type=float, default=0.45)
    ap.add_argument("--containment", type=float, default=0.80)
    args = ap.parse_args(argv)

    xlsx = args.xlsx or args.infile
    if xlsx is None:
        ap.error("indique o Excel: --xlsx IN.xlsx  ou  positional IN.xlsx")
    if not xlsx.is_file():
        raise SystemExit(f"Ficheiro inexistente: {xlsx}")
    if args.apply is False and args.demote_duplicates:
        raise SystemExit("--demote-duplicates requer --apply")

    report = run_qa(
        xlsx,
        apply=args.apply,
        demote_duplicates=args.demote_duplicates,
        outfile=args.outfile,
        jaccard=args.jaccard,
        containment=args.containment,
    )
    print(json.dumps(report, indent=2, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
