#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
textura_gui.py — interface TEXTURA (pesquisa → revisão → análise → apêndice)
===========================================================================

Pipeline:
  1. Pesquisar / Extrair NEAR → Excel
  2. Rever no Excel (relacao_sintactica, nuclear, …)
  3. Analisar Excel… → estatística/gráficos
  4. Apêndice DOCX… → concordância das atribuições genuínas (fase 3)

    python textura_gui.py
"""

from __future__ import annotations

import json
import os
import queue
import re
import subprocess
import sys
import tempfile
import threading
from pathlib import Path
import tkinter as tk
from tkinter import filedialog, messagebox, ttk

AQUI = Path(__file__).resolve().parent
if str(AQUI) not in sys.path:
    sys.path.insert(0, str(AQUI))
MOTOR = AQUI / "textura_search.py"
MOTOR_ANALISE = AQUI / "textura_analise.py"
MOTOR_APENDICE = AQUI / "textura_apendice.py"
MOTOR_DOCTOR = AQUI / "textura_doctor.py"
MOTOR_APA7 = AQUI / "textura_apa7.py"

try:
    from textura_search import DEFAULT_SAIDA
except Exception:  # noqa: BLE001
    DEFAULT_SAIDA = Path(r"C:\Users\lmr20\Desktop\EXCEL_list") / "resultado_pesquisa.xlsx"

try:
    from textura_legendas import (
        PADRAO as LEG_PADRAO,
        carregar_defeito_utilizador,
        guardar_defeito_utilizador,
    )
except Exception:  # noqa: BLE001
    LEG_PADRAO = {
        "rodape": "TEXTURA  ·  pesquisa bibliográfica de termos",
        "sankey": {"titulo": "Sankey: termos → documentos", "subtitulo": "",
                   "eixo_esq": "Termos", "eixo_dir": "Documentos"},
        "nuvem": {"titulo": "Nuvem de palavras", "subtitulo": ""},
        "docs": {"titulo": "Dispersão lexical pelos documentos", "subtitulo": "",
                 "xlabel": "Ocorrências"},
        "formas": {"titulo": "Termos associados", "subtitulo": "",
                   "xlabel": "Ocorrências"},
        "near": {"titulo": "Distâncias NEAR", "subtitulo": "",
                 "xlabel": "Distância (tokens)", "ylabel": "Nº de pares",
                 "mediana": "Mediana", "media": "Média"},
    }

    def carregar_defeito_utilizador():
        return LEG_PADRAO

    def guardar_defeito_utilizador(legendas):
        return Path(".") / "legendas_defeito.json"

EXEMPLOS = [
    "music* NEAR/4 texture*",
    "(uniform* OR constant*) AND NOT varied*",
    "dense* OR densit*",
    "texture* AND (orchestr* OR string*)",
    "timbre* NOR texture*",
]

AJUDA = (
    "1) Pesquisar → Excel   2) Rever   3) Analisar Excel…   "
    "4) Apêndice DOCX… (fase 3)   ·   AND OR NOR NOT NEAR/4 * ?"
)


class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("TEXTURA — pesquisa → revisão → análise → apêndice")
        self.geometry("920x820")
        self.minsize(780, 640)
        self.configure(bg="#f4f2ee")

        self.proc: subprocess.Popen | None = None
        self.fila: queue.Queue[str] = queue.Queue()

        self.v_xlsx = tk.StringVar(value="")  # escolher na GUI — sem default
        self.v_consulta = tk.StringVar()
        self.v_saida = tk.StringVar(value=str(DEFAULT_SAIDA))
        self.v_estado = tk.StringVar(value="Pronto")
        self.v_limite = tk.StringVar(value="")
        self.v_near_extra = tk.BooleanVar(value=False)
        self.v_mesma_frase = tk.BooleanVar(value=True)
        self.v_sintaxe = tk.BooleanVar(value=True)
        self.v_folha = tk.StringVar(value="Neighbor Contexts")
        # Omissão: todas as nucleares (nao passa --relacao).
        # Restringir e opcional e COMPOE-SE com nuclear=True (nunca o substitui).
        self.v_rel_todas_nuc = tk.BooleanVar(value=True)
        self.v_rel_atr = tk.BooleanVar(value=False)
        self.v_rel_pred = tk.BooleanVar(value=False)
        self.v_rel_pred_sec = tk.BooleanVar(value=False)
        self.v_rel_gen = tk.BooleanVar(value=False)
        self.v_rel_comp = tk.BooleanVar(value=False)
        self.v_rel_adv = tk.BooleanVar(value=False)
        self.v_rel_indet = tk.BooleanVar(value=False)

        # legendas editáveis — carrega defeitos do utilizador se existirem
        try:
            leg0 = carregar_defeito_utilizador()
        except Exception:  # noqa: BLE001
            leg0 = LEG_PADRAO
        self.v_leg_rodape = tk.StringVar(
            value=str(leg0.get("rodape") or LEG_PADRAO["rodape"])
        )
        self.v_leg = {}
        for chave in ("sankey", "nuvem", "docs", "formas", "near"):
            base = dict(LEG_PADRAO.get(chave) or {})
            base.update(leg0.get(chave) or {})
            self.v_leg[chave] = {
                k: tk.StringVar(value=str(v)) for k, v in base.items()
            }

        self._estilo()
        self._constroi()
        self.after(100, self._drena)

    def _estilo(self):
        sty = ttk.Style(self)
        if "vista" in sty.theme_names():
            sty.theme_use("vista")
        sty.configure("Title.TLabel", font=("Segoe UI Semibold", 16),
                      background="#f4f2ee")
        sty.configure("Hint.TLabel", font=("Segoe UI", 9), foreground="#555",
                      background="#f4f2ee")
        sty.configure("Card.TFrame", background="#f4f2ee")
        sty.configure("Big.TButton", font=("Segoe UI Semibold", 11), padding=8)

    def _constroi(self):
        root = ttk.Frame(self, padding=18, style="Card.TFrame")
        root.pack(fill="both", expand=True)

        ttk.Label(root, text="TEXTURA", style="Title.TLabel").pack(anchor="w")
        ttk.Label(root, text=AJUDA, style="Hint.TLabel", wraplength=860).pack(
            anchor="w", pady=(2, 14))

        # —— consulta (Text: aceita colar Ctrl+V / botão direito) ——
        box = ttk.LabelFrame(
            root, text="Consulta  (Ctrl+V ou botão direito para colar · Ctrl+Enter para pesquisar)",
            padding=12)
        box.pack(fill="x")
        self.txt_consulta = tk.Text(
            box, height=3, wrap="word", font=("Segoe UI", 12),
            relief="solid", borderwidth=1, padx=6, pady=4,
            undo=True, maxundo=-1)
        self.txt_consulta.pack(fill="x")
        self.txt_consulta.focus_set()
        self._activa_colar(self.txt_consulta)
        self.txt_consulta.bind("<Control-Return>", lambda _e: self._corre())
        self.txt_consulta.bind("<Control-KP_Enter>", lambda _e: self._corre())

        ex = ttk.Frame(box)
        ex.pack(fill="x", pady=(8, 0))
        ttk.Label(ex, text="Exemplos:", style="Hint.TLabel").pack(side="left")
        for e in EXEMPLOS[:3]:
            ttk.Button(ex, text=e, style="Toolbutton",
                       command=lambda s=e: self._define_consulta(s)).pack(
                side="left", padx=3)
        ttk.Button(ex, text="Colar", style="Toolbutton",
                   command=lambda: self._colar(self.txt_consulta)).pack(
            side="right", padx=3)
        ttk.Button(ex, text="Limpar", style="Toolbutton",
                   command=lambda: self._define_consulta("")).pack(side="right")

        filt = ttk.Frame(box)
        filt.pack(fill="x", pady=(10, 0))
        ttk.Checkbutton(
            filt, variable=self.v_mesma_frase,
            text="Recusar termos separados por ponto (frases diferentes)"
        ).pack(anchor="w")
        ttk.Checkbutton(
            filt, variable=self.v_sintaxe,
            text="Exigir que um termo defina/caracterize o outro "
                 "(atributivo, predicativo, «of/as/with», etc.)"
        ).pack(anchor="w", pady=(2, 0))
        ttk.Label(
            filt, style="Hint.TLabel", wraplength=740,
            text="Duplicados da coluna O (contexto) são sempre removidos "
                 "antes da análise — a matriz original não é alterada."
        ).pack(anchor="w", pady=(4, 0))

        # —— ficheiros ——
        arq = ttk.LabelFrame(root, text="Ficheiros", padding=12)
        arq.pack(fill="x", pady=(12, 0))
        arq.columnconfigure(1, weight=1)

        ttk.Label(arq, text="Matriz").grid(row=0, column=0, sticky="w", padx=(0, 8))
        ttk.Entry(arq, textvariable=self.v_xlsx).grid(row=0, column=1, sticky="ew")
        ttk.Button(arq, text="…", width=3,
                   command=self._escolhe_matriz).grid(row=0, column=2, padx=(6, 0))
        ttk.Label(arq, text="Escolha o .xlsx (sem ficheiro pré-definido)",
                  style="Hint.TLabel").grid(row=1, column=1, sticky="w", pady=(2, 0))

        ttk.Label(arq, text="Guardar em").grid(row=2, column=0, sticky="w",
                                               padx=(0, 8), pady=(8, 0))
        ttk.Entry(arq, textvariable=self.v_saida).grid(row=2, column=1, sticky="ew",
                                                       pady=(8, 0))
        ttk.Button(arq, text="…", width=3,
                   command=self._escolhe_saida).grid(row=2, column=2, padx=(6, 0),
                                                     pady=(8, 0))

        # —— acções ——
        ac = ttk.Frame(root)
        ac.pack(fill="x", pady=(14, 0))
        self.btn_corre = ttk.Button(ac, text="1. Pesquisar", style="Big.TButton",
                                    command=self._corre)
        self.btn_corre.pack(side="left")
        self.btn_analisa = ttk.Button(
            ac, text="3. Analisar Excel…", style="Big.TButton",
            command=self._analisar_excel)
        self.btn_analisa.pack(side="left", padx=(8, 0))
        self.btn_apendice = ttk.Button(
            ac, text="4. Apêndice DOCX…", style="Big.TButton",
            command=self._apendice_docx)
        self.btn_apendice.pack(side="left", padx=(8, 0))
        self.btn_para = ttk.Button(ac, text="Parar", command=self._para,
                                   state="disabled")
        self.btn_para.pack(side="left", padx=8)
        self.btn_abre = ttk.Button(ac, text="Abrir resultado",
                                   command=self._abre_saida, state="disabled")
        self.btn_abre.pack(side="left")
        self.barra = ttk.Progressbar(ac, mode="indeterminate", length=140)
        self.barra.pack(side="left", padx=14)
        ttk.Label(ac, textvariable=self.v_estado, style="Hint.TLabel").pack(
            side="left")
        ttk.Label(
            root, style="Hint.TLabel", wraplength=860,
            text="2. Rever: abra o Excel, edite relacao_sintactica / nuclear. "
                 "3. Analisar → estatística. "
                 "4. Apêndice DOCX → concordância legível (fase 3)."
        ).pack(anchor="w", pady=(6, 0))

        # —— utilitários aditivos (não alteram o fluxo 1–4) ——
        util = ttk.Frame(root)
        util.pack(fill="x", pady=(8, 0))
        ttk.Button(
            util, text="Doctor (checklist…)", style="Toolbutton",
            command=self._doctor_excel,
        ).pack(side="left")
        ttk.Button(
            util, text="APA7 catálogo…", style="Toolbutton",
            command=self._apa7_catalogo,
        ).pack(side="left", padx=(8, 0))
        ttk.Label(
            util, style="Hint.TLabel",
            text="opcional · antes da análise / do apêndice",
        ).pack(side="left", padx=(10, 0))

        # —— legendas (recolhido, mas bem visível) ——
        self.leg_aberto = tk.BooleanVar(value=False)
        cab_leg = ttk.Frame(root)
        cab_leg.pack(fill="x", pady=(12, 0))
        self.btn_leg = ttk.Button(
            cab_leg, text="▸ Legendas dos gráficos (editar títulos · PT)",
            style="Toolbutton", command=self._toggle_leg)
        self.btn_leg.pack(anchor="w")

        self.frm_leg = ttk.LabelFrame(
            root, text="Legendas (aplicadas aos PNG e HTML na próxima pesquisa)",
            padding=10)
        self.frm_leg.columnconfigure(1, weight=1)
        self.frm_leg.columnconfigure(3, weight=1)
        row = 0
        ttk.Label(self.frm_leg, text="Rodapé comum").grid(
            row=row, column=0, sticky="w", padx=4, pady=2)
        ttk.Entry(self.frm_leg, textvariable=self.v_leg_rodape).grid(
            row=row, column=1, columnspan=2, sticky="ew", padx=4, pady=2)
        btns_leg = ttk.Frame(self.frm_leg)
        btns_leg.grid(row=row, column=3, sticky="e", padx=4, pady=2)
        ttk.Button(
            btns_leg, text="Guardar como defeito", style="Toolbutton",
            command=self._guarda_legendas_defeito,
        ).pack(side="right")
        ttk.Button(
            btns_leg, text="Restaurar PT", style="Toolbutton",
            command=self._restaura_legendas,
        ).pack(side="right", padx=(0, 6))
        row += 1
        for chave, campo, defeito in (
            ("docs", "xlabel", "Ocorrências"),
            ("formas", "xlabel", "Ocorrências"),
            ("near", "xlabel", "Distância (tokens)"),
            ("near", "ylabel", "Nº de pares"),
            ("near", "mediana", "Mediana"),
            ("near", "media", "Média"),
        ):
            if campo not in self.v_leg[chave]:
                self.v_leg[chave][campo] = tk.StringVar(
                    value=str(LEG_PADRAO.get(chave, {}).get(campo, defeito)))
        campos = [
            ("Sankey · título", self.v_leg["sankey"]["titulo"]),
            ("Sankey · subtítulo", self.v_leg["sankey"]["subtitulo"]),
            ("Sankey · eixo esq.", self.v_leg["sankey"]["eixo_esq"]),
            ("Sankey · eixo dir.", self.v_leg["sankey"]["eixo_dir"]),
            ("Nuvem · título", self.v_leg["nuvem"]["titulo"]),
            ("Nuvem · subtítulo", self.v_leg["nuvem"]["subtitulo"]),
            ("Docs · título", self.v_leg["docs"]["titulo"]),
            ("Docs · subtítulo", self.v_leg["docs"]["subtitulo"]),
            ("Docs · eixo X", self.v_leg["docs"]["xlabel"]),
            ("Formas · título", self.v_leg["formas"]["titulo"]),
            ("Formas · subtítulo", self.v_leg["formas"]["subtitulo"]),
            ("Formas · eixo X", self.v_leg["formas"]["xlabel"]),
            ("NEAR · título", self.v_leg["near"]["titulo"]),
            ("NEAR · subtítulo", self.v_leg["near"]["subtitulo"]),
            ("NEAR · eixo X", self.v_leg["near"]["xlabel"]),
            ("NEAR · eixo Y", self.v_leg["near"]["ylabel"]),
        ]
        for i, (rot, var) in enumerate(campos):
            r, c = row + i // 2, (i % 2) * 2
            ttk.Label(self.frm_leg, text=rot).grid(
                row=r, column=c, sticky="w", padx=4, pady=2)
            ttk.Entry(self.frm_leg, textvariable=var).grid(
                row=r, column=c + 1, sticky="ew", padx=4, pady=2)
        ttk.Label(
            self.frm_leg, style="Hint.TLabel", wraplength=740,
            text="Os títulos entram nos PNG e nos HTML (pasta *_figs). "
                 "Nos HTML Plotly pode ainda editar no browser (modebar)."
        ).grid(row=row + (len(campos) + 1) // 2 + 1, column=0, columnspan=4,
               sticky="w", padx=4, pady=(8, 0))

        # —— avançado (recolhido) ——
        self.adv_aberto = tk.BooleanVar(value=False)
        cab_adv = ttk.Frame(root)
        cab_adv.pack(fill="x", pady=(8, 0))
        self.btn_adv = ttk.Button(cab_adv, text="▸ Opções avançadas",
                                  style="Toolbutton", command=self._toggle_adv)
        self.btn_adv.pack(anchor="w")

        self.frm_adv = ttk.Frame(root, padding=(8, 6))
        ttk.Label(self.frm_adv, text="Folha").grid(row=0, column=0, sticky="w")
        ttk.Entry(self.frm_adv, textvariable=self.v_folha, width=28).grid(
            row=0, column=1, sticky="w", padx=8)
        ttk.Label(self.frm_adv, text="Limite de linhas (teste)").grid(
            row=0, column=2, sticky="w", padx=(16, 0))
        ttk.Entry(self.frm_adv, textvariable=self.v_limite, width=10).grid(
            row=0, column=3, sticky="w", padx=8)
        ttk.Checkbutton(
            self.frm_adv, variable=self.v_near_extra,
            text="Após pesquisar, extrair também concordância NEAR "
                 "(Excel para revisão — sem estatística/gráficos)"
        ).grid(row=1, column=0, columnspan=4, sticky="w", pady=(8, 0))
        ttk.Label(
            self.frm_adv, style="Hint.TLabel", wraplength=700,
            text="A pesquisa e a extracção NEAR só geram Excel. "
                 "Estatística e gráficos ficam em «Analisar Excel…»."
        ).grid(row=2, column=0, columnspan=4, sticky="w", pady=(6, 0))

        # —— log ——
        logf = ttk.LabelFrame(root, text="Registo", padding=8)
        logf.pack(fill="both", expand=True, pady=(12, 0))
        mono = ("Consolas", 9) if os.name == "nt" else ("Menlo", 10)
        self.txt = tk.Text(logf, wrap="word", height=12, state="disabled",
                           background="#faf9f7", font=mono, relief="flat")
        sy = ttk.Scrollbar(logf, orient="vertical", command=self.txt.yview)
        self.txt.configure(yscrollcommand=sy.set)
        self.txt.pack(side="left", fill="both", expand=True)
        sy.pack(side="right", fill="y")

        self._log(
            "Pipeline: 1 Pesquisar → 2 Rever Excel → 3 Analisar → "
            "4 Apêndice DOCX (fase 3)")
        if self.v_xlsx.get():
            self._log(f"Matriz: {self.v_xlsx.get()}")
        else:
            self._log("Escolha a matriz KWIC (…) antes de pesquisar.")

    def _toggle_adv(self):
        if self.adv_aberto.get():
            self.frm_adv.pack_forget()
            self.btn_adv.config(text="▸ Opções avançadas")
            self.adv_aberto.set(False)
        else:
            self.frm_adv.pack(fill="x", after=self.btn_adv.master)
            self.btn_adv.config(text="▾ Opções avançadas")
            self.adv_aberto.set(True)

    def _toggle_leg(self):
        if self.leg_aberto.get():
            self.frm_leg.pack_forget()
            self.btn_leg.config(
                text="▸ Legendas dos gráficos (editar títulos · PT)")
            self.leg_aberto.set(False)
        else:
            self.frm_leg.pack(fill="x", after=self.btn_leg.master, pady=(4, 0))
            self.btn_leg.config(
                text="▾ Legendas dos gráficos (editar títulos · PT)")
            self.leg_aberto.set(True)

    def _restaura_legendas(self):
        self.v_leg_rodape.set(LEG_PADRAO["rodape"])
        for chave, campos in LEG_PADRAO.items():
            if chave == "rodape" or not isinstance(campos, dict):
                continue
            for k, v in campos.items():
                if k not in self.v_leg.get(chave, {}):
                    self.v_leg.setdefault(chave, {})[k] = tk.StringVar()
                self.v_leg[chave][k].set(str(v))

    def _guarda_legendas_defeito(self) -> Path | None:
        """Persiste as legendas actuais para a próxima sessão."""
        try:
            path = guardar_defeito_utilizador(self._serializa_legendas())
            self._log(f"Legendas guardadas como defeito: {path}")
            messagebox.showinfo(
                "Legendas",
                f"Defeitos guardados.\nNa próxima abertura da GUI "
                f"estes títulos serão carregados automaticamente.\n\n{path}",
            )
            return path
        except Exception as exc:  # noqa: BLE001
            messagebox.showerror("Legendas", str(exc))
            return None

    def _serializa_legendas(self) -> dict:
        out = {"rodape": self.v_leg_rodape.get().strip()}
        for chave, campos in self.v_leg.items():
            out[chave] = {k: v.get().strip() for k, v in campos.items()}
        return out

    def _persistir_legendas_silencioso(self) -> None:
        """Auto-grava defeitos ao analisar/pesquisar (sem diálogo)."""
        try:
            path = guardar_defeito_utilizador(self._serializa_legendas())
            self._log(f"Legendas (defeito) actualizadas: {path.name}")
        except Exception as exc:  # noqa: BLE001
            self._log(f"Aviso: não foi possível gravar legendas defeito: {exc}")

    # ------------------------------------------------------------------ colar
    def _define_consulta(self, texto: str) -> None:
        self.txt_consulta.delete("1.0", "end")
        if texto:
            self.txt_consulta.insert("1.0", texto)
        self.v_consulta.set(texto)

    def _consulta_actual(self) -> str:
        # uma só linha lógica: quebras → espaço (útil ao colar de Word/PDF)
        bruto = self.txt_consulta.get("1.0", "end-1c")
        return " ".join(bruto.split())

    def _colar(self, widget: tk.Text) -> None:
        try:
            texto = self.clipboard_get()
        except tk.TclError:
            return
        try:
            widget.delete("sel.first", "sel.last")
        except tk.TclError:
            pass
        widget.insert("insert", texto)
        return "break"

    def _activa_colar(self, widget: tk.Text) -> None:
        """Garante Ctrl+V / Shift+Ins / menu de contexto no Windows."""
        def _paste(event=None):
            self._colar(widget)
            return "break"

        for seq in ("<<Paste>>", "<Control-v>", "<Control-V>",
                    "<Shift-Insert>", "<Control-Insert>"):
            widget.bind(seq, _paste)

        menu = tk.Menu(widget, tearoff=0)
        menu.add_command(label="Colar", command=lambda: self._colar(widget))
        menu.add_command(label="Copiar",
                         command=lambda: widget.event_generate("<<Copy>>"))
        menu.add_command(label="Cortar",
                         command=lambda: widget.event_generate("<<Cut>>"))
        menu.add_separator()
        menu.add_command(label="Seleccionar tudo",
                         command=lambda: widget.tag_add("sel", "1.0", "end-1c"))

        def _menu(event):
            try:
                menu.tk_popup(event.x_root, event.y_root)
            finally:
                menu.grab_release()
            return "break"

        widget.bind("<Button-3>", _menu)       # Windows / Linux
        widget.bind("<Button-2>", _menu)       # alguns ratos
        if sys.platform == "darwin":
            widget.bind("<Button-2>", _menu)
            widget.bind("<Control-Button-1>", _menu)

    def _escolhe_matriz(self):
        c = filedialog.askopenfilename(
            title="Matriz KWIC",
            filetypes=[("Excel", "*.xlsx *.xlsm"), ("Todos", "*.*")],
            initialdir=str(Path(self.v_xlsx.get()).parent)
            if self.v_xlsx.get() else str(Path.home()))
        if c:
            self.v_xlsx.set(c)
            self.v_saida.set(str(Path(c).with_name("resultado_pesquisa.xlsx")))

    def _escolhe_saida(self):
        c = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
            initialfile=Path(self.v_saida.get()).name)
        if c:
            self.v_saida.set(c)

    def _abre_saida(self):
        p = Path(self.v_saida.get())
        if not p.exists():
            messagebox.showwarning("Abrir", "O ficheiro ainda não existe.")
            return
        if os.name == "nt":
            os.startfile(p)  # noqa: S606
        elif sys.platform == "darwin":
            subprocess.Popen(["open", str(p)])
        else:
            subprocess.Popen(["xdg-open", str(p)])

    def _corre(self):
        if self.proc is not None:
            return
        if not MOTOR.exists():
            messagebox.showerror("Erro", f"Motor em falta:\n{MOTOR}")
            return
        xlsx = self.v_xlsx.get().strip()
        consulta = self._consulta_actual()
        self.v_consulta.set(consulta)
        if not xlsx or not Path(xlsx).exists():
            messagebox.showwarning("Matriz", "Indique um ficheiro Excel válido.")
            return
        if not consulta:
            messagebox.showwarning("Consulta", "Escreva ou cole uma consulta de pesquisa.")
            return

        # gravar legendas editadas para o motor
        leg_path = Path(tempfile.gettempdir()) / "_textura_legendas.json"
        leg_path.write_text(
            json.dumps(self._serializa_legendas(), ensure_ascii=False, indent=2),
            encoding="utf-8")
        self._persistir_legendas_silencioso()

        cmd = [sys.executable, str(MOTOR),
               "--xlsx", xlsx,
               "--folha", self.v_folha.get().strip() or "Neighbor Contexts",
               "--consulta", consulta,
               "--saida", self.v_saida.get().strip(),
               "--legendas", str(leg_path)]
        if self.v_limite.get().strip().isdigit():
            cmd += ["--limite", self.v_limite.get().strip()]
        if not self.v_mesma_frase.get():
            cmd += ["--permitir-outra-frase"]
        if not self.v_sintaxe.get():
            cmd += ["--sem-sintaxe"]
        if self.v_near_extra.get():
            cmd += ["--extrair-near"]

        self._log("\n" + "-" * 60)
        self._log(" ".join(f'"{a}"' if " " in a else a for a in cmd))
        self._arranca_cmd(cmd, "A pesquisar…")

    def _analisar_excel(self):
        if self.proc is not None:
            return
        if not MOTOR_ANALISE.exists():
            messagebox.showerror("Erro", f"Motor em falta:\n{MOTOR_ANALISE}")
            return
        dlg = tk.Toplevel(self)
        dlg.title("Analisar Excel revisto")
        dlg.transient(self)
        dlg.grab_set()
        dlg.geometry("620x480")
        frm = ttk.Frame(dlg, padding=14)
        frm.pack(fill="both", expand=True)
        frm.columnconfigure(1, weight=1)

        v_in = tk.StringVar(value="")
        v_out = tk.StringVar(value="")
        v_plano = tk.StringVar(value="")
        v_kappa = tk.StringVar(value="")
        v_dedupe = tk.StringVar(value="contexto")
        ttk.Label(frm, text="Excel revisto").grid(row=0, column=0, sticky="w")
        ttk.Entry(frm, textvariable=v_in).grid(row=0, column=1, sticky="ew",
                                               padx=6)
        ttk.Button(
            frm, text="…", width=3,
            command=lambda: self._escolhe_ficheiro(
                v_in, v_out, abrir=True)
        ).grid(row=0, column=2)

        ttk.Label(frm, text="Guardar análise em").grid(
            row=1, column=0, sticky="w", pady=(8, 0))
        ttk.Entry(frm, textvariable=v_out).grid(
            row=1, column=1, sticky="ew", padx=6, pady=(8, 0))
        ttk.Button(
            frm, text="…", width=3,
            command=lambda: self._escolhe_ficheiro(
                v_out, v_out, abrir=False)
        ).grid(row=1, column=2, pady=(8, 0))

        ttk.Label(frm, text="Desduplicação primária").grid(
            row=2, column=0, sticky="w", pady=(8, 0))
        ttk.Combobox(
            frm, textvariable=v_dedupe, state="readonly",
            values=("candidatos", "nenhuma", "obra_termo", "contexto"),
            width=18,
        ).grid(row=2, column=1, sticky="w", padx=6, pady=(8, 0))

        ttk.Label(frm, text="Plano a priori (YAML)").grid(
            row=3, column=0, sticky="w", pady=(8, 0))
        ttk.Entry(frm, textvariable=v_plano).grid(
            row=3, column=1, sticky="ew", padx=6, pady=(8, 0))
        ttk.Button(
            frm, text="…", width=3,
            command=lambda: v_plano.set(
                filedialog.askopenfilename(
                    title="plano_analitico.yaml",
                    filetypes=[("YAML", "*.yaml;*.yml"), ("Todos", "*.*")])
                or v_plano.get()),
        ).grid(row=3, column=2, pady=(8, 0))

        ttk.Label(frm, text="κ cego (Excel/JSON)").grid(
            row=4, column=0, sticky="w", pady=(8, 0))
        ttk.Entry(frm, textvariable=v_kappa).grid(
            row=4, column=1, sticky="ew", padx=6, pady=(8, 0))
        ttk.Button(
            frm, text="…", width=3,
            command=lambda: v_kappa.set(
                filedialog.askopenfilename(
                    title="Resultado --avaliar-kappa",
                    filetypes=[("Excel/JSON", "*.xlsx;*.json"),
                               ("Todos", "*.*")])
                or v_kappa.get()),
        ).grid(row=4, column=2, pady=(8, 0))

        ttk.Label(frm, text="Elegibilidade (nuclear=False e absoluto):").grid(
            row=5, column=0, columnspan=3, sticky="w", pady=(14, 4))
        op0 = ttk.Frame(frm)
        op0.grid(row=6, column=0, columnspan=3, sticky="w")

        def _toggle_todas():
            if self.v_rel_todas_nuc.get():
                for v in (self.v_rel_atr, self.v_rel_pred, self.v_rel_pred_sec,
                          self.v_rel_gen, self.v_rel_comp, self.v_rel_adv,
                          self.v_rel_indet):
                    v.set(False)

        def _toggle_subset():
            if any(v.get() for v in (
                    self.v_rel_atr, self.v_rel_pred, self.v_rel_pred_sec,
                    self.v_rel_gen, self.v_rel_comp, self.v_rel_adv,
                    self.v_rel_indet)):
                self.v_rel_todas_nuc.set(False)

        ttk.Checkbutton(
            op0, text="todas as nucleares (recomendado)",
            variable=self.v_rel_todas_nuc, command=_toggle_todas,
        ).pack(anchor="w")
        op = ttk.Frame(frm)
        op.grid(row=7, column=0, columnspan=3, sticky="w", pady=(6, 0))
        for txt, var in (
            ("atributiva", self.v_rel_atr),
            ("predicativa", self.v_rel_pred),
            ("pred._secundaria", self.v_rel_pred_sec),
            ("nominal_genitiva", self.v_rel_gen),
            ("nominal_composto", self.v_rel_comp),
            ("adverbial", self.v_rel_adv),
            ("indeterminada", self.v_rel_indet),
        ):
            ttk.Checkbutton(
                op, text=txt, variable=var, command=_toggle_subset,
            ).pack(side="left", padx=(0, 8))

        ttk.Label(
            frm, style="Hint.TLabel", wraplength=580,
            text="Por omissão: nucleares + desduplicação «candidatos» "
                 "(só snippets de contexto exactamente iguais). "
                 "«obra_termo» é só sensibilidade — não use para limpar "
                 "falsos duplicados de títulos."
        ).grid(row=8, column=0, columnspan=3, sticky="w", pady=(12, 0))

        def ok():
            xin = v_in.get().strip()
            if not xin or not Path(xin).exists():
                messagebox.showwarning("Excel", "Escolha o Excel revisto.",
                                       parent=dlg)
                return
            pin = Path(xin)
            # Evitar *_analise.xlsx (saída da fase 2) como entrada
            if "_analise" in pin.stem.lower() or pin.name.lower().endswith(
                "_analise.xlsx"
            ):
                alt = pin.with_name(
                    pin.name.replace("_analise.xlsx", ".xlsx")
                    .replace("_analise.XLSX", ".xlsx")
                )
                if not alt.exists():
                    # …_v2_analise → …_v2.xlsx
                    alt = pin.with_name(
                        re.sub(r"_analise\.xlsx$", ".xlsx", pin.name, flags=re.I)
                    )
                msg = (
                    "Seleccionou um Excel de *análise* (*_analise.xlsx), "
                    "não o Excel de *revisão* (*_near* / *_revisto*).\n\n"
                    "A fase 3 «Analisar» exige 8_Concordancia com "
                    "relacao_sintactica."
                )
                if alt.exists():
                    msg += f"\n\nUsar em vez disso?\n{alt}"
                    if messagebox.askyesno("Excel errado", msg, parent=dlg):
                        xin = str(alt)
                        v_in.set(xin)
                        v_out.set(str(alt.with_name(alt.stem + "_analise.xlsx")))
                    else:
                        return
                else:
                    messagebox.showerror("Excel errado", msg, parent=dlg)
                    return
            rels = []
            if not self.v_rel_todas_nuc.get():
                for flag, nome in (
                    (self.v_rel_atr, "atributiva"),
                    (self.v_rel_pred, "predicativa"),
                    (self.v_rel_pred_sec, "predicativa_secundaria"),
                    (self.v_rel_gen, "nominal_genitiva"),
                    (self.v_rel_comp, "nominal_composto"),
                    (self.v_rel_adv, "adverbial"),
                    (self.v_rel_indet, "indeterminada"),
                ):
                    if flag.get():
                        rels.append(nome)
                if not rels:
                    messagebox.showwarning(
                        "Relações",
                        "Seleccione «todas as nucleares» ou pelo menos "
                        "uma relação.",
                        parent=dlg)
                    return
            xout = v_out.get().strip() or str(
                Path(xin).with_name(Path(xin).stem + "_analise.xlsx"))
            leg_path = Path(tempfile.gettempdir()) / "_textura_legendas.json"
            leg_path.write_text(
                json.dumps(self._serializa_legendas(), ensure_ascii=False,
                           indent=2),
                encoding="utf-8")
            self._persistir_legendas_silencioso()
            cmd = [sys.executable, str(MOTOR_ANALISE),
                   "--xlsx", xin, "--saida", xout,
                   "--legendas", str(leg_path),
                   "--desduplicacao", v_dedupe.get().strip() or "candidatos"]
            if rels:
                cmd += ["--relacao", ",".join(rels)]
            if v_plano.get().strip():
                cmd += ["--plano-a-priori", v_plano.get().strip()]
            if v_kappa.get().strip():
                cmd += ["--kappa-cego", v_kappa.get().strip()]
            self.v_saida.set(xout)
            dlg.destroy()
            self._log("\n" + "-" * 60)
            self._log(" ".join(f'"{a}"' if " " in a else a for a in cmd))
            self._arranca_cmd(cmd, "A analisar…")

        btns = ttk.Frame(frm)
        btns.grid(row=9, column=0, columnspan=3, sticky="e", pady=(16, 0))
        ttk.Button(btns, text="Cancelar", command=dlg.destroy).pack(side="right")
        ttk.Button(btns, text="Analisar", command=ok).pack(side="right",
                                                           padx=(0, 8))

    def _sugestao_excel_apendice(self) -> str:
        """Prefere Excel NEAR/revisto/análise — nunca Results da pesquisa."""
        cand = self.v_saida.get().strip()
        if cand:
            nome = Path(cand).name.lower()
            if any(k in nome for k in ("_near", "revisto", "_analise",
                                       "concordancia", "concordância")):
                if Path(cand).exists():
                    return cand
            # resultado_pesquisa / UNIFORME.xlsx (sem _near) → não sugerir
            if "pesquisa" in nome or nome.endswith("uniforme.xlsx"):
                return ""
            # se tiver 0_Instrucoes, ainda é válido
            try:
                from openpyxl import load_workbook
                wb = load_workbook(cand, read_only=True)
                ok = "0_Instrucoes" in wb.sheetnames and (
                    "8_Concordancia" in wb.sheetnames)
                wb.close()
                if ok:
                    return cand
            except Exception:
                pass
        return ""

    def _doctor_excel(self):
        """Checklist pré-análise (avisos; não altera o ficheiro)."""
        if self.proc is not None:
            return
        if not MOTOR_DOCTOR.exists():
            messagebox.showerror("Erro", f"Motor em falta:\n{MOTOR_DOCTOR}")
            return
        xin = filedialog.askopenfilename(
            title="Excel revisto (8_Concordancia)",
            filetypes=[("Excel", "*.xlsx"), ("Todos", "*.*")],
        )
        if not xin:
            return
        cmd = [sys.executable, str(MOTOR_DOCTOR), "--xlsx", xin]
        self._log("\n" + "-" * 60)
        self._log(" ".join(f'"{a}"' if " " in a else a for a in cmd))
        self._arranca_cmd(cmd, "Doctor…")

    def _apa7_catalogo(self):
        """Gera catálogo APA7 opcional para --refs no apêndice."""
        if self.proc is not None:
            return
        if not MOTOR_APA7.exists():
            messagebox.showerror("Erro", f"Motor em falta:\n{MOTOR_APA7}")
            return
        xin = filedialog.askopenfilename(
            title="Excel com fontes (NEAR / revisto)",
            filetypes=[("Excel", "*.xlsx"), ("Todos", "*.*")],
        )
        if not xin:
            return
        xout = filedialog.asksaveasfilename(
            title="Guardar catálogo APA7",
            defaultextension=".xlsx",
            initialfile=Path(xin).stem + "_refs_apa7.xlsx",
            filetypes=[("Excel", "*.xlsx"), ("CSV", "*.csv"), ("Todos", "*.*")],
        )
        if not xout:
            return
        cmd = [sys.executable, str(MOTOR_APA7),
               "--xlsx", xin, "--saida", xout]
        self.v_saida.set(xout)
        self._log("\n" + "-" * 60)
        self._log(" ".join(f'"{a}"' if " " in a else a for a in cmd))
        self._arranca_cmd(cmd, "APA7…")

    def _apendice_docx(self):
        """Fase 3 — projecção legível (DOCX) das atribuições nucleares."""
        if self.proc is not None:
            return
        if not MOTOR_APENDICE.exists():
            messagebox.showerror("Erro", f"Motor em falta:\n{MOTOR_APENDICE}")
            return
        dlg = tk.Toplevel(self)
        dlg.title("Apêndice de concordância (fase 3)")
        dlg.transient(self)
        dlg.grab_set()
        dlg.geometry("640x380")
        frm = ttk.Frame(dlg, padding=14)
        frm.pack(fill="both", expand=True)
        frm.columnconfigure(1, weight=1)

        v_in = tk.StringVar(value=self._sugestao_excel_apendice())
        v_out = tk.StringVar(value="")
        v_refs = tk.StringVar(value="")
        v_ap = tk.StringVar(value="X")
        v_agrupar = tk.StringVar(value="query_pattern")

        ttk.Label(frm, text="Excel revisto / NEAR").grid(
            row=0, column=0, sticky="w")
        ttk.Entry(frm, textvariable=v_in).grid(
            row=0, column=1, sticky="ew", padx=6)
        ttk.Button(
            frm, text="…", width=3,
            command=lambda: v_in.set(
                filedialog.askopenfilename(
                    title="Excel revisto (*_near*.xlsx / 8_Concordancia)",
                    filetypes=[("Excel", "*.xlsx"), ("Todos", "*.*")])
                or v_in.get()),
        ).grid(row=0, column=2)

        ttk.Label(frm, text="Guardar DOCX em").grid(
            row=1, column=0, sticky="w", pady=(8, 0))
        ttk.Entry(frm, textvariable=v_out).grid(
            row=1, column=1, sticky="ew", padx=6, pady=(8, 0))
        ttk.Button(
            frm, text="…", width=3,
            command=lambda: v_out.set(
                filedialog.asksaveasfilename(
                    title="Apêndice DOCX",
                    defaultextension=".docx",
                    filetypes=[("Word", "*.docx"), ("Todos", "*.*")])
                or v_out.get()),
        ).grid(row=1, column=2, pady=(8, 0))

        ttk.Label(frm, text="Catálogo APA7 (opcional)").grid(
            row=2, column=0, sticky="w", pady=(8, 0))
        ttk.Entry(frm, textvariable=v_refs).grid(
            row=2, column=1, sticky="ew", padx=6, pady=(8, 0))
        ttk.Button(
            frm, text="…", width=3,
            command=lambda: v_refs.set(
                filedialog.askopenfilename(
                    title="Catálogo APA7",
                    filetypes=[
                        ("CSV/TSV/Excel", "*.csv;*.tsv;*.xlsx"),
                        ("Todos", "*.*"),
                    ])
                or v_refs.get()),
        ).grid(row=2, column=2, pady=(8, 0))

        ttk.Label(frm, text="Agrupar por").grid(
            row=3, column=0, sticky="w", pady=(8, 0))
        ttk.Combobox(
            frm, textvariable=v_agrupar, state="readonly",
            values=("query_pattern", "canonical_term", "termo_tipo"),
            width=18,
        ).grid(row=3, column=1, sticky="w", padx=6, pady=(8, 0))

        ttk.Label(frm, text="Nº do apêndice").grid(
            row=4, column=0, sticky="w", pady=(8, 0))
        ttk.Entry(frm, textvariable=v_ap, width=8).grid(
            row=4, column=1, sticky="w", padx=6, pady=(8, 0))

        v_pdf = tk.BooleanVar(value=True)
        ttk.Checkbutton(
            frm, variable=v_pdf,
            text="Localizar página no PDF (recomendado; paralelo; pymupdf). "
                 "Página a seguir ao excerto: (p. N)=obra; (PDF p. N)=folha",
        ).grid(row=5, column=0, columnspan=3, sticky="w", pady=(8, 0))

        ttk.Label(
            frm, style="Hint.TLabel", wraplength=580,
            text="Use o Excel REVISTO da extracção NEAR "
                 "(*_near*.xlsx / *_revisto*.xlsx), NÃO o Results da pesquisa "
                 "nem resultado_pesquisa.xlsx. Só nuclear=TRUE. "
                 "Gera dois DOCX: publicação + «…_links.docx». "
                 "Excertos usam aspas \"…\"; cotos KWIC finais são limpos "
                 "sem remover os termos de pesquisa.",
        ).grid(row=6, column=0, columnspan=3, sticky="w", pady=(12, 0))

        def ok():
            xin = v_in.get().strip()
            if not xin or not Path(xin).exists():
                messagebox.showwarning(
                    "Excel",
                    "Escolha o Excel revisto da extracção NEAR "
                    "(ex.: UNIFORME_near_revisto_LR.xlsx).",
                    parent=dlg)
                return
            nome = Path(xin).name.lower()
            if ("pesquisa" in nome or nome.endswith("uniforme.xlsx")
                    or (not any(k in nome for k in
                                ("_near", "revisto", "_analise"))
                        and "concord" not in nome)):
                if not messagebox.askyesno(
                        "Ficheiro suspeito",
                        "Este caminho parece o Excel da PESQUISA "
                        "(Results), não o NEAR revisto.\n\n"
                        "O apêndice deve usar o ficheiro com "
                        "8_Concordancia após revisão "
                        "(ex.: …_near_revisto….xlsx).\n\n"
                        "Continuar mesmo assim?",
                        parent=dlg):
                    return
            xout = v_out.get().strip() or str(
                Path(xin).with_name("Apendice_Concordancia.docx"))
            cmd = [sys.executable, str(MOTOR_APENDICE),
                   "--xlsx", xin, "--saida", xout,
                   "--agrupar", v_agrupar.get().strip() or "query_pattern",
                   "--apendice", v_ap.get().strip() or "X"]
            if v_refs.get().strip():
                cmd += ["--refs", v_refs.get().strip()]
            # CLI tem --paginas-pdf ligado por omissão; desligar explicitamente.
            cmd.append("--paginas-pdf" if v_pdf.get() else "--no-paginas-pdf")
            self.v_saida.set(xout)
            dlg.destroy()
            self._log("\n" + "-" * 60)
            self._log(" ".join(f'"{a}"' if " " in a else a for a in cmd))
            self._arranca_cmd(cmd, "A gerar apêndice…")

        btns = ttk.Frame(frm)
        btns.grid(row=7, column=0, columnspan=3, sticky="e", pady=(16, 0))
        ttk.Button(btns, text="Cancelar", command=dlg.destroy).pack(
            side="right")
        ttk.Button(btns, text="Gerar DOCX", command=ok).pack(
            side="right", padx=(0, 8))

    def _escolhe_ficheiro(self, var_in: tk.StringVar, var_out: tk.StringVar,
                          *, abrir: bool):
        if abrir:
            c = filedialog.askopenfilename(
                title="Excel revisto (concordância)",
                filetypes=[("Excel", "*.xlsx"), ("Todos", "*.*")],
                initialdir=str(Path(var_in.get()).parent)
                if var_in.get() else str(Path.home()))
            if c:
                var_in.set(c)
                if not var_out.get():
                    var_out.set(str(Path(c).with_name(
                        Path(c).stem + "_analise.xlsx")))
        else:
            c = filedialog.asksaveasfilename(
                title="Guardar análise",
                defaultextension=".xlsx",
                filetypes=[("Excel", "*.xlsx")],
                initialfile=Path(var_out.get()).name if var_out.get()
                else "analise.xlsx")
            if c:
                var_out.set(c)

    def _arranca_cmd(self, cmd: list[str], estado: str):
        self.btn_corre.config(state="disabled")
        self.btn_analisa.config(state="disabled")
        self.btn_para.config(state="normal")
        self.btn_abre.config(state="disabled")
        self.barra.start(12)
        self.v_estado.set(estado)
        threading.Thread(target=self._trabalha, args=(cmd,), daemon=True).start()

    def _trabalha(self, cmd):
        cod = 1
        try:
            env = os.environ.copy()
            # Garantir imports locais (textura_legendas, etc.) mesmo se o
            # processo herdar um PYTHONPATH diferente.
            prev = env.get("PYTHONPATH", "")
            env["PYTHONPATH"] = str(AQUI) + (os.pathsep + prev if prev else "")
            self.proc = subprocess.Popen(
                cmd, stdout=subprocess.PIPE, stderr=subprocess.STDOUT,
                text=True, encoding="utf-8", errors="replace", bufsize=1,
                cwd=str(AQUI), env=env)
            for linha in self.proc.stdout:  # type: ignore[union-attr]
                self.fila.put(linha.rstrip())
            cod = self.proc.wait()
        except Exception as exc:  # noqa: BLE001
            self.fila.put(f"ERRO: {exc}")
        finally:
            self.proc = None
        self.fila.put(f"__FIM__{cod}")

    def _para(self):
        if self.proc is not None:
            self.proc.terminate()
            self._log("Interrompido.")

    def _drena(self):
        while True:
            try:
                item = self.fila.get_nowait()
            except queue.Empty:
                break
            if item.startswith("__FIM__"):
                cod = item.removeprefix("__FIM__")
                self.barra.stop()
                self.btn_corre.config(state="normal")
                self.btn_analisa.config(state="normal")
                self.btn_para.config(state="disabled")
                ok = cod == "0"
                self.v_estado.set("Concluído" if ok else f"Erro ({cod})")
                if ok:
                    self.btn_abre.config(state="normal")
            else:
                self._log(item)
        self.after(100, self._drena)

    def _log(self, texto: str):
        self.txt.config(state="normal")
        self.txt.insert("end", texto + "\n")
        self.txt.see("end")
        self.txt.config(state="disabled")


if __name__ == "__main__":
    App().mainloop()
